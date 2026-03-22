"""
v2_chart_data_excel.py — Generate a "Chart Data" Excel workbook from dashboard JSON.

Produces structured tabular data suitable for PPT chart creation:
  1. MOIC Distribution     MOIC heatmap grid data (SOC & EQ bases)
  2. IRR Analysis           Per-claim E[XIRR], P(IRR>30%), conditional metrics
  3. Return Sensitivity     E[MOIC] & P(IRR>30%) vs Upfront% at Tata Tail 10/20/30%
  4. Cashflow Timeline      Annual & quarterly cashflow projections
  5. J-Curve Data           Monthly cumulative NAV curves
  6. Waterfall              Nominal & PV waterfall steps
  7. Risk Metrics           P(loss), VaR, CVaR across grid
  8. Per-Claim Summary      Per-claim breakdown for reference scenario
  9. Stochastic Grid        Full stochastic pricing grid (MOIC, IRR, P(loss))
  10. Simulation Summary    Meta info, probability inputs, quantum summary
  11. Scenario Comparison   Comparison across scenarios
  12. Breakeven Analysis    Breakeven surfaces and per-claim breakeven
  13. IRR Histogram         IRR & MOIC distribution bins (Full + SIAC)
  14. Portfolio Comparison   Full vs SIAC side-by-side metrics
  15. Loss Sensitivity      P(Loss) / VaR by Upfront% (Full + SIAC)
  16. Scenario Bands        MOIC band probabilities (5x+, 3-5x, 2-3x, 1-2x, <1x)
  17. Timeline Summary      Per-claim duration stats + lifecycle phases
  18. IRR Probability Table  P(IRR>40/30/25/0%) for both portfolios
"""

from __future__ import annotations
import json
import os
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Styling constants ──
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
PCT_FMT = '0.0%'
NUM_FMT = '#,##0.00'
INT_FMT = '#,##0'
CR_FMT = '#,##0.00'


def _style_header_row(ws, row, max_col, start_col=1):
    """Apply header styling to a row."""
    for col in range(start_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_cell(ws, row, col, fmt=None):
    """Apply data cell styling."""
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    if fmt:
        cell.number_format = fmt
    return cell


def _auto_width(ws, min_width=10, max_width=25):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ===================================================================
# Sheet builders
# ===================================================================

def _write_moic_distribution(wb, data):
    """Sheet 1: MOIC Distribution Grid (SOC & EQ bases)."""
    for basis_key, label in [("investment_grid_soc", "MOIC Grid (SOC)"),
                              ("investment_grid_eq", "MOIC Grid (EQ)")]:
        grid = data.get(basis_key, [])
        if not grid:
            continue

        ws = wb.create_sheet(title=label[:31])

        # Headers
        headers = [
            "Upfront %", "Tata Tail %", "Award Share %",
            "E[MOIC]", "Median MOIC", "Std MOIC",
            "E[XIRR]", "Median XIRR",
            "E[Net Return] (Cr)", "P(Loss)",
            "P(IRR>30%)", "P(IRR>25%)",
            "VaR 1%", "CVaR 1%",
        ]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        _style_header_row(ws, 1, len(headers))

        for ri, cell in enumerate(grid, 2):
            ws.cell(row=ri, column=1, value=cell.get("upfront_pct", 0))
            _style_data_cell(ws, ri, 1, PCT_FMT)
            ws.cell(row=ri, column=2, value=cell.get("tata_tail_pct", 0))
            _style_data_cell(ws, ri, 2, PCT_FMT)
            ws.cell(row=ri, column=3, value=cell.get("award_share_pct", 0))
            _style_data_cell(ws, ri, 3, PCT_FMT)
            ws.cell(row=ri, column=4, value=cell.get("mean_moic", 0))
            _style_data_cell(ws, ri, 4, NUM_FMT)
            ws.cell(row=ri, column=5, value=cell.get("median_moic", 0))
            _style_data_cell(ws, ri, 5, NUM_FMT)
            ws.cell(row=ri, column=6, value=cell.get("std_moic", 0))
            _style_data_cell(ws, ri, 6, NUM_FMT)
            ws.cell(row=ri, column=7, value=cell.get("mean_xirr", 0))
            _style_data_cell(ws, ri, 7, PCT_FMT)
            ws.cell(row=ri, column=8, value=cell.get("median_xirr", 0))
            _style_data_cell(ws, ri, 8, PCT_FMT)
            ws.cell(row=ri, column=9, value=cell.get("mean_net_return_cr", 0))
            _style_data_cell(ws, ri, 9, CR_FMT)
            ws.cell(row=ri, column=10, value=cell.get("p_loss", 0))
            _style_data_cell(ws, ri, 10, PCT_FMT)
            ws.cell(row=ri, column=11, value=cell.get("p_irr_gt_30", 0))
            _style_data_cell(ws, ri, 11, PCT_FMT)
            ws.cell(row=ri, column=12, value=cell.get("p_irr_gt_25", 0))
            _style_data_cell(ws, ri, 12, PCT_FMT)
            ws.cell(row=ri, column=13, value=cell.get("var_1", 0))
            _style_data_cell(ws, ri, 13, CR_FMT)
            ws.cell(row=ri, column=14, value=cell.get("cvar_1", 0))
            _style_data_cell(ws, ri, 14, CR_FMT)

        _auto_width(ws)


def _write_irr_analysis(wb, data):
    """Sheet 2: IRR Analysis — Per-claim IRR metrics from per_claim_grid.

    per_claim_grid is dict[claim_id → list[grid_entries]].
    We pick the reference scenario (10% upfront / 20% tail) for each claim.
    """
    pcg = data.get("per_claim_grid", {})
    if not pcg or not isinstance(pcg, dict):
        return

    ws = wb.create_sheet(title="IRR Analysis")

    headers = [
        "Claim", "E[XIRR]", "Median XIRR", "Conditional E[XIRR|Win]",
        "P(XIRR>0)", "P(IRR>30%)", "E[MOIC]", "Median MOIC",
        "P(Loss)", "E[Net Return] (Cr)",
        "Economically Viable", "Mean Legal Cost (Cr)",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header_row(ws, 1, len(headers))

    ri = 2
    for claim_id, entries in pcg.items():
        if not isinstance(entries, list) or not entries:
            continue
        # Find reference scenario: 10% upfront, 20% tail
        ref = None
        for e in entries:
            if (abs(e.get("upfront_pct", 0) - 0.10) < 0.01 and
                    abs(e.get("tata_tail_pct", 0) - 0.20) < 0.01):
                ref = e
                break
        if ref is None:
            ref = entries[0]  # fallback to first entry

        ws.cell(row=ri, column=1, value=claim_id)
        ws.cell(row=ri, column=2, value=ref.get("mean_xirr", 0))
        _style_data_cell(ws, ri, 2, PCT_FMT)
        ws.cell(row=ri, column=3, value=ref.get("median_xirr", 0))
        _style_data_cell(ws, ri, 3, PCT_FMT)
        ws.cell(row=ri, column=4, value=ref.get("conditional_xirr_win", 0))
        _style_data_cell(ws, ri, 4, PCT_FMT)
        ws.cell(row=ri, column=5, value=ref.get("p_xirr_gt_0", 0))
        _style_data_cell(ws, ri, 5, PCT_FMT)
        ws.cell(row=ri, column=6, value=ref.get("p_irr_gt_30", 0))
        _style_data_cell(ws, ri, 6, PCT_FMT)
        ws.cell(row=ri, column=7, value=ref.get("mean_moic", 0))
        _style_data_cell(ws, ri, 7, NUM_FMT)
        ws.cell(row=ri, column=8, value=ref.get("median_moic", 0))
        _style_data_cell(ws, ri, 8, NUM_FMT)
        ws.cell(row=ri, column=9, value=ref.get("p_loss", 0))
        _style_data_cell(ws, ri, 9, PCT_FMT)
        ws.cell(row=ri, column=10, value=ref.get("mean_net_return_cr", 0))
        _style_data_cell(ws, ri, 10, CR_FMT)
        ws.cell(row=ri, column=11, value="Yes" if ref.get("economically_viable") else "No")
        _style_data_cell(ws, ri, 11)
        ws.cell(row=ri, column=12, value=ref.get("mean_legal_cost_cr", 0))
        _style_data_cell(ws, ri, 12, CR_FMT)
        ri += 1

    _auto_width(ws)


def _write_return_sensitivity(wb, data):
    """Sheet 3: Return Sensitivity — E[MOIC] and P(IRR>30%) vs Upfront%.

    X-axis: Upfront % (5%–30%)
    Line series: Tata Tail % (10%, 20%, 30%)
    Matches PPT slide: "Return Sensitivity — MOIC & IRR"
    """
    grid = data.get("investment_grid_soc", [])
    if not grid:
        return

    ws = wb.create_sheet(title="Return Sensitivity")

    # Build lookup: (upfront_pct, tata_tail_pct) → entry
    lookup = {}
    for c in grid:
        key = (c.get("upfront_pct", 0), c.get("tata_tail_pct", 0))
        lookup[key] = c

    upfronts = sorted(set(c.get("upfront_pct", 0) for c in grid))
    tail_series = [0.10, 0.15, 0.20, 0.25, 0.30]  # the tails to chart

    # ── Section 1: E[MOIC] by Upfront % ──
    ws.cell(row=1, column=1, value="Expected Multiple (E[MOIC]) by Upfront %")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

    ws.cell(row=2, column=1, value="Upfront %")
    for ci, t in enumerate(tail_series, 2):
        ws.cell(row=2, column=ci, value=f"Tail {t:.0%}")
    _style_header_row(ws, 2, 1 + len(tail_series))

    for ri, up in enumerate(upfronts, 3):
        ws.cell(row=ri, column=1, value=up)
        _style_data_cell(ws, ri, 1, PCT_FMT)
        for ci, t in enumerate(tail_series, 2):
            val = lookup.get((up, t), {}).get("mean_moic", 0)
            ws.cell(row=ri, column=ci, value=val)
            _style_data_cell(ws, ri, ci, NUM_FMT)

    # ── Section 2: P(IRR > 30%) by Upfront % ──
    gap_row = len(upfronts) + 5
    ws.cell(row=gap_row, column=1, value="Chance of 30%+ IRR by Upfront %")
    ws.cell(row=gap_row, column=1).font = Font(bold=True, size=12, color="1F4E79")

    ws.cell(row=gap_row + 1, column=1, value="Upfront %")
    for ci, t in enumerate(tail_series, 2):
        ws.cell(row=gap_row + 1, column=ci, value=f"Tail {t:.0%}")
    _style_header_row(ws, gap_row + 1, 1 + len(tail_series))

    for ri_off, up in enumerate(upfronts):
        ri = gap_row + 2 + ri_off
        ws.cell(row=ri, column=1, value=up)
        _style_data_cell(ws, ri, 1, PCT_FMT)
        for ci, t in enumerate(tail_series, 2):
            val = lookup.get((up, t), {}).get("p_irr_gt_30", 0)
            ws.cell(row=ri, column=ci, value=val)
            _style_data_cell(ws, ri, ci, PCT_FMT)

    # ── Section 3: IRR > 30% Sensitivity Heatmap (Upfront × Tata Tail) ──
    all_tails = sorted(set(c.get("tata_tail_pct", 0) for c in grid))
    gap2 = gap_row + len(upfronts) + 4
    ws.cell(row=gap2, column=1, value="IRR > 30% Sensitivity — Full Grid")
    ws.cell(row=gap2, column=1).font = Font(bold=True, size=12, color="1F4E79")

    ws.cell(row=gap2 + 1, column=1, value="Upfront ↓ / Tail →")
    for ci, t in enumerate(all_tails, 2):
        ws.cell(row=gap2 + 1, column=ci, value=f"{t:.0%}")
    _style_header_row(ws, gap2 + 1, 1 + len(all_tails))

    for ri_off, up in enumerate(upfronts):
        ri = gap2 + 2 + ri_off
        ws.cell(row=ri, column=1, value=f"{up:.0%}")
        ws.cell(row=ri, column=1).font = Font(bold=True)
        for ci, t in enumerate(all_tails, 2):
            val = lookup.get((up, t), {}).get("p_irr_gt_30", 0)
            ws.cell(row=ri, column=ci, value=val)
            _style_data_cell(ws, ri, ci, PCT_FMT)

    _auto_width(ws)


def _write_cashflow_timeline(wb, data):
    """Sheet 4 & 5: Annual and Quarterly Cashflow Projections."""
    cf = data.get("cashflow_analysis", {})

    # ── Annual timeline ──
    annual = cf.get("annual_timeline", [])
    if annual:
        ws = wb.create_sheet(title="Annual Cashflows")
        headers = ["Year", "Month Range", "% Resolving", "% Cumulative",
                   "E[Recovery] (Cr)", "Cumulative Recovery (Cr)", "Phase"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        _style_header_row(ws, 1, len(headers))

        for ri, row in enumerate(annual, 2):
            ws.cell(row=ri, column=1, value=row.get("year", ""))
            _style_data_cell(ws, ri, 1, INT_FMT)
            ws.cell(row=ri, column=2, value=row.get("month_range", ""))
            ws.cell(row=ri, column=3, value=row.get("pct_resolving", 0))
            _style_data_cell(ws, ri, 3, PCT_FMT)
            ws.cell(row=ri, column=4, value=row.get("pct_cumulative", 0))
            _style_data_cell(ws, ri, 4, PCT_FMT)
            ws.cell(row=ri, column=5, value=row.get("e_recovery_cr", 0))
            _style_data_cell(ws, ri, 5, CR_FMT)
            ws.cell(row=ri, column=6, value=row.get("cumul_recovery_cr", 0))
            _style_data_cell(ws, ri, 6, CR_FMT)
            ws.cell(row=ri, column=7, value=row.get("phase", ""))
        _auto_width(ws)

    # ── Quarterly timeline ──
    quarterly = cf.get("quarterly_timeline", [])
    if quarterly:
        ws = wb.create_sheet(title="Quarterly Cashflows")
        headers = ["Quarter", "Label", "% Resolving",
                   "E[Recovery] (Cr)", "Cumulative Recovery (Cr)"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        _style_header_row(ws, 1, len(headers))

        for ri, row in enumerate(quarterly, 2):
            ws.cell(row=ri, column=1, value=row.get("quarter", ""))
            ws.cell(row=ri, column=2, value=row.get("label", ""))
            ws.cell(row=ri, column=3, value=row.get("pct_resolving", 0))
            _style_data_cell(ws, ri, 3, PCT_FMT)
            ws.cell(row=ri, column=4, value=row.get("e_recovery_cr", 0))
            _style_data_cell(ws, ri, 4, CR_FMT)
            ws.cell(row=ri, column=5, value=row.get("cumul_recovery_cr", 0))
            _style_data_cell(ws, ri, 5, CR_FMT)
        _auto_width(ws)

    # ── Cashflow decomposition (waterfall steps) ──
    decomp = cf.get("decomposition", [])
    if decomp and isinstance(decomp, list):
        ws = wb.create_sheet(title="Cashflow Decomposition")

        headers = ["Step", "Label", "Factor", "Value (Cr)", "Note"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        _style_header_row(ws, 1, len(headers))

        for ri, row in enumerate(decomp, 2):
            ws.cell(row=ri, column=1, value=row.get("step", ""))
            ws.cell(row=ri, column=2, value=row.get("label", ""))
            ws.cell(row=ri, column=3, value=row.get("factor", ""))
            ws.cell(row=ri, column=4, value=row.get("value_cr", 0))
            _style_data_cell(ws, ri, 4, CR_FMT)
            ws.cell(row=ri, column=5, value=row.get("note", ""))
        _auto_width(ws)

    # ── Investor scenarios ──
    inv_scenarios = cf.get("investor_scenarios", [])
    if inv_scenarios:
        ws = wb.create_sheet(title="Investor Scenarios")
        if inv_scenarios:
            headers = list(inv_scenarios[0].keys())
            for ci, h in enumerate(headers, 1):
                ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
            _style_header_row(ws, 1, len(headers))

            for ri, sc in enumerate(inv_scenarios, 2):
                for ci, h in enumerate(headers, 1):
                    val = sc.get(h, "")
                    ws.cell(row=ri, column=ci, value=val)
                    if isinstance(val, float):
                        if abs(val) < 1:
                            _style_data_cell(ws, ri, ci, PCT_FMT)
                        else:
                            _style_data_cell(ws, ri, ci, NUM_FMT)
            _auto_width(ws)


def _write_jcurve(wb, data):
    """Sheet 6: J-Curve Data — Monthly cumulative NAV for various scenarios."""
    jc = data.get("jcurve_data", {})
    scenarios = jc.get("scenarios", {})
    if not scenarios:
        return

    ws = wb.create_sheet(title="J-Curve Data")

    # Write each scenario as a block
    # Each scenario value is a list of {month, label, p5, p25, median, p75, p95, mean}
    col_offset = 0
    for key, points in scenarios.items():
        if not isinstance(points, list) or not points:
            continue

        start_col = col_offset + 1
        ws.cell(row=1, column=start_col, value=f"Scenario: {key}")
        ws.cell(row=1, column=start_col).font = Font(bold=True, size=10, color="1F4E79")

        headers = ["Month", "Label", "Mean NAV (Cr)", "P5", "P25", "Median", "P75", "P95"]
        for ci, h in enumerate(headers):
            ws.cell(row=2, column=start_col + ci, value=h)
        _style_header_row(ws, 2, start_col + len(headers) - 1)

        for ri, pt in enumerate(points, 3):
            ws.cell(row=ri, column=start_col, value=pt.get("month", 0))
            _style_data_cell(ws, ri, start_col, INT_FMT)
            ws.cell(row=ri, column=start_col + 1, value=pt.get("label", ""))

            for ci, metric in enumerate(["mean", "p5", "p25", "median", "p75", "p95"], 2):
                val = pt.get(metric, 0)
                ws.cell(row=ri, column=start_col + ci, value=val)
                _style_data_cell(ws, ri, start_col + ci, CR_FMT)

        col_offset += len(headers) + 1  # gap between scenarios

    _auto_width(ws)


def _write_waterfall(wb, data):
    """Sheet 7: Waterfall — Step-by-step value waterfall."""
    wf = data.get("waterfall", {})
    if not wf:
        return

    ws = wb.create_sheet(title="Waterfall")

    # Waterfall is a dict of step_name → value
    headers = ["Step", "Value"]
    ws.cell(row=1, column=1, value="Step")
    ws.cell(row=1, column=2, value="Value (Cr)")
    ws.cell(row=1, column=3, value="Description")
    _style_header_row(ws, 1, 3)

    step_labels = {
        "soc_cr": "SOC Value (Cr)",
        "pv_factor": "PV Discount Factor",
        "pv_soc_cr": "PV-Adjusted SOC (Cr)",
        "win_rate": "Probability-Weighted Win Rate",
        "eq_multiplier": "Expected Quantum Multiplier",
        "prob_adjusted_cr": "Probability-Adjusted Value (Cr)",
        "legal_costs_cr": "Legal Costs (Cr)",
        "net_after_legal_cr": "Net After Legal (Cr)",
        "reference_tail_pct": "Reference Tata Tail %",
        "tata_receives_cr": "Tata Receives (Cr)",
        "fund_net_profit_cr": "Fund Net Profit (Cr)",
    }

    ri = 2
    for key, label in step_labels.items():
        if key in wf and not isinstance(wf[key], (dict, list)):
            ws.cell(row=ri, column=1, value=key)
            ws.cell(row=ri, column=2, value=wf[key])
            if isinstance(wf[key], (int, float)):
                _style_data_cell(ws, ri, 2, CR_FMT if abs(wf[key]) > 1 else NUM_FMT)
            ws.cell(row=ri, column=3, value=label)
            ri += 1

    # Write nominal/present_value sub-tables if present
    for section_key, section_label in [("nominal", "Nominal Valuation"), ("present_value", "Present Value Adjusted")]:
        section = wf.get(section_key)
        if isinstance(section, dict):
            ri += 1
            ws.cell(row=ri, column=1, value=section_label)
            ws.cell(row=ri, column=1).font = Font(bold=True, size=10, color="1F4E79")
            ri += 1
            for sk, sv in section.items():
                if isinstance(sv, (int, float, str)):
                    ws.cell(row=ri, column=1, value=sk)
                    ws.cell(row=ri, column=2, value=sv)
                    if isinstance(sv, float):
                        _style_data_cell(ws, ri, 2, CR_FMT if abs(sv) > 1 else NUM_FMT)
                    ws.cell(row=ri, column=3, value=step_labels.get(sk, sk.replace("_", " ").title()))
                    ri += 1

    _auto_width(ws)


def _write_risk_metrics(wb, data):
    """Sheet 8: Risk Metrics — P(loss), VaR, CVaR across the grid."""
    ws = wb.create_sheet(title="Risk Metrics")

    grid = data.get("investment_grid_soc", [])
    if not grid:
        return

    headers = [
        "Upfront %", "Tata Tail %",
        "P(Loss)", "VaR 1% (Cr)", "CVaR 1% (Cr)",
        "E[MOIC]", "Std MOIC",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header_row(ws, 1, len(headers))

    for ri, c in enumerate(grid, 2):
        ws.cell(row=ri, column=1, value=c.get("upfront_pct", 0))
        _style_data_cell(ws, ri, 1, PCT_FMT)
        ws.cell(row=ri, column=2, value=c.get("tata_tail_pct", 0))
        _style_data_cell(ws, ri, 2, PCT_FMT)
        ws.cell(row=ri, column=3, value=c.get("p_loss", 0))
        _style_data_cell(ws, ri, 3, PCT_FMT)
        ws.cell(row=ri, column=4, value=c.get("var_1", 0))
        _style_data_cell(ws, ri, 4, CR_FMT)
        ws.cell(row=ri, column=5, value=c.get("cvar_1", 0))
        _style_data_cell(ws, ri, 5, CR_FMT)
        ws.cell(row=ri, column=6, value=c.get("mean_moic", 0))
        _style_data_cell(ws, ri, 6, NUM_FMT)
        ws.cell(row=ri, column=7, value=c.get("std_moic", 0))
        _style_data_cell(ws, ri, 7, NUM_FMT)

    _auto_width(ws)


def _write_per_claim_summary(wb, data):
    """Sheet 9: Per-Claim Summary from per_claim_grid.
    
    per_claim_grid is a dict: claim_id → list of grid entries.
    Each entry has upfront_pct, tata_tail_pct, award_share_pct, basis,
    mean_moic, median_moic, mean_xirr, etc.
    """
    pcg = data.get("per_claim_grid", {})
    if not pcg:
        return

    ws = wb.create_sheet(title="Per-Claim Detail")

    if isinstance(pcg, dict) and pcg:
        # Determine columns from first claim's first entry
        first_claim = next(iter(pcg.values()))
        if not first_claim or not isinstance(first_claim, list):
            return
        sample = first_claim[0] if first_claim else {}
        entry_keys = [k for k in sample.keys()]

        headers = ["Claim ID"] + [k.replace("_", " ").title() for k in entry_keys]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        _style_header_row(ws, 1, len(headers))

        ri = 2
        for claim_id, entries in pcg.items():
            if not isinstance(entries, list):
                continue
            for entry in entries:
                ws.cell(row=ri, column=1, value=claim_id)
                for ci, k in enumerate(entry_keys, 2):
                    val = entry.get(k, "")
                    ws.cell(row=ri, column=ci, value=val)
                    if isinstance(val, float):
                        if abs(val) < 1:
                            _style_data_cell(ws, ri, ci, PCT_FMT)
                        else:
                            _style_data_cell(ws, ri, ci, NUM_FMT)
                ri += 1

    _auto_width(ws)


def _write_stochastic_grid(wb, data):
    """Sheet 10: Full stochastic pricing grid data."""
    sp = data.get("stochastic_pricing", {})
    if not sp:
        # Try loading from stochastic_pricing.json in same directory
        return

    grid = sp.get("grid", {})
    if not grid:
        return

    ws = wb.create_sheet(title="Stochastic Grid")

    # Grid is dict: "upfront_tail" → metrics
    headers = [
        "Upfront %", "Tata Tail %",
        "E[MOIC]", "Median MOIC",
        "E[IRR]", "Median IRR",
        "P5 MOIC", "P25 MOIC", "P75 MOIC", "P95 MOIC",
        "P5 IRR", "P25 IRR", "P75 IRR", "P95 IRR",
        "P(Loss)", "P(Hurdle)",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header_row(ws, 1, len(headers))

    ri = 2
    for key, cell in sorted(grid.items()):
        parts = key.split("_")
        if len(parts) >= 2:
            try:
                up = int(parts[0]) / 100
                tail = int(parts[1]) / 100
            except ValueError:
                continue
        else:
            continue

        ws.cell(row=ri, column=1, value=up)
        _style_data_cell(ws, ri, 1, PCT_FMT)
        ws.cell(row=ri, column=2, value=tail)
        _style_data_cell(ws, ri, 2, PCT_FMT)
        ws.cell(row=ri, column=3, value=cell.get("e_moic", cell.get("mean_moic", 0)))
        _style_data_cell(ws, ri, 3, NUM_FMT)
        ws.cell(row=ri, column=4, value=cell.get("p50_moic", cell.get("median_moic", 0)))
        _style_data_cell(ws, ri, 4, NUM_FMT)
        ws.cell(row=ri, column=5, value=cell.get("e_irr", cell.get("mean_xirr", 0)))
        _style_data_cell(ws, ri, 5, PCT_FMT)
        ws.cell(row=ri, column=6, value=cell.get("p50_irr", cell.get("median_xirr", 0)))
        _style_data_cell(ws, ri, 6, PCT_FMT)
        ws.cell(row=ri, column=7, value=cell.get("p5_moic", 0))
        _style_data_cell(ws, ri, 7, NUM_FMT)
        ws.cell(row=ri, column=8, value=cell.get("p25_moic", 0))
        _style_data_cell(ws, ri, 8, NUM_FMT)
        ws.cell(row=ri, column=9, value=cell.get("p75_moic", 0))
        _style_data_cell(ws, ri, 9, NUM_FMT)
        ws.cell(row=ri, column=10, value=cell.get("p95_moic", 0))
        _style_data_cell(ws, ri, 10, NUM_FMT)
        ws.cell(row=ri, column=11, value=cell.get("p5_irr", 0))
        _style_data_cell(ws, ri, 11, PCT_FMT)
        ws.cell(row=ri, column=12, value=cell.get("p25_irr", 0))
        _style_data_cell(ws, ri, 12, PCT_FMT)
        ws.cell(row=ri, column=13, value=cell.get("p75_irr", 0))
        _style_data_cell(ws, ri, 13, PCT_FMT)
        ws.cell(row=ri, column=14, value=cell.get("p95_irr", 0))
        _style_data_cell(ws, ri, 14, PCT_FMT)
        ws.cell(row=ri, column=15, value=cell.get("prob_loss", cell.get("p_loss", 0)))
        _style_data_cell(ws, ri, 15, PCT_FMT)
        ws.cell(row=ri, column=16, value=cell.get("prob_hurdle", 0))
        _style_data_cell(ws, ri, 16, PCT_FMT)
        ri += 1

    _auto_width(ws)


def _write_simulation_summary(wb, data):
    """Sheet 11: Simulation Summary — meta, quantum, probability, claims."""
    ws = wb.create_sheet(title="Simulation Summary")

    # ── Meta ──
    meta = data.get("simulation_meta", {})
    ws.cell(row=1, column=1, value="Simulation Meta")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

    meta_fields = [
        ("Number of Paths", meta.get("n_paths", "")),
        ("Random Seed", meta.get("seed", "")),
        ("Portfolio", meta.get("portfolio_label", "")),
        ("Number of Claims", meta.get("n_claims", "")),
        ("Timestamp", meta.get("timestamp", "")),
    ]
    for ri, (label, val) in enumerate(meta_fields, 2):
        ws.cell(row=ri, column=1, value=label)
        ws.cell(row=ri, column=1).font = SUBHEADER_FONT
        ws.cell(row=ri, column=2, value=val)

    # ── Probability Summary ──
    row = len(meta_fields) + 3
    ws.cell(row=row, column=1, value="Probability Summary")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")

    prob = data.get("probability_summary", {})
    prob_items = [
        ("Arbitration Win Probability", prob.get("arb_win_probability", prob.get("arb_win_prob", ""))),
        ("Re-Arb Win Probability", prob.get("re_arb_win_probability", prob.get("re_arb_win_prob", ""))),
    ]
    for ri_off, (label, val) in enumerate(prob_items):
        ws.cell(row=row + 1 + ri_off, column=1, value=label)
        ws.cell(row=row + 1 + ri_off, column=1).font = SUBHEADER_FONT
        ws.cell(row=row + 1 + ri_off, column=2, value=val)

    # ── Quantum Summary ──
    row = row + len(prob_items) + 2
    ws.cell(row=row, column=1, value="Quantum Summary")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")

    qs = data.get("quantum_summary", {})
    q_items = [
        ("Expected Quantum % of SOC", qs.get("expected_quantum_pct_of_soc", "")),
        ("Quantum Bands", str(qs.get("bands", []))),
    ]
    for ri_off, (label, val) in enumerate(q_items):
        ws.cell(row=row + 1 + ri_off, column=1, value=label)
        ws.cell(row=row + 1 + ri_off, column=1).font = SUBHEADER_FONT
        ws.cell(row=row + 1 + ri_off, column=2, value=val)

    # ── Claims overview ──
    row = row + len(q_items) + 2
    ws.cell(row=row, column=1, value="Claims")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")

    claims = data.get("claims", [])
    if claims:
        claim_headers = ["Claim ID", "Description", "SOC (Cr)", "Jurisdiction"]
        for ci, h in enumerate(claim_headers, 1):
            ws.cell(row=row + 1, column=ci, value=h)
        _style_header_row(ws, row + 1, len(claim_headers))

        for ri_off, cl in enumerate(claims):
            r = row + 2 + ri_off
            ws.cell(row=r, column=1, value=cl.get("id", ""))
            ws.cell(row=r, column=2, value=cl.get("description", ""))
            ws.cell(row=r, column=3, value=cl.get("soc_value_cr", 0))
            _style_data_cell(ws, r, 3, CR_FMT)
            ws.cell(row=r, column=4, value=cl.get("jurisdiction", ""))

    _auto_width(ws)


def _write_scenario_comparison(wb, data):
    """Sheet: Scenario Comparison — verdict matrix."""
    sc = data.get("scenario_comparison", [])
    if not sc:
        return

    ws = wb.create_sheet(title="Scenario Comparison")

    if isinstance(sc, list) and sc:
        headers = list(sc[0].keys())
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
        _style_header_row(ws, 1, len(headers))

        for ri, row in enumerate(sc, 2):
            for ci, h in enumerate(headers, 1):
                val = row.get(h, "")
                ws.cell(row=ri, column=ci, value=val)
                if isinstance(val, float):
                    if abs(val) < 1:
                        _style_data_cell(ws, ri, ci, PCT_FMT)
                    else:
                        _style_data_cell(ws, ri, ci, NUM_FMT)

    _auto_width(ws)


# ===================================================================
# New PPT-oriented sheets
# ===================================================================

def _find_stochastic_ref(grid, target_up=10, target_tail=20):
    """Find reference scenario in stochastic grid.

    Handles both decimal (0.10) and integer (10) percentage formats.
    Returns (key, entry) or (first_key, first_entry) as fallback.
    """
    for k, v in grid.items():
        up = v.get("upfront_pct", 0)
        tail = v.get("tata_tail_pct", 0)
        # Integer format (5, 10, 20)
        if abs(up - target_up) < 0.5 and abs(tail - target_tail) < 0.5:
            return k, v
        # Decimal format (0.05, 0.10, 0.20)
        if abs(up - target_up / 100) < 0.005 and abs(tail - target_tail / 100) < 0.005:
            return k, v
    first_key = next(iter(grid))
    return first_key, grid[first_key]


def _write_irr_histogram(wb, data):
    """IRR & MOIC Distribution Histograms from stochastic pricing grid.

    Writes bin-level data for the reference scenario (10% upfront / 20% tail)
    for both Full and SIAC portfolios (if available).
    """
    stoch = data.get("stochastic_pricing", {})
    grid = stoch.get("grid", {})
    if not grid:
        return

    # Find reference scenario key (handles both integer and decimal pct formats)
    ref_key, ref = _find_stochastic_ref(grid)

    # ── IRR Histogram ──
    irr_hist = ref.get("irr_hist", [])
    if irr_hist:
        ws = wb.create_sheet(title="IRR Histogram")
        ws.cell(row=1, column=1, value=f"IRR Distribution — Full Portfolio ({ref_key})")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

        headers = ["Bin Edge (IRR)", "Count", "Bin Edge (%)", "Frequency %"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=2, column=ci, value=h)
        _style_header_row(ws, 2, len(headers))

        total = sum(b.get("count", 0) for b in irr_hist)
        for ri, b in enumerate(irr_hist, 3):
            edge = b.get("edge", 0)
            count = b.get("count", 0)
            ws.cell(row=ri, column=1, value=edge)
            _style_data_cell(ws, ri, 1, NUM_FMT)
            ws.cell(row=ri, column=2, value=count)
            _style_data_cell(ws, ri, 2, INT_FMT)
            ws.cell(row=ri, column=3, value=edge)
            _style_data_cell(ws, ri, 3, PCT_FMT)
            freq = count / total if total > 0 else 0
            ws.cell(row=ri, column=4, value=freq)
            _style_data_cell(ws, ri, 4, PCT_FMT)

        # SIAC histogram if available
        siac_stoch = data.get("siac_stochastic_pricing", {})
        siac_grid = siac_stoch.get("grid", {})
        if siac_grid:
            _, siac_ref = _find_stochastic_ref(siac_grid)
            siac_irr_hist = siac_ref.get("irr_hist", [])
            if siac_irr_hist:
                col_off = 6
                ws.cell(row=1, column=col_off, value="IRR Distribution — SIAC Portfolio")
                ws.cell(row=1, column=col_off).font = Font(bold=True, size=12, color="1F4E79")
                for ci, h in enumerate(["Bin Edge (IRR)", "Count", "Frequency %"], col_off):
                    ws.cell(row=2, column=ci, value=h)
                _style_header_row(ws, 2, col_off + 2, start_col=col_off)

                stotal = sum(b.get("count", 0) for b in siac_irr_hist)
                for ri, b in enumerate(siac_irr_hist, 3):
                    ws.cell(row=ri, column=col_off, value=b.get("edge", 0))
                    _style_data_cell(ws, ri, col_off, PCT_FMT)
                    ws.cell(row=ri, column=col_off + 1, value=b.get("count", 0))
                    _style_data_cell(ws, ri, col_off + 1, INT_FMT)
                    freq = b.get("count", 0) / stotal if stotal > 0 else 0
                    ws.cell(row=ri, column=col_off + 2, value=freq)
                    _style_data_cell(ws, ri, col_off + 2, PCT_FMT)

        _auto_width(ws)

    # ── MOIC Histogram ──
    moic_hist = ref.get("moic_hist", [])
    if moic_hist:
        ws = wb.create_sheet(title="MOIC Histogram")
        ws.cell(row=1, column=1, value=f"MOIC Distribution — Full Portfolio ({ref_key})")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

        headers = ["Bin Edge (MOIC)", "Count", "Frequency %"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=2, column=ci, value=h)
        _style_header_row(ws, 2, len(headers))

        total = sum(b.get("count", 0) for b in moic_hist)
        for ri, b in enumerate(moic_hist, 3):
            edge = b.get("edge", 0)
            count = b.get("count", 0)
            ws.cell(row=ri, column=1, value=edge)
            _style_data_cell(ws, ri, 1, NUM_FMT)
            ws.cell(row=ri, column=2, value=count)
            _style_data_cell(ws, ri, 2, INT_FMT)
            freq = count / total if total > 0 else 0
            ws.cell(row=ri, column=3, value=freq)
            _style_data_cell(ws, ri, 3, PCT_FMT)

        # SIAC MOIC histogram
        siac_stoch = data.get("siac_stochastic_pricing", {})
        siac_grid = siac_stoch.get("grid", {})
        if siac_grid:
            _, siac_ref = _find_stochastic_ref(siac_grid)
            siac_moic_hist = siac_ref.get("moic_hist", [])
            if siac_moic_hist:
                col_off = 5
                ws.cell(row=1, column=col_off, value="MOIC Distribution — SIAC Portfolio")
                ws.cell(row=1, column=col_off).font = Font(bold=True, size=12, color="1F4E79")
                for ci, h in enumerate(["Bin Edge (MOIC)", "Count", "Frequency %"], col_off):
                    ws.cell(row=2, column=ci, value=h)
                _style_header_row(ws, 2, col_off + 2, start_col=col_off)

                stotal = sum(b.get("count", 0) for b in siac_moic_hist)
                for ri, b in enumerate(siac_moic_hist, 3):
                    ws.cell(row=ri, column=col_off, value=b.get("edge", 0))
                    _style_data_cell(ws, ri, col_off, NUM_FMT)
                    ws.cell(row=ri, column=col_off + 1, value=b.get("count", 0))
                    _style_data_cell(ws, ri, col_off + 1, INT_FMT)
                    freq = b.get("count", 0) / stotal if stotal > 0 else 0
                    ws.cell(row=ri, column=col_off + 2, value=freq)
                    _style_data_cell(ws, ri, col_off + 2, PCT_FMT)

        _auto_width(ws)


def _write_portfolio_comparison(wb, data):
    """Cross-portfolio comparison: Full vs SIAC side-by-side metrics.

    Used for PPT slide: "Full Portfolio vs SIAC — Comparison"
    """
    siac_data = data.get("siac_dashboard_data", {})
    if not siac_data:
        return

    ws = wb.create_sheet(title="Portfolio Comparison")

    ws.cell(row=1, column=1, value="Full Portfolio vs SIAC — Comparison")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E79")

    headers = ["Metric", "Full Portfolio", "SIAC Portfolio"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=2, column=ci, value=h)
    _style_header_row(ws, 2, len(headers))

    def _get_ref_grid_entry(d, target_up=0.10, target_tail=0.20):
        """Get reference scenario from investment_grid_soc."""
        grid = d.get("investment_grid_soc", [])
        for e in grid:
            if (abs(e.get("upfront_pct", 0) - target_up) < 0.01 and
                    abs(e.get("tata_tail_pct", 0) - target_tail) < 0.01):
                return e
        return grid[0] if grid else {}

    full_ref = _get_ref_grid_entry(data)
    siac_ref = _get_ref_grid_entry(siac_data)

    full_ps = data.get("cashflow_analysis", {}).get("portfolio_summary", {})
    siac_ps = siac_data.get("cashflow_analysis", {}).get("portfolio_summary", {})

    full_prob = data.get("probability_summary", {})
    siac_prob = siac_data.get("probability_summary", {})

    # Define metric rows: (label, full_val, siac_val, fmt)
    metrics = [
        ("Total SOC (Cr)", full_ps.get("total_soc_cr", 0), siac_ps.get("total_soc_cr", 0), CR_FMT),
        ("Total Equity (Cr)", full_ps.get("total_eq_cr", 0), siac_ps.get("total_eq_cr", 0), CR_FMT),
        ("Equity / SOC", full_ps.get("eq_over_soc", 0), siac_ps.get("eq_over_soc", 0), PCT_FMT),
        ("Arb Win Probability", full_prob.get("arb_win_probability", 0), siac_prob.get("arb_win_probability", 0), PCT_FMT),
        ("Avg Win Rate (Award/SOC)", full_ps.get("avg_win_rate", 0), siac_ps.get("avg_win_rate", 0), PCT_FMT),
        ("E[Total Collected] (Cr)", full_ps.get("total_e_collected_cr", 0), siac_ps.get("total_e_collected_cr", 0), CR_FMT),
        ("E[Total Legal Cost] (Cr)", full_ps.get("total_e_legal_cr", 0), siac_ps.get("total_e_legal_cr", 0), CR_FMT),
        ("E[Net Profit] (Cr)", full_ps.get("total_e_net_cr", 0), siac_ps.get("total_e_net_cr", 0), CR_FMT),
        ("E[MOIC] (10%/20%)", full_ref.get("mean_moic", 0), siac_ref.get("mean_moic", 0), NUM_FMT),
        ("E[XIRR] (10%/20%)", full_ref.get("mean_xirr", 0), siac_ref.get("mean_xirr", 0), PCT_FMT),
        ("P(IRR > 30%)", full_ref.get("p_irr_gt_30", 0), siac_ref.get("p_irr_gt_30", 0), PCT_FMT),
        ("P(IRR > 25%)", full_ref.get("p_irr_gt_25", 0), siac_ref.get("p_irr_gt_25", 0), PCT_FMT),
        ("P(Loss)", full_ref.get("p_loss", 0), siac_ref.get("p_loss", 0), PCT_FMT),
        ("VaR (1%)", full_ref.get("var_1", 0), siac_ref.get("var_1", 0), NUM_FMT),
        ("CVaR (1%)", full_ref.get("cvar_1", 0), siac_ref.get("cvar_1", 0), NUM_FMT),
        ("Simulation Paths", full_ps.get("n_paths", 0), siac_ps.get("n_paths", 0), INT_FMT),
    ]

    for ri, (label, full_val, siac_val, fmt) in enumerate(metrics, 3):
        ws.cell(row=ri, column=1, value=label)
        ws.cell(row=ri, column=1).font = Font(bold=True)
        ws.cell(row=ri, column=2, value=full_val)
        _style_data_cell(ws, ri, 2, fmt)
        ws.cell(row=ri, column=3, value=siac_val)
        _style_data_cell(ws, ri, 3, fmt)

    _auto_width(ws)


def _write_loss_sensitivity(wb, data):
    """Loss Sensitivity: P(Loss) and VaR by Upfront% for both portfolios.

    Used for PPT slide: "Risk Analysis — P(Loss) / VaR / CVaR"
    """
    grid = data.get("investment_grid_soc", [])
    if not grid:
        return

    ws = wb.create_sheet(title="Loss Sensitivity")
    ws.cell(row=1, column=1, value="Loss Sensitivity — P(Loss) by Upfront %")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

    # Build lookup
    lookup = {}
    for c in grid:
        key = (c.get("upfront_pct", 0), c.get("tata_tail_pct", 0))
        lookup[key] = c

    upfronts = sorted(set(c.get("upfront_pct", 0) for c in grid))
    tail_series = [0.10, 0.20, 0.30]

    # ── Section 1: P(Loss) ──
    ws.cell(row=2, column=1, value="Upfront %")
    col = 2
    for t in tail_series:
        ws.cell(row=2, column=col, value=f"Full Tail {t:.0%}")
        col += 1
    # SIAC columns
    siac_grid = data.get("siac_dashboard_data", {}).get("investment_grid_soc", [])
    siac_lookup = {}
    for c in siac_grid:
        key = (c.get("upfront_pct", 0), c.get("tata_tail_pct", 0))
        siac_lookup[key] = c
    for t in tail_series:
        ws.cell(row=2, column=col, value=f"SIAC Tail {t:.0%}")
        col += 1
    _style_header_row(ws, 2, col - 1)

    for ri_off, up in enumerate(upfronts):
        ri = 3 + ri_off
        ws.cell(row=ri, column=1, value=up)
        _style_data_cell(ws, ri, 1, PCT_FMT)
        col = 2
        for t in tail_series:
            val = lookup.get((up, t), {}).get("p_loss", 0)
            ws.cell(row=ri, column=col, value=val)
            _style_data_cell(ws, ri, col, PCT_FMT)
            col += 1
        for t in tail_series:
            val = siac_lookup.get((up, t), {}).get("p_loss", 0)
            ws.cell(row=ri, column=col, value=val)
            _style_data_cell(ws, ri, col, PCT_FMT)
            col += 1

    # ── Section 2: VaR (1%) ──
    gap = len(upfronts) + 5
    ws.cell(row=gap, column=1, value="VaR (1%) by Upfront %")
    ws.cell(row=gap, column=1).font = Font(bold=True, size=12, color="1F4E79")

    ws.cell(row=gap + 1, column=1, value="Upfront %")
    col = 2
    for t in tail_series:
        ws.cell(row=gap + 1, column=col, value=f"Full Tail {t:.0%}")
        col += 1
    for t in tail_series:
        ws.cell(row=gap + 1, column=col, value=f"SIAC Tail {t:.0%}")
        col += 1
    _style_header_row(ws, gap + 1, col - 1)

    for ri_off, up in enumerate(upfronts):
        ri = gap + 2 + ri_off
        ws.cell(row=ri, column=1, value=up)
        _style_data_cell(ws, ri, 1, PCT_FMT)
        col = 2
        for t in tail_series:
            val = lookup.get((up, t), {}).get("var_1", 0)
            ws.cell(row=ri, column=col, value=val)
            _style_data_cell(ws, ri, col, NUM_FMT)
            col += 1
        for t in tail_series:
            val = siac_lookup.get((up, t), {}).get("var_1", 0)
            ws.cell(row=ri, column=col, value=val)
            _style_data_cell(ws, ri, col, NUM_FMT)
            col += 1

    _auto_width(ws)


def _write_scenario_bands(wb, data):
    """Scenario Band Probabilities from MOIC histogram.

    Computes probability mass in MOIC bands:
      Best Case 5x+, Strong Upside 3-5x, Base Case 2-3x,
      Moderate 1-2x, Loss <1x
    For reference scenario (10%/20%) — both Full and SIAC.
    """
    stoch = data.get("stochastic_pricing", {})
    grid = stoch.get("grid", {})
    if not grid:
        return

    def _compute_bands(g):
        """Find ref scenario and compute MOIC bands."""
        _, ref = _find_stochastic_ref(g)

        moic_hist = ref.get("moic_hist", [])
        if not moic_hist:
            return None, ref

        total = sum(b.get("count", 0) for b in moic_hist)
        if total == 0:
            return None, ref

        bands = {"5x+ (Best Case)": 0, "3x-5x (Strong Upside)": 0,
                 "2x-3x (Base Case)": 0, "1x-2x (Moderate)": 0,
                 "<1x (Loss)": 0}

        for b in moic_hist:
            edge = b.get("edge", 0)
            count = b.get("count", 0)
            if edge >= 5.0:
                bands["5x+ (Best Case)"] += count
            elif edge >= 3.0:
                bands["3x-5x (Strong Upside)"] += count
            elif edge >= 2.0:
                bands["2x-3x (Base Case)"] += count
            elif edge >= 1.0:
                bands["1x-2x (Moderate)"] += count
            else:
                bands["<1x (Loss)"] += count

        return {k: v / total for k, v in bands.items()}, ref

    ws = wb.create_sheet(title="Scenario Bands")
    ws.cell(row=1, column=1, value="MOIC Scenario Band Probabilities (10% Upfront / 20% Tail)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

    headers = ["Scenario Band", "Full Portfolio", "SIAC Portfolio"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=2, column=ci, value=h)
    _style_header_row(ws, 2, len(headers))

    full_bands, _ = _compute_bands(grid)
    siac_stoch = data.get("siac_stochastic_pricing", {})
    siac_grid = siac_stoch.get("grid", {})
    siac_bands = None
    if siac_grid:
        siac_bands, _ = _compute_bands(siac_grid)

    band_order = ["5x+ (Best Case)", "3x-5x (Strong Upside)", "2x-3x (Base Case)",
                  "1x-2x (Moderate)", "<1x (Loss)"]

    for ri, band in enumerate(band_order, 3):
        ws.cell(row=ri, column=1, value=band)
        ws.cell(row=ri, column=1).font = Font(bold=True)
        if full_bands:
            ws.cell(row=ri, column=2, value=full_bands.get(band, 0))
            _style_data_cell(ws, ri, 2, PCT_FMT)
        if siac_bands:
            ws.cell(row=ri, column=3, value=siac_bands.get(band, 0))
            _style_data_cell(ws, ri, 3, PCT_FMT)

    _auto_width(ws)


def _write_timeline_summary(wb, data):
    """Timeline Summary: Per-claim duration stats + lifecycle phases.

    Used for PPT slide: "Cash Flow Timeline — Recovery Schedule"
    """
    cf = data.get("cashflow_analysis", {})
    ts = cf.get("timeline_summary", {})
    per_claim = ts.get("per_claim", {})

    if not per_claim:
        return

    ws = wb.create_sheet(title="Timeline Summary")
    ws.cell(row=1, column=1, value="Per-Claim Duration Analysis (Months)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

    headers = ["Claim ID", "Pipeline", "Jurisdiction", "Mean", "Median",
               "P5", "P25", "P75", "P95", "Max", "P(>96 mo)"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=2, column=ci, value=h)
    _style_header_row(ws, 2, len(headers))

    ri = 3
    for claim_id, stats in per_claim.items():
        ws.cell(row=ri, column=1, value=claim_id)
        ws.cell(row=ri, column=2, value=stats.get("pipeline", ""))
        ws.cell(row=ri, column=3, value=stats.get("jurisdiction", ""))
        ws.cell(row=ri, column=4, value=stats.get("mean", 0))
        _style_data_cell(ws, ri, 4, NUM_FMT)
        ws.cell(row=ri, column=5, value=stats.get("median", 0))
        _style_data_cell(ws, ri, 5, NUM_FMT)
        ws.cell(row=ri, column=6, value=stats.get("p5", 0))
        _style_data_cell(ws, ri, 6, NUM_FMT)
        ws.cell(row=ri, column=7, value=stats.get("p25", 0))
        _style_data_cell(ws, ri, 7, NUM_FMT)
        ws.cell(row=ri, column=8, value=stats.get("p75", 0))
        _style_data_cell(ws, ri, 8, NUM_FMT)
        ws.cell(row=ri, column=9, value=stats.get("p95", 0))
        _style_data_cell(ws, ri, 9, NUM_FMT)
        ws.cell(row=ri, column=10, value=stats.get("max", 0))
        _style_data_cell(ws, ri, 10, NUM_FMT)
        ws.cell(row=ri, column=11, value=stats.get("pct_above_96m", 0))
        _style_data_cell(ws, ri, 11, PCT_FMT)
        ri += 1

    # Lifecycle phases from annual timeline
    annual = cf.get("annual_timeline", [])
    if annual:
        ri += 2
        ws.cell(row=ri, column=1, value="Lifecycle Phases")
        ws.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
        ri += 1

        phase_headers = ["Year", "Month Range", "Phase", "% Resolving", "Cumulative %"]
        for ci, h in enumerate(phase_headers, 1):
            ws.cell(row=ri, column=ci, value=h)
        _style_header_row(ws, ri, len(phase_headers))
        ri += 1

        for row in annual:
            ws.cell(row=ri, column=1, value=row.get("year", ""))
            ws.cell(row=ri, column=2, value=row.get("month_range", ""))
            ws.cell(row=ri, column=3, value=row.get("phase", ""))
            ws.cell(row=ri, column=4, value=row.get("pct_resolving", 0))
            _style_data_cell(ws, ri, 4, PCT_FMT)
            ws.cell(row=ri, column=5, value=row.get("pct_cumulative", 0))
            _style_data_cell(ws, ri, 5, PCT_FMT)
            ri += 1

    _auto_width(ws)


def _write_irr_probability_table(wb, data):
    """IRR Probability Comparison table for both portfolios.

    Computes P(IRR>40%), P(IRR>30%), P(IRR>25%), P(IRR<0%) from IRR histogram.
    Used for PPT slide: "IRR Probability Comparison" table.
    """
    stoch = data.get("stochastic_pricing", {})
    grid = stoch.get("grid", {})
    if not grid:
        return

    def _irr_probs_from_grid(g):
        """Compute IRR probability thresholds from histogram."""
        _, ref = _find_stochastic_ref(g)

        irr_hist = ref.get("irr_hist", [])
        if not irr_hist:
            return {}

        total = sum(b.get("count", 0) for b in irr_hist)
        if total == 0:
            return {}

        thresholds = {
            "P(IRR > 40%)": sum(b["count"] for b in irr_hist if b.get("edge", 0) > 0.40),
            "P(IRR > 30%)": sum(b["count"] for b in irr_hist if b.get("edge", 0) > 0.30),
            "P(IRR > 25%)": sum(b["count"] for b in irr_hist if b.get("edge", 0) > 0.25),
            "P(IRR > 0%)": sum(b["count"] for b in irr_hist if b.get("edge", 0) > 0.0),
            "P(IRR < 0%)": sum(b["count"] for b in irr_hist if b.get("edge", 0) < 0.0),
            "P(Loss — MOIC < 1x)": ref.get("prob_loss", 0),
        }
        # Convert counts to proportions (except prob_loss which is already a probability)
        for k in thresholds:
            if k != "P(Loss — MOIC < 1x)":
                thresholds[k] = thresholds[k] / total
        return thresholds

    ws = wb.create_sheet(title="IRR Probability Table")
    ws.cell(row=1, column=1, value="IRR Probability Comparison (10% Upfront / 20% Tail)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")

    headers = ["Threshold", "Full Portfolio", "SIAC Portfolio"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=2, column=ci, value=h)
    _style_header_row(ws, 2, len(headers))

    full_probs = _irr_probs_from_grid(grid)
    siac_stoch = data.get("siac_stochastic_pricing", {})
    siac_grid = siac_stoch.get("grid", {})
    siac_probs = _irr_probs_from_grid(siac_grid) if siac_grid else {}

    row_order = ["P(IRR > 40%)", "P(IRR > 30%)", "P(IRR > 25%)",
                 "P(IRR > 0%)", "P(IRR < 0%)", "P(Loss — MOIC < 1x)"]

    for ri, label in enumerate(row_order, 3):
        ws.cell(row=ri, column=1, value=label)
        ws.cell(row=ri, column=1).font = Font(bold=True)
        if label in full_probs:
            ws.cell(row=ri, column=2, value=full_probs[label])
            _style_data_cell(ws, ri, 2, PCT_FMT)
        if label in siac_probs:
            ws.cell(row=ri, column=3, value=siac_probs[label])
            _style_data_cell(ws, ri, 3, PCT_FMT)

    _auto_width(ws)


def _write_breakeven(wb, data):
    """Sheet: Breakeven Analysis."""
    be = data.get("breakeven_data", {})
    if not be:
        return

    ws = wb.create_sheet(title="Breakeven Analysis")
    ri = 1

    # Write breakeven surfaces
    surfaces = be.get("surfaces", {})
    for basis_key, basis_data in surfaces.items():
        ws.cell(row=ri, column=1, value=f"Breakeven Surface — {basis_key.upper()}")
        ws.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
        ri += 1

        surface_list = basis_data.get("surface", []) if isinstance(basis_data, dict) else []
        if surface_list:
            headers = ["Tata Tail %", "Award Share %", "Max Upfront %"]
            for ci, h in enumerate(headers, 1):
                ws.cell(row=ri, column=ci, value=h)
            _style_header_row(ws, ri, len(headers))
            ri += 1

            for entry in surface_list:
                ws.cell(row=ri, column=1, value=entry.get("tata_tail_pct", 0))
                _style_data_cell(ws, ri, 1, PCT_FMT)
                ws.cell(row=ri, column=2, value=entry.get("award_share_pct", 0))
                _style_data_cell(ws, ri, 2, PCT_FMT)
                ws.cell(row=ri, column=3, value=entry.get("max_upfront_pct", 0))
                _style_data_cell(ws, ri, 3, PCT_FMT)
                ri += 1
        ri += 1

    # Write per-claim breakeven at reference tail
    per_claim = be.get("per_claim_at_30_tata_tail", {})
    if per_claim:
        ws.cell(row=ri, column=1, value="Per-Claim Breakeven (at 30% Tata Tail)")
        ws.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
        ri += 1

        headers = ["Claim ID", "SOC (Cr)", "Archetype", "SOC Breakeven %", "EQ Breakeven %"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=h)
        _style_header_row(ws, ri, len(headers))
        ri += 1

        for claim_id, claim_data in per_claim.items():
            ws.cell(row=ri, column=1, value=claim_id)
            ws.cell(row=ri, column=2, value=claim_data.get("soc_cr", 0))
            _style_data_cell(ws, ri, 2, CR_FMT)
            ws.cell(row=ri, column=3, value=claim_data.get("archetype", ""))
            ws.cell(row=ri, column=4, value=claim_data.get("soc_breakeven_pct", 0))
            _style_data_cell(ws, ri, 4, PCT_FMT)
            ws.cell(row=ri, column=5, value=claim_data.get("eq_breakeven_pct", 0))
            _style_data_cell(ws, ri, 5, PCT_FMT)
            ri += 1

    _auto_width(ws)


# ===================================================================
# Main entry point
# ===================================================================

def generate_chart_data_excel(
    output_dir: str,
    filename: str = "Chart_Data.xlsx",
    dashboard_json_name: str = "dashboard_data.json",
    stochastic_json_name: str = "stochastic_pricing.json",
    alt_output_dir: Optional[str] = None,
) -> Optional[str]:
    """Generate a Chart Data Excel workbook from dashboard JSON output.

    Parameters
    ----------
    output_dir : str
        Directory containing dashboard_data.json and stochastic_pricing.json.
    filename : str
        Output Excel filename.
    dashboard_json_name : str
        Name of the dashboard JSON file.
    stochastic_json_name : str
        Name of the stochastic pricing JSON file.
    alt_output_dir : str or None
        Optional second output directory (e.g. SIAC portfolio) for
        cross-portfolio comparison sheets. If None, attempts to
        auto-detect outputs_siac/ sibling directory.

    Returns
    -------
    str or None
        Path to generated Excel file, or None on failure.
    """
    if not HAS_OPENPYXL:
        print("  [WARN] openpyxl not installed — skipping chart data Excel")
        return None

    dashboard_path = os.path.join(output_dir, dashboard_json_name)
    if not os.path.exists(dashboard_path):
        print(f"  [WARN] {dashboard_json_name} not found in {output_dir}")
        return None

    with open(dashboard_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Also load stochastic pricing if available
    stochastic_path = os.path.join(output_dir, stochastic_json_name)
    if os.path.exists(stochastic_path):
        with open(stochastic_path, "r", encoding="utf-8") as f:
            stoch_data = json.load(f)
        data["stochastic_pricing"] = stoch_data

    # ── Load SIAC / comparison portfolio data ──
    if alt_output_dir is None:
        # Auto-detect sibling outputs_siac/ directory
        parent = os.path.dirname(os.path.abspath(output_dir))
        candidate = os.path.join(parent, "outputs_siac")
        if os.path.isdir(candidate):
            alt_output_dir = candidate

    if alt_output_dir:
        siac_dash = os.path.join(alt_output_dir, dashboard_json_name)
        siac_stoch = os.path.join(alt_output_dir, stochastic_json_name)
        if os.path.exists(siac_dash):
            with open(siac_dash, "r", encoding="utf-8") as f:
                data["siac_dashboard_data"] = json.load(f)
            print(f"  Loaded SIAC dashboard data from {alt_output_dir}")
        if os.path.exists(siac_stoch):
            with open(siac_stoch, "r", encoding="utf-8") as f:
                data["siac_stochastic_pricing"] = json.load(f)
            print(f"  Loaded SIAC stochastic data from {alt_output_dir}")

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print(f"  Generating Chart Data Excel: {filename}")

    _write_moic_distribution(wb, data)
    _write_irr_analysis(wb, data)
    _write_return_sensitivity(wb, data)
    _write_cashflow_timeline(wb, data)
    _write_jcurve(wb, data)
    _write_waterfall(wb, data)
    _write_risk_metrics(wb, data)
    _write_per_claim_summary(wb, data)
    _write_stochastic_grid(wb, data)
    _write_simulation_summary(wb, data)
    _write_scenario_comparison(wb, data)
    _write_breakeven(wb, data)

    # ── New PPT-oriented sheets ──
    _write_irr_histogram(wb, data)
    _write_portfolio_comparison(wb, data)
    _write_loss_sensitivity(wb, data)
    _write_scenario_bands(wb, data)
    _write_timeline_summary(wb, data)
    _write_irr_probability_table(wb, data)

    out_path = os.path.join(output_dir, filename)
    wb.save(out_path)
    print(f"  Chart Data Excel saved: {out_path}")
    print(f"    Sheets: {len(wb.sheetnames)}")
    return out_path


# ── CLI entry point (standalone) ──
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Generate Chart Data Excel from dashboard JSON")
    parser.add_argument("--output-dir", default="outputs", help="Directory containing dashboard_data.json")
    parser.add_argument("--filename", default="Chart_Data.xlsx", help="Output filename")
    parser.add_argument("--alt-output-dir", default=None, help="SIAC/comparison portfolio output directory")
    args = parser.parse_args()

    result = generate_chart_data_excel(
        output_dir=args.output_dir,
        filename=args.filename,
        alt_output_dir=args.alt_output_dir,
    )
    if result:
        print(f"\nDone: {result}")
    else:
        print("\nFailed to generate chart data Excel.")
