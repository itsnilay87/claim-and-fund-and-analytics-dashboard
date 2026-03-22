"""
TATA_code_v2/v2_excel_writer.py — 14-sheet Excel workbook generator.
=====================================================================

Produces TATA_V2_Valuation_Model.xlsx with EXACTLY 14 sheets:
  1.  Executive Summary    — portfolio overview, key investment scenario
  2.  Model Assumptions    — every parameter (MI sections 1-10)
  3.  Probability Analysis — domestic/SIAC tree paths, combined outcomes
  4.  TP-301-6             — per-claim detail  (×6 sheets: 4–9)
  5.  TP-302-3
  6.  TP-302-5
  7.  TP-CTP11-2
  8.  TP-CTP11-4
  9.  TP-CTP13-2
  10. Portfolio Analysis    — aggregated MOIC grid, conditional formatting
  11. Breakeven Analysis    — max upfront where E[MOIC] >= 1.0
  12. Scenario Comparison   — verdicts (Strong Buy / Attractive / Marginal / Avoid)
  13. Legal Costs           — per-claim per-stage breakdown
  14. Glossary & Notes      — definitions, methodology, caveats

Formatting follows CLAUDE.md: Arial, blue font for inputs, yellow key
results, white on #2E75B6 headers, thin grey borders, col A = 3-width spacer.
All monetary values in ₹ Crore.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import v2_master_inputs as MI
from .v2_config import ClaimConfig, SimulationResults
from .v2_investment_analysis import InvestmentGridResults


# ===================================================================
# Style Constants
# ===================================================================

# Fonts
TITLE_FONT    = Font(name="Arial", bold=True, size=14, color="1F4E79")
SECTION_FONT  = Font(name="Arial", bold=True, size=12, color="2E75B6")
SUBSECTION_FONT = Font(name="Arial", bold=True, size=10, color="404040")
HEADER_FONT   = Font(name="Arial", bold=True, size=9, color="FFFFFF")
NORMAL_FONT   = Font(name="Arial", size=9)
BLUE_FONT     = Font(name="Arial", size=9, color="0000FF")
BOLD_FONT     = Font(name="Arial", bold=True, size=9)
SMALL_FONT    = Font(name="Arial", size=8, color="808080")
WARNING_FONT  = Font(name="Arial", size=9, color="FF4444")

# Fills
HEADER_FILL   = PatternFill("solid", fgColor="2E75B6")
BLUE_FILL     = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL    = PatternFill("solid", fgColor="E2EFDA")
YELLOW_FILL   = PatternFill("solid", fgColor="FFF2CC")
RED_FILL      = PatternFill("solid", fgColor="FCE4EC")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color="2E75B6"))

# Alignments
CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_W   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_A  = Alignment(horizontal="right", vertical="center")

# Number formats
PCT_FMT  = "0.0%"
CR_FMT   = "#,##0.00"
INT_FMT  = "#,##0"
MOIC_FMT = '0.00"×"'
MONTH_FMT = "0.0"


# ===================================================================
# Helper Utilities
# ===================================================================

def _setup_sheet(ws, title: str) -> int:
    """Standard sheet setup: col A = 3 width spacer, title in B1. Returns row=3."""
    ws.sheet_properties.tabColor = "2E75B6"
    ws.column_dimensions["A"].width = 3
    ws["B1"].value = title
    ws["B1"].font = TITLE_FONT
    ws["B1"].alignment = LEFT_W
    return 3


def _header_row(ws, row: int, cols: list[str], start_col: int = 2) -> int:
    """Write a header row with bold white text on blue fill. Returns row+1."""
    for j, hdr in enumerate(cols):
        cell = ws.cell(row=row, column=start_col + j, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    return row + 1


def _data_row(ws, row: int, values: list, start_col: int = 2,
              fonts: Optional[list] = None, fmts: Optional[list] = None,
              fills: Optional[list] = None) -> int:
    """Write one data row. Returns row+1."""
    for j, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + j, value=val)
        cell.font = (fonts[j] if fonts and j < len(fonts) else NORMAL_FONT)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if fmts and j < len(fmts) and fmts[j]:
            cell.number_format = fmts[j]
        if fills and j < len(fills) and fills[j]:
            cell.fill = fills[j]
    return row + 1


def _section_header(ws, row: int, text: str, col: int = 2) -> int:
    """Write a section header. Returns row+1."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SECTION_FONT
    cell.alignment = LEFT_W
    return row + 1


def _kv_row(ws, row: int, key: str, value, col: int = 2,
            val_font=None, val_fmt: str = "") -> int:
    """Write key-value pair in two columns. Returns row+1."""
    ws.cell(row=row, column=col, value=key).font = BOLD_FONT
    ws.cell(row=row, column=col, value=key).border = THIN_BORDER
    c = ws.cell(row=row, column=col + 1, value=value)
    c.font = val_font or BLUE_FONT
    c.border = THIN_BORDER
    if val_fmt:
        c.number_format = val_fmt
    return row + 1


def _auto_width(ws, min_width: float = 10, max_width: float = 35) -> None:
    """Auto-approximate column widths (skip col A spacer)."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        if col_letter == "A":
            continue
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) * 1.15, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _verdict(moic: float, p_loss: float) -> str:
    """Classify into verdict label."""
    if moic > 2.5 and p_loss < 0.10:
        return "Strong Buy"
    elif moic > 1.5 and p_loss < 0.25:
        return "Attractive"
    elif moic > 1.0 and p_loss < 0.40:
        return "Marginal"
    return "Avoid"


def _verdict_fill(verdict: str) -> PatternFill:
    if verdict == "Strong Buy":
        return GREEN_FILL
    elif verdict == "Attractive":
        return BLUE_FILL
    elif verdict == "Marginal":
        return YELLOW_FILL
    return RED_FILL


# ===================================================================
# Sheet Builders
# ===================================================================

def _build_executive_summary(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    claims: list[ClaimConfig],
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 1: Executive Summary."""
    ws = wb.active
    ws.title = "Executive Summary"
    row = _setup_sheet(ws, "TATA v2 Valuation Model — Executive Summary")

    # Portfolio Overview
    row = _section_header(ws, row, "Portfolio Overview")
    total_soc = sum(c.soc_value_cr for c in claims)
    row = _kv_row(ws, row, "Number of Claims", len(claims))
    row = _kv_row(ws, row, "Total SOC (₹ Cr)", total_soc, val_fmt=CR_FMT)
    row = _kv_row(ws, row, "Monte Carlo Paths", f"{sim.n_paths:,}")
    row = _kv_row(ws, row, "Random Seed", sim.seed)
    row = _kv_row(ws, row, "Pricing Basis", basis.upper())
    row = _kv_row(ws, row, "Start Date", MI.START_DATE)
    row += 1

    # Per-claim overview table
    row = _section_header(ws, row, "Claim Overview")
    cols = ["Claim", "Archetype", "SOC (₹ Cr)", "Jurisdiction",
            "Win Rate", "E[Duration] (mo)", "E[Q] (₹ Cr)"]
    row = _header_row(ws, row, cols)

    claim_map = {c.claim_id: c for c in claims}
    for cid in sim.claim_ids:
        c = claim_map[cid]
        wr = sim.win_rate_map.get(cid, 0)
        dur = sim.mean_duration_map.get(cid, 0)
        eq = sim.expected_quantum_map.get(cid, 0)
        row = _data_row(ws, row,
                        [cid, c.archetype.title(), c.soc_value_cr,
                         c.jurisdiction.upper(), wr, dur, eq],
                        fmts=[None, None, CR_FMT, None, PCT_FMT, MONTH_FMT, CR_FMT])
    row += 1

    # Reference Scenario
    ref_up, ref_aw = 0.10, 0.80
    key = (ref_up, ref_aw, basis)
    cell = grid.cells.get(key)
    if cell:
        tata_tail = 1.0 - ref_aw
        row = _section_header(ws, row, f"Reference Scenario: {ref_up:.0%} Upfront / {tata_tail:.0%} Tata Tail ({basis.upper()})")
        row = _kv_row(ws, row, "E[MOIC]", cell.mean_moic, val_fmt=MOIC_FMT)
        row = _kv_row(ws, row, "Median MOIC", cell.median_moic, val_fmt=MOIC_FMT)
        row = _kv_row(ws, row, "E[XIRR]", cell.mean_xirr, val_fmt=PCT_FMT)
        row = _kv_row(ws, row, "Median XIRR", cell.median_xirr, val_fmt=PCT_FMT)
        row = _kv_row(ws, row, "E[Net Return] (₹ Cr)", cell.mean_net_return_cr, val_fmt=CR_FMT)
        row = _kv_row(ws, row, "P(Capital Loss)", cell.p_loss, val_fmt=PCT_FMT)
        row = _kv_row(ws, row, "P(IRR > 30%)", cell.p_irr_gt_30, val_fmt=PCT_FMT)
        row = _kv_row(ws, row, "P(IRR > 25%)", cell.p_irr_gt_25, val_fmt=PCT_FMT)
        row = _kv_row(ws, row, "VaR 1% (₹ Cr)", cell.var_1, val_fmt=CR_FMT)
        row = _kv_row(ws, row, "CVaR 1% (₹ Cr)", cell.cvar_1, val_fmt=CR_FMT)

        v = _verdict(cell.mean_moic, cell.p_loss)
        vr = _kv_row(ws, row, "Investment Verdict", v)
        ws.cell(row=row, column=3).fill = _verdict_fill(v)
        row = vr

    # Highlight key result cells
    for r in range(3, row):
        c3 = ws.cell(row=r, column=3)
        if c3.font == BLUE_FONT:
            c3.fill = BLUE_FILL

    _auto_width(ws)


def _build_model_assumptions(
    wb: openpyxl.Workbook,
    claims: list[ClaimConfig],
) -> None:
    """Sheet 2: Model Assumptions — every MI parameter."""
    ws = wb.create_sheet("Model Assumptions")
    row = _setup_sheet(ws, "Model Assumptions")

    # Section 1: Simulation
    row = _section_header(ws, row, "1. Simulation Engine")
    cols = ["Parameter", "Value", "Description"]
    row = _header_row(ws, row, cols)
    _params = [
        ("N_SIMULATIONS", MI.N_SIMULATIONS, "Monte Carlo paths"),
        ("RANDOM_SEED", MI.RANDOM_SEED, "Base RNG seed"),
        ("MAX_TIMELINE_MONTHS", MI.MAX_TIMELINE_MONTHS, "Re-arb cutoff (months)"),
        ("START_DATE", MI.START_DATE, "Investment anchor date"),
    ]
    for p, v, d in _params:
        row = _data_row(ws, row, [p, v, d],
                        fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT])
    row += 1

    # Section 4: Arb Outcome
    row = _section_header(ws, row, "2. Arbitration Outcome")
    row = _header_row(ws, row, cols)
    row = _data_row(ws, row, ["ARB_WIN_PROBABILITY", MI.ARB_WIN_PROBABILITY, "P(TATA wins arb)"],
                    fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT], fmts=[None, PCT_FMT, None])
    row = _data_row(ws, row, ["RE_ARB_WIN_PROBABILITY", MI.RE_ARB_WIN_PROBABILITY, "P(TATA wins re-arb)"],
                    fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT], fmts=[None, PCT_FMT, None])
    row += 1

    # Section 5: Quantum Bands
    row = _section_header(ws, row, "3. Quantum Bands (conditional on WIN)")
    row = _header_row(ws, row, ["Band", "Low", "High", "Probability", "E[Q] contribution"])
    for i, b in enumerate(MI.QUANTUM_BANDS):
        e_contrib = b["probability"] * (b["low"] + b["high"]) / 2
        row = _data_row(ws, row,
                        [f"Band {i+1}", b["low"], b["high"], b["probability"], e_contrib],
                        fmts=[None, PCT_FMT, PCT_FMT, PCT_FMT, PCT_FMT],
                        fonts=[NORMAL_FONT, BLUE_FONT, BLUE_FONT, BLUE_FONT, NORMAL_FONT])
    total_eq = sum(b["probability"] * (b["low"] + b["high"]) / 2 for b in MI.QUANTUM_BANDS)
    row = _data_row(ws, row, ["Total E[Q|WIN]", "", "", "", total_eq],
                    fonts=[BOLD_FONT, None, None, None, BOLD_FONT],
                    fmts=[None, None, None, None, PCT_FMT],
                    fills=[YELLOW_FILL, None, None, None, YELLOW_FILL])
    row += 1

    # Section 3: Timeline durations
    row = _section_header(ws, row, "4. Timeline Durations (months)")
    row = _header_row(ws, row, ["Stage", "Low", "High", "Note"])
    _stages = [
        ("DAB Duration", MI.DAB_DURATION["low"], MI.DAB_DURATION["high"], "Uniform"),
        ("Arbitration Duration", MI.ARB_DURATION["low"], MI.ARB_DURATION["high"], "Uniform"),
        ("Arb Remaining (302-5)", MI.ARB_REMAINING_302_5["low"], MI.ARB_REMAINING_302_5["high"], "302-5 only"),
        ("Re-referral (CTP11-2)", MI.RE_REFERRAL_CTP11_2["low"], MI.RE_REFERRAL_CTP11_2["high"], "CTP11-2 only"),
        ("S.34 Duration", MI.S34_DURATION["low"], MI.S34_DURATION["high"], "Domestic courts"),
        ("S.37 Duration", MI.S37_DURATION["low"], MI.S37_DURATION["high"], "Domestic courts"),
        ("SLP Dismissed", MI.SLP_DISMISSED_DURATION, MI.SLP_DISMISSED_DURATION, "Fixed"),
        ("SLP Admitted", MI.SLP_ADMITTED_DURATION, MI.SLP_ADMITTED_DURATION, "Fixed"),
        ("SIAC HC Duration", MI.SIAC_HC_DURATION, MI.SIAC_HC_DURATION, "Fixed"),
        ("SIAC COA Duration", MI.SIAC_COA_DURATION, MI.SIAC_COA_DURATION, "Fixed"),
        ("Domestic Payment Delay", MI.DOMESTIC_PAYMENT_DELAY, MI.DOMESTIC_PAYMENT_DELAY, "Post-resolution"),
        ("SIAC Payment Delay", MI.SIAC_PAYMENT_DELAY, MI.SIAC_PAYMENT_DELAY, "Post-resolution"),
    ]
    for name, lo, hi, note in _stages:
        row = _data_row(ws, row, [name, lo, hi, note],
                        fonts=[NORMAL_FONT, BLUE_FONT, BLUE_FONT, SMALL_FONT],
                        fmts=[None, MONTH_FMT, MONTH_FMT, None])
    row += 1

    # Section 8: Investment structure
    row = _section_header(ws, row, "5. Investment Grid")
    row = _kv_row(ws, row, "Upfront % levels", str(MI.UPFRONT_PCT_SOC))
    row = _kv_row(ws, row, "Tata Tail % levels", str(MI.TATA_TAIL_PCT))
    row = _kv_row(ws, row, "Grid size", f"{len(MI.UPFRONT_PCT_SOC)} × {len(MI.TATA_TAIL_PCT)} = "
                  f"{len(MI.UPFRONT_PCT_SOC) * len(MI.TATA_TAIL_PCT)}")
    row += 1

    # Section 10: Financial
    row = _section_header(ws, row, "6. Financial Parameters")
    row = _header_row(ws, row, cols)
    row = _data_row(ws, row, ["DISCOUNT_RATE", MI.DISCOUNT_RATE, "Annual hurdle rate"],
                    fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT], fmts=[None, PCT_FMT, None])
    row = _data_row(ws, row, ["RISK_FREE_RATE", MI.RISK_FREE_RATE, "Annualized risk-free rate"],
                    fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT], fmts=[None, PCT_FMT, None])

    _auto_width(ws)


def _build_probability_analysis(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
) -> None:
    """Sheet 3: Probability Analysis — prob tree paths & combined outcomes."""
    ws = wb.create_sheet("Probability Analysis")
    row = _setup_sheet(ws, "Probability Analysis")

    # Domestic paths
    row = _section_header(ws, row, "Domestic Challenge Tree Paths")
    row = _header_row(ws, row, ["Path", "Outcome", "Probability", "Duration (mo)", "Description"])

    for paths_list, label in [
        (MI.DOMESTIC_PATHS_A, "Scenario A (TATA won arb)"),
        (MI.DOMESTIC_PATHS_B, "Scenario B (TATA lost arb)"),
    ]:
        # Sub-header
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = SUBSECTION_FONT
        row += 1
        for p in paths_list:
            outcome_fill = GREEN_FILL if p["outcome"] == "TRUE_WIN" else (
                YELLOW_FILL if p["outcome"] == "RESTART" else RED_FILL
            )
            row = _data_row(ws, row,
                            [p["path_id"], p["outcome"],
                             p["conditional_prob"],
                             p.get("slp_duration_months", 0),
                             p.get("description", "")],
                            fmts=[None, None, PCT_FMT, MONTH_FMT, None],
                            fills=[None, outcome_fill, None, None, None])
        row += 1

    # SIAC paths
    row = _section_header(ws, row, "SIAC Challenge Tree Paths")
    row = _header_row(ws, row, ["Path", "Outcome", "Probability", "Duration (mo)", "Description"])

    for paths_list, label in [
        (MI.SIAC_PATHS_A, "Scenario A (TATA won arb)"),
        (MI.SIAC_PATHS_B, "Scenario B (TATA lost arb)"),
    ]:
        cell = ws.cell(row=row, column=2, value=label)
        cell.font = SUBSECTION_FONT
        row += 1
        for p in paths_list:
            outcome_fill = GREEN_FILL if p["outcome"] == "TRUE_WIN" else (
                YELLOW_FILL if p["outcome"] == "RESTART" else RED_FILL
            )
            # SIAC paths have fixed duration (HC + COA)
            dur = MI.SIAC_HC_DURATION + MI.SIAC_COA_DURATION
            row = _data_row(ws, row,
                            [p["path_id"], p["outcome"],
                             p["conditional_prob"], dur,
                             p.get("description", "")],
                            fmts=[None, None, PCT_FMT, MONTH_FMT, None],
                            fills=[None, outcome_fill, None, None, None])
        row += 1

    # Combined outcomes summary
    row = _section_header(ws, row, "Combined Outcome Summary (from MC)")
    row = _header_row(ws, row, ["Claim", "P(TRUE_WIN)", "P(RESTART)", "P(LOSE)",
                                "Win Rate (post-reArb)"])
    for cid in sim.claim_ids:
        paths = sim.results[cid]
        n = len(paths)
        p_tw = sum(1 for p in paths if p.final_outcome == "TRUE_WIN") / n
        p_re = sum(1 for p in paths if p.final_outcome == "RESTART") / n
        p_lo = sum(1 for p in paths if p.final_outcome == "LOSE") / n
        wr = sim.win_rate_map.get(cid, 0)
        row = _data_row(ws, row, [cid, p_tw, p_re, p_lo, wr],
                        fmts=[None, PCT_FMT, PCT_FMT, PCT_FMT, PCT_FMT])

    _auto_width(ws)


def _build_per_claim_sheet(
    wb: openpyxl.Workbook,
    cid: str,
    sim: SimulationResults,
    claim: ClaimConfig,
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheets 4–9: Per-claim detail."""
    ws = wb.create_sheet(cid)
    row = _setup_sheet(ws, f"Claim Analysis — {cid}")

    # Identity
    row = _section_header(ws, row, "Claim Identity")
    row = _kv_row(ws, row, "Claim ID", cid)
    row = _kv_row(ws, row, "Archetype", claim.archetype.title())
    row = _kv_row(ws, row, "SOC (₹ Cr)", claim.soc_value_cr, val_fmt=CR_FMT)
    row = _kv_row(ws, row, "Jurisdiction", claim.jurisdiction.upper())
    row = _kv_row(ws, row, "Current Gate", claim.current_gate)
    row = _kv_row(ws, row, "TPL Share", claim.tpl_share, val_fmt=PCT_FMT)
    row = _kv_row(ws, row, "Pipeline", " → ".join(claim.pipeline))
    row += 1

    # Simulation results
    row = _section_header(ws, row, "Monte Carlo Results")
    paths = sim.results.get(cid, [])
    n = len(paths)
    if n == 0:
        ws.cell(row=row, column=2, value="No simulation data").font = WARNING_FONT
        _auto_width(ws)
        return

    arb_win_rate = sum(1 for p in paths if p.arb_won) / n
    final_win_rate = sim.win_rate_map.get(cid, 0)
    durations = [p.total_duration_months for p in paths]
    legal_costs = [p.legal_cost_total_cr for p in paths]
    collected = [p.collected_cr for p in paths]

    eq_cr = sim.expected_quantum_map.get(cid, 0)

    row = _kv_row(ws, row, "Arb Win Rate", arb_win_rate, val_fmt=PCT_FMT)
    row = _kv_row(ws, row, "Final Win Rate", final_win_rate, val_fmt=PCT_FMT)
    row = _kv_row(ws, row, "E[Duration] (mo)", float(np.mean(durations)), val_fmt=MONTH_FMT)
    row = _kv_row(ws, row, "Median Duration (mo)", float(np.median(durations)), val_fmt=MONTH_FMT)
    row = _kv_row(ws, row, "E[Legal Cost] (₹ Cr)", float(np.mean(legal_costs)), val_fmt=CR_FMT)
    row = _kv_row(ws, row, "E[Collected] (₹ Cr)", float(np.mean(collected)), val_fmt=CR_FMT)
    row = _kv_row(ws, row, "E[Q] analytical (₹ Cr)", eq_cr, val_fmt=CR_FMT)
    row += 1

    # Percentile table
    row = _section_header(ws, row, "Duration & Quantum Percentiles")
    pctiles = [5, 25, 50, 75, 95]
    row = _header_row(ws, row, ["Percentile"] + [f"P{p}" for p in pctiles])
    dur_arr = np.array(durations)
    coll_arr = np.array(collected)
    dur_pcts = [float(np.percentile(dur_arr, p)) for p in pctiles]
    coll_pcts = [float(np.percentile(coll_arr, p)) for p in pctiles]
    row = _data_row(ws, row, ["Duration (mo)"] + dur_pcts,
                    fmts=[None] + [MONTH_FMT] * len(pctiles))
    row = _data_row(ws, row, ["Collected (₹ Cr)"] + coll_pcts,
                    fmts=[None] + [CR_FMT] * len(pctiles))
    row += 1

    # Investment grid sub-table (for this claim only)
    row = _section_header(ws, row, f"Per-Claim Investment Grid ({basis.upper()} Pricing)")
    row = _header_row(ws, row, ["Upfront %", "Tata Tail %", "E[MOIC]",
                                "E[XIRR]", "P(Loss)", "E[Net Return ₹ Cr]"])

    # Show a selection of grid cells for this claim
    ref_ups = [0.05, 0.10, 0.15, 0.20]
    ref_aws = [0.90, 0.80, 0.70, 0.50]  # Tata tail: 10%, 20%, 30%, 50%
    for up in ref_ups:
        for aw in ref_aws:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            if cell_data and cid in cell_data.per_claim:
                pc = cell_data.per_claim[cid]
                moic_val = pc["E[MOIC]"]
                fill = GREEN_FILL if moic_val > 1.5 else (
                    YELLOW_FILL if moic_val > 1.0 else RED_FILL
                )
                row = _data_row(ws, row,
                                [up, 1-aw, moic_val, pc["E[XIRR]"],
                                 pc["P(loss)"], pc["E[net_return_cr]"]],
                                fmts=[PCT_FMT, PCT_FMT, MOIC_FMT, PCT_FMT,
                                      PCT_FMT, CR_FMT],
                                fills=[None, None, fill, None, None, None])

    _auto_width(ws)


def _build_portfolio_analysis(
    wb: openpyxl.Workbook,
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 10: Portfolio Analysis — full MOIC grid with conditional formatting."""
    ws = wb.create_sheet("Portfolio Analysis")
    row = _setup_sheet(ws, f"Portfolio Analysis — {basis.upper()} Pricing")

    # Full E[MOIC] grid
    row = _section_header(ws, row, "E[MOIC] Grid (Portfolio Level)")
    up_pcts = sorted(grid.upfront_pcts)
    aw_pcts = sorted(grid.award_share_pcts)

    # Header row — show Tata Tail % (complement of fund's award share)
    headers = ["Upfront \\ Tata Tail"] + [f"{1-p:.0%}" for p in aw_pcts]
    row = _header_row(ws, row, headers)

    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fmts_row = [None]
        fills_row = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            moic = cell_data.mean_moic if cell_data else float("nan")
            vals.append(moic)
            fmts_row.append(MOIC_FMT)
            if moic > 1.5:
                fills_row.append(GREEN_FILL)
            elif moic >= 1.0:
                fills_row.append(YELLOW_FILL)
            else:
                fills_row.append(RED_FILL)
        row = _data_row(ws, row, vals, fmts=fmts_row, fills=fills_row)
    row += 1

    # E[XIRR] grid
    row = _section_header(ws, row, "E[XIRR] Grid (Portfolio Level)")
    row = _header_row(ws, row, headers)
    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fmts_row = [None]
        fills_row = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            xirr = cell_data.mean_xirr if cell_data else float("nan")
            vals.append(xirr)
            fmts_row.append(PCT_FMT)
            if xirr > 0.30:
                fills_row.append(GREEN_FILL)
            elif xirr >= 0.0:
                fills_row.append(YELLOW_FILL)
            else:
                fills_row.append(RED_FILL)
        row = _data_row(ws, row, vals, fmts=fmts_row, fills=fills_row)
    row += 1

    # P(Loss) grid
    row = _section_header(ws, row, "P(Capital Loss) Grid")
    row = _header_row(ws, row, headers)
    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fmts_row = [None]
        fills_row = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            ploss = cell_data.p_loss if cell_data else float("nan")
            vals.append(ploss)
            fmts_row.append(PCT_FMT)
            if ploss < 0.10:
                fills_row.append(GREEN_FILL)
            elif ploss < 0.25:
                fills_row.append(YELLOW_FILL)
            else:
                fills_row.append(RED_FILL)
        row = _data_row(ws, row, vals, fmts=fmts_row, fills=fills_row)
    row += 1

    # VaR grid
    row = _section_header(ws, row, "VaR 1% Grid (₹ Crore)")
    row = _header_row(ws, row, headers)
    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fmts_row = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            var5 = cell_data.var_1 if cell_data else float("nan")
            vals.append(var5)
            fmts_row.append(CR_FMT)
        row = _data_row(ws, row, vals, fmts=fmts_row)

    _auto_width(ws)


def _build_breakeven_analysis(
    wb: openpyxl.Workbook,
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 11: Breakeven Analysis — max upfront where E[MOIC] >= 1.0."""
    ws = wb.create_sheet("Breakeven Analysis")
    row = _setup_sheet(ws, "Breakeven Analysis")

    be = grid.breakeven.get(basis, {})
    aw_pcts = sorted(grid.award_share_pcts)

    row = _section_header(ws, row, f"Maximum Viable Upfront % ({basis.upper()} Pricing)")
    row = _header_row(ws, row, ["Tata Tail %", "Max Upfront %",
                                "E[MOIC] at Breakeven", "E[XIRR] at Breakeven", "Verdict"])

    for aw in aw_pcts:
        max_up = be.get(aw, 0)
        key = (max_up, aw, basis)
        cell_data = grid.cells.get(key)
        moic = cell_data.mean_moic if cell_data else 0
        xirr = cell_data.mean_xirr if cell_data else 0
        v = _verdict(moic, cell_data.p_loss if cell_data else 1.0)
        vf = _verdict_fill(v)

        row = _data_row(ws, row, [1-aw, max_up, moic, xirr, v],
                        fmts=[PCT_FMT, PCT_FMT, MOIC_FMT, PCT_FMT, None],
                        fills=[None, YELLOW_FILL, None, None, vf])
    row += 1

    # Summary
    row = _section_header(ws, row, "Interpretation")
    notes = [
        "Breakeven = maximum upfront % of SOC where the portfolio E[MOIC] ≥ 1.0.",
        "Higher award shares allow larger upfront investments while maintaining profitability.",
        "Values below 5% indicate the scenario is not viable at any tested upfront level.",
    ]
    for note in notes:
        ws.cell(row=row, column=2, value=note).font = SMALL_FONT
        row += 1

    _auto_width(ws)


def _build_scenario_comparison(
    wb: openpyxl.Workbook,
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 12: Scenario Comparison — verdicts across the grid."""
    ws = wb.create_sheet("Scenario Comparison")
    row = _setup_sheet(ws, "Scenario Comparison — Investment Decision Matrix")

    up_pcts = sorted(grid.upfront_pcts)
    aw_pcts = sorted(grid.award_share_pcts)

    row = _section_header(ws, row, f"Verdict Grid ({basis.upper()} Pricing)")
    headers = ["Upfront \\ Tata Tail"] + [f"{1-p:.0%}" for p in aw_pcts]
    row = _header_row(ws, row, headers)

    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fills_row = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            if cell_data:
                v = _verdict(cell_data.mean_moic, cell_data.p_loss)
            else:
                v = "N/A"
            vals.append(v)
            fills_row.append(_verdict_fill(v) if v != "N/A" else None)
        row = _data_row(ws, row, vals, fills=fills_row)
    row += 1

    # Detail table for key scenarios
    row = _section_header(ws, row, "Key Scenario Details")
    row = _header_row(ws, row, ["Upfront %", "Tata Tail %", "E[MOIC]", "E[XIRR]", "Std[MOIC]",
                                "P(Loss)", "P(IRR>30%)", "VaR 1%", "Verdict"])

    key_scenarios = [
        (0.05, 0.80), (0.10, 0.80), (0.15, 0.80),
        (0.10, 0.90), (0.10, 0.70), (0.10, 0.50),
        (0.20, 0.70), (0.25, 0.55), (0.30, 0.50),
    ]
    for up, aw in key_scenarios:
        key = (up, aw, basis)
        cell_data = grid.cells.get(key)
        if cell_data:
            v = _verdict(cell_data.mean_moic, cell_data.p_loss)
            row = _data_row(ws, row,
                            [up, 1-aw, cell_data.mean_moic, cell_data.mean_xirr,
                             cell_data.std_moic,
                             cell_data.p_loss, cell_data.p_irr_gt_30,
                             cell_data.var_1, v],
                            fmts=[PCT_FMT, PCT_FMT, MOIC_FMT, PCT_FMT, MOIC_FMT,
                                  PCT_FMT, PCT_FMT, CR_FMT, None],
                            fills=[None, None, None, None, None, None, None, None,
                                   _verdict_fill(v)])

    _auto_width(ws)


def _build_legal_costs(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    claims: list[ClaimConfig],
) -> None:
    """Sheet 13: Legal Costs — per-claim per-stage breakdown."""
    ws = wb.create_sheet("Legal Costs")
    row = _setup_sheet(ws, "Legal Cost Analysis")

    claim_map = {c.claim_id: c for c in claims}

    # Per-claim summary
    row = _section_header(ws, row, "Legal Cost Summary (from MC Simulation)")
    row = _header_row(ws, row, ["Claim", "E[Total Cost] (₹ Cr)",
                                "Median (₹ Cr)", "Std (₹ Cr)",
                                "P5 (₹ Cr)", "P95 (₹ Cr)"])

    portfolio_total = 0
    for cid in sim.claim_ids:
        paths = sim.results.get(cid, [])
        costs = [p.legal_cost_total_cr for p in paths]
        if costs:
            e_cost = float(np.mean(costs))
            portfolio_total += e_cost
            row = _data_row(ws, row,
                            [cid, e_cost, float(np.median(costs)),
                             float(np.std(costs)),
                             float(np.percentile(costs, 5)),
                             float(np.percentile(costs, 95))],
                            fmts=[None, CR_FMT, CR_FMT, CR_FMT, CR_FMT, CR_FMT])
    row += 1

    # Portfolio total
    row = _data_row(ws, row, ["PORTFOLIO TOTAL", portfolio_total, "", "", "", ""],
                    fonts=[BOLD_FONT, BOLD_FONT, None, None, None, None],
                    fmts=[None, CR_FMT, None, None, None, None],
                    fills=[YELLOW_FILL, YELLOW_FILL, None, None, None, None])
    row += 1

    # Default cost rates from MI (new model: one-time + duration-based)
    row = _section_header(ws, row, "Cost Structure (from master_inputs.py)")

    # One-time costs
    row = _header_row(ws, row, ["Category", "Component", "Cost (₹ Cr)", "", ""])
    onetime = MI.LEGAL_COSTS["onetime"]
    for component, cost in onetime.items():
        row = _data_row(ws, row,
                        ["One-time (Month 0)", component, cost, "", ""],
                        fmts=[None, None, CR_FMT, None, None],
                        fonts=[NORMAL_FONT, NORMAL_FONT, BLUE_FONT, None, None])
    total_onetime = sum(onetime.values())
    row = _data_row(ws, row,
                    ["One-time Total", "", total_onetime, "", ""],
                    fmts=[None, None, CR_FMT, None, None],
                    fonts=[BOLD_FONT, None, BOLD_FONT, None, None])
    row += 1

    # Duration-based costs
    row = _header_row(ws, row, ["Stage", "Low (₹ Cr)", "High (₹ Cr)",
                                "Midpoint (₹ Cr)", "Notes"])
    db = MI.LEGAL_COSTS["duration_based"]
    for stage_key, val in db.items():
        if isinstance(val, dict):
            low = val["low"]
            high = val["high"]
            mid = (low + high) / 2
            row = _data_row(ws, row,
                            [stage_key, low, high, mid, "range"],
                            fmts=[None, CR_FMT, CR_FMT, CR_FMT, None],
                            fonts=[NORMAL_FONT, BLUE_FONT, BLUE_FONT,
                                   NORMAL_FONT, NORMAL_FONT])
        else:
            row = _data_row(ws, row,
                            [stage_key, val, val, val, "fixed"],
                            fmts=[None, CR_FMT, CR_FMT, CR_FMT, None],
                            fonts=[NORMAL_FONT, BLUE_FONT, BLUE_FONT,
                                   NORMAL_FONT, NORMAL_FONT])
    row += 1

    # Overrun parameters
    row = _section_header(ws, row, "Legal Cost Overrun Distribution")
    ov = MI.LEGAL_COST_OVERRUN
    row = _kv_row(ws, row, "Distribution", f"ScaledBeta(α={ov['alpha']}, β={ov['beta']})")
    row = _kv_row(ws, row, "Range", f"[{ov['low']:.0%}, {ov['high']:.0%}]")
    e_overrun = ov["low"] + (ov["alpha"] / (ov["alpha"] + ov["beta"])) * (ov["high"] - ov["low"])
    row = _kv_row(ws, row, "E[Overrun]", e_overrun, val_fmt=PCT_FMT)

    _auto_width(ws)


def _build_glossary(wb: openpyxl.Workbook) -> None:
    """Sheet 14: Glossary & Notes."""
    ws = wb.create_sheet("Glossary & Notes")
    row = _setup_sheet(ws, "Glossary & Notes")

    # Terms
    row = _section_header(ws, row, "Key Terms")
    terms = [
        ("SOC", "Statement of Claim — the total claimed amount (₹ Crore)."),
        ("E[Q]", "Expected Quantum — analytical mean of award conditional on winning arbitration."),
        ("E[Q|WIN]", "Quantum as fraction of SOC, conditional on arb win (= 72.00%)."),
        ("MOIC", "Multiple on Invested Capital = total return ÷ total invested."),
        ("XIRR", "Extended Internal Rate of Return — annualized time-weighted return."),
        ("P(Loss)", "Probability of capital loss (MOIC < 1.0)."),
        ("VaR 1%", "Value at Risk — 1st percentile of net return distribution."),
        ("CVaR 1%", "Conditional VaR — expected loss given we are in the worst 1%."),
        ("TPL Share", "Third Party Litigation funder's ownership share of the claim."),
        ("DAB", "Dispute Adjudication Board — first-instance dispute resolution."),
        ("S.34", "Section 34 of Indian Arbitration Act — application to set aside award."),
        ("S.37", "Section 37 — appeal against S.34 decision."),
        ("SLP", "Special Leave Petition — Supreme Court discretionary review."),
        ("SIAC", "Singapore International Arbitration Centre."),
        ("HC / COA", "High Court / Court of Appeal (SIAC jurisdiction)."),
        ("TRUE_WIN", "Final collectible outcome — TATA ultimately collects quantum."),
        ("RESTART", "Award set aside → re-arbitration with independent fresh tribunal."),
        ("LOSE", "Final loss — TATA collects nothing."),
        ("Upfront %", "Upfront investment as fraction of SOC (or E[Q] for EQ pricing)."),
        ("Tata Tail %", "Percentage of award received paid to TATA as tail payment. Fund keeps (1 − Tata Tail %)."),
        ("Strong Buy", "E[MOIC] > 2.5× AND P(Loss) < 10%."),
        ("Attractive", "E[MOIC] > 1.5× AND P(Loss) < 25%."),
        ("Marginal", "E[MOIC] > 1.0× AND P(Loss) < 40%."),
        ("Avoid", "Does not meet Marginal thresholds."),
    ]
    row = _header_row(ws, row, ["Term", "Definition"])
    for term, defn in terms:
        row = _data_row(ws, row, [term, defn],
                        fonts=[BOLD_FONT, NORMAL_FONT])
    row += 1

    # Methodology
    row = _section_header(ws, row, "Methodology Notes")
    notes = [
        "Monte Carlo simulation draws N independent paths per claim.",
        "Each path traverses: Timeline → Arbitration → Challenge Tree → Re-Arb → Payment.",
        "Quantum is drawn from 5 discrete bands conditional on arbitration WIN.",
        "Legal costs are built from per-claim per-stage annual burn rates with stochastic overrun.",
        "Investment cashflow: upfront (month 0) + legal costs (monthly) → award receipt (terminal).",
        "XIRR computed using scipy.optimize.brentq with monthly bounds.",
        "Probability trees implement full 4-level domestic (24 paths) and 2-level SIAC (8 paths) traversal.",
        "Re-arbitration modelled for RESTART outcomes with independent P(win) = 70%.",
    ]
    for note in notes:
        ws.cell(row=row, column=2, value=f"• {note}").font = NORMAL_FONT
        row += 1
    row += 1

    # Caveats
    row = _section_header(ws, row, "Important Caveats")
    caveats = [
        "TPL shares are set to 100% — actual JV shares (50% / 33.33%) are UNCONFIRMED.",
        "Quantum bands and arbitration win probability are expert-judgment calibrated.",
        "Legal cost overrun distribution assumes mean +10% — may vary by claim complexity.",
        "IRR is undefined for paths with no positive cashflows (set to -100%).",
        "Collection efficiency is 100% of awarded quantum (no default modelled in v2).",
    ]
    for caveat in caveats:
        ws.cell(row=row, column=2, value=f"⚠ {caveat}").font = WARNING_FONT
        row += 1

    _auto_width(ws)


# ===================================================================
# Main Entry Point
# ===================================================================

def generate_excel_report(
    sim: SimulationResults,
    claims: list[ClaimConfig],
    grid: InvestmentGridResults,
    basis: str = "soc",
    output_dir: Optional[str] = None,
    filename: Optional[str] = None,
    ctx=None,
) -> str:
    """Generate the complete 14-sheet Excel workbook.

    Parameters
    ----------
    sim : SimulationResults
    claims : list[ClaimConfig]
    grid : InvestmentGridResults
    basis : str — pricing basis for portfolio analysis
    output_dir : str — output directory (default: MI.REPORT_OUTPUT_DIR)
    filename : str — output filename (default: MI.EXCEL_OUTPUT_NAME)
    ctx : PortfolioContext, optional — if provided, uses ctx.output_dir

    Returns
    -------
    str — full path to generated Excel file.
    """
    out_dir = output_dir or (ctx.output_dir if ctx else MI.REPORT_OUTPUT_DIR)
    fname = filename or MI.EXCEL_OUTPUT_NAME
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, fname)

    wb = openpyxl.Workbook()

    print("  Generating Excel report...")

    # Sheet 1: Executive Summary
    _build_executive_summary(wb, sim, claims, grid, basis)
    print("    [1/14] Executive Summary")

    # Sheet 2: Model Assumptions
    _build_model_assumptions(wb, claims)
    print("    [2/14] Model Assumptions")

    # Sheet 3: Probability Analysis
    _build_probability_analysis(wb, sim)
    print("    [3/14] Probability Analysis")

    # Sheets 4–9: Per-claim
    claim_map = {c.claim_id: c for c in claims}
    for i, cid in enumerate(sim.claim_ids):
        _build_per_claim_sheet(wb, cid, sim, claim_map[cid], grid, basis)
        print(f"    [{i+4}/14] {cid}")

    # Sheet 10: Portfolio Analysis
    _build_portfolio_analysis(wb, grid, basis)
    print("    [10/14] Portfolio Analysis")

    # Sheet 11: Breakeven Analysis
    _build_breakeven_analysis(wb, grid, basis)
    print("    [11/14] Breakeven Analysis")

    # Sheet 12: Scenario Comparison
    _build_scenario_comparison(wb, grid, basis)
    print("    [12/14] Scenario Comparison")

    # Sheet 13: Legal Costs
    _build_legal_costs(wb, sim, claims)
    print("    [13/14] Legal Costs")

    # Sheet 14: Glossary & Notes
    _build_glossary(wb)
    print("    [14/14] Glossary & Notes")

    wb.save(out_path)
    print(f"  Excel report saved: {out_path}")
    return out_path
