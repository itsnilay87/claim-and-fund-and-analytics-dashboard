"""
TATA_code_v2/v2_comparison_excel.py — Cross-Portfolio Comparison Workbook
==========================================================================

Generates a comparison workbook when --portfolio compare is used.
Takes results from all three portfolio modes (all, siac, domestic)
and produces a 7-sheet Excel workbook with side-by-side analysis.

SHEETS:
  1. Side-by-Side Summary — key metrics for all 3 portfolios
  2. MOIC Comparison      — E[MOIC] grid for each portfolio mode
  3. Risk Comparison      — P(loss), VaR, CVaR comparison
  4. Claim Attribution    — which claims drive value in each portfolio
  5. Optimal Structures   — best investment structure per portfolio
  6. Jurisdiction Impact  — domestic vs SIAC performance
  7. Notes & Methodology  — definitions, caveats

All monetary values in ₹ Crore.
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Optional

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import v2_master_inputs as MI


# ===================================================================
# Style Constants (matching v2_excel_writer.py)
# ===================================================================

TITLE_FONT    = Font(name="Arial", bold=True, size=14, color="1F4E79")
SECTION_FONT  = Font(name="Arial", bold=True, size=12, color="2E75B6")
SUBSECTION_FONT = Font(name="Arial", bold=True, size=10, color="404040")
HEADER_FONT   = Font(name="Arial", bold=True, size=9, color="FFFFFF")
NORMAL_FONT   = Font(name="Arial", size=9)
BLUE_FONT     = Font(name="Arial", size=9, color="0000FF")
BOLD_FONT     = Font(name="Arial", bold=True, size=9)
SMALL_FONT    = Font(name="Arial", size=8, color="808080")
WARNING_FONT  = Font(name="Arial", size=9, color="FF4444")

HEADER_FILL   = PatternFill("solid", fgColor="2E75B6")
BLUE_FILL     = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL    = PatternFill("solid", fgColor="E2EFDA")
YELLOW_FILL   = PatternFill("solid", fgColor="FFF2CC")
RED_FILL      = PatternFill("solid", fgColor="FCE4EC")
LIGHT_GREY    = PatternFill("solid", fgColor="F5F5F5")

# Theme colors per portfolio
PORTFOLIO_FILLS = {
    "all":      PatternFill("solid", fgColor="2E75B6"),   # Primary blue
    "siac":     PatternFill("solid", fgColor="7030A0"),   # Purple
    "domestic": PatternFill("solid", fgColor="548235"),   # Green
}
PORTFOLIO_FONTS = {
    "all":      Font(name="Arial", bold=True, size=9, color="FFFFFF"),
    "siac":     Font(name="Arial", bold=True, size=9, color="FFFFFF"),
    "domestic": Font(name="Arial", bold=True, size=9, color="FFFFFF"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Number formats
PCT_FMT = "0.0%"
CR_FMT = "#,##0.00"
MOIC_FMT = "0.00x"
INT_FMT = "#,##0"


# ===================================================================
# Helpers
# ===================================================================

def _setup_sheet(ws, title: str) -> int:
    """Standard sheet setup: col A = 3-wide spacer, title in B1."""
    ws.column_dimensions["A"].width = 3
    ws.sheet_properties.tabColor = "2E75B6"
    cell = ws.cell(row=1, column=2, value=title)
    cell.font = TITLE_FONT
    return 3  # next row to write


def _section_header(ws, row: int, text: str) -> int:
    """Write bold section header in column B."""
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = SECTION_FONT
    return row + 1


def _header_row(ws, row: int, headers: list[str]) -> int:
    """Write header row with styled cells."""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=2 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    return row + 1


def _data_row(ws, row: int, values: list, fmts=None, fills=None, fonts=None) -> int:
    """Write a data row with optional formats and fills."""
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=2 + i, value=v)
        cell.font = (fonts[i] if fonts and i < len(fonts) and fonts[i] else NORMAL_FONT)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if fmts and i < len(fmts) and fmts[i]:
            cell.number_format = fmts[i]
        if fills and i < len(fills) and fills[i]:
            cell.fill = fills[i]
    return row + 1


def _auto_width(ws, min_width=8, max_width=30):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        widths = []
        for cell in col_cells:
            if cell.value is not None:
                widths.append(len(str(cell.value)))
        if widths:
            adjusted = min(max(max(widths) + 2, min_width), max_width)
            ws.column_dimensions[get_column_letter(cell.column)].width = adjusted


# ===================================================================
# Main Generator
# ===================================================================

def generate_comparison_report(
    all_results: dict,
    output_dir: str,
    filename: str = "Portfolio_Comparison.xlsx",
) -> str:
    """Generate the cross-portfolio comparison workbook.

    Parameters
    ----------
    all_results : dict
        Keys: "all", "siac", "domestic"
        Values: dict with "sim", "grid", "claims", "context"
    output_dir : str
        Output directory.
    filename : str
        Output filename.

    Returns
    -------
    str: path to generated file.
    """
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()

    print("  Generating Portfolio Comparison Report...")

    # Sheet 1: Side-by-Side Summary
    _build_summary(wb, all_results)
    print("    [1/7] Side-by-Side Summary")

    # Sheet 2: MOIC Comparison
    _build_moic_comparison(wb, all_results)
    print("    [2/7] MOIC Comparison")

    # Sheet 3: Risk Comparison
    _build_risk_comparison(wb, all_results)
    print("    [3/7] Risk Comparison")

    # Sheet 4: Claim Attribution
    _build_claim_attribution(wb, all_results)
    print("    [4/7] Claim Attribution")

    # Sheet 5: Optimal Structures
    _build_optimal_structures(wb, all_results)
    print("    [5/7] Optimal Structures")

    # Sheet 6: Jurisdiction Impact
    _build_jurisdiction_impact(wb, all_results)
    print("    [6/7] Jurisdiction Impact")

    # Sheet 7: Notes & Methodology
    _build_notes(wb)
    print("    [7/7] Notes & Methodology")

    wb.save(out_path)
    print(f"  Comparison report saved: {out_path}")
    return out_path


# ===================================================================
# Sheet 1: Side-by-Side Summary
# ===================================================================

def _build_summary(wb: openpyxl.Workbook, all_results: dict) -> None:
    """Side-by-side key metrics for 3 portfolios."""
    ws = wb.active
    ws.title = "Summary"
    row = _setup_sheet(ws, "Cross-Portfolio Comparison Summary")

    # Date
    ws.cell(row=2, column=2, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = SMALL_FONT
    row = 4

    # Portfolio overview table
    row = _section_header(ws, row, "Portfolio Overview")
    row = _header_row(ws, row, ["Metric", "Full (6 Claims)", "SIAC (3 Claims)", "Domestic (3 Claims)"])

    modes = ["all", "siac", "domestic"]
    mode_labels = ["Full (6 Claims)", "SIAC (3 Claims)", "Domestic (3 Claims)"]

    # Number of claims
    vals = ["# Claims"]
    for m in modes:
        ctx = all_results[m]["context"]
        vals.append(ctx.n_claims)
    row = _data_row(ws, row, vals, fmts=[None, INT_FMT, INT_FMT, INT_FMT])

    # Total SOC
    vals = ["Total SOC (₹ Cr)"]
    for m in modes:
        ctx = all_results[m]["context"]
        vals.append(ctx.portfolio_soc_cr)
    row = _data_row(ws, row, vals, fmts=[None, CR_FMT, CR_FMT, CR_FMT])

    # Claim IDs
    vals = ["Claims"]
    for m in modes:
        ctx = all_results[m]["context"]
        vals.append(", ".join(ctx.claim_ids))
    row = _data_row(ws, row, vals)

    # Jurisdiction mix
    vals = ["Jurisdiction"]
    for m in modes:
        ctx = all_results[m]["context"]
        mix_str = ", ".join(f"{k}: {v}" for k, v in ctx.jurisdiction_mix.items())
        vals.append(mix_str)
    row = _data_row(ws, row, vals)

    row += 1

    # Key metrics at reference scenario (10% upfront, 30% tail)
    row = _section_header(ws, row, "Key Metrics at Reference Scenario (10% Upfront / 30% Tata Tail)")

    ref_up = 0.10
    ref_aw = 0.70  # 30% tail = 70% award share
    basis = "soc"

    row = _header_row(ws, row, ["Metric", "Full Portfolio", "SIAC Portfolio", "Domestic Portfolio"])

    metric_rows = [
        ("E[MOIC]", "mean_moic", MOIC_FMT),
        ("Median MOIC", "median_moic", MOIC_FMT),
        ("E[XIRR]", "mean_xirr", PCT_FMT),
        ("Median XIRR", "median_xirr", PCT_FMT),
        ("P(Loss)", "p_loss", PCT_FMT),
        ("P(IRR > 30%)", "p_irr_gt_30", PCT_FMT),
        ("P(IRR > 25%)", "p_irr_gt_25", PCT_FMT),
        ("E[Net Return] (₹ Cr)", "mean_net_return_cr", CR_FMT),
        ("VaR 1% (₹ Cr)", "var_1", CR_FMT),
        ("CVaR 1% (₹ Cr)", "cvar_1", CR_FMT),
    ]

    for label, attr, fmt in metric_rows:
        vals = [label]
        for m in modes:
            grid = all_results[m]["grid"]
            if grid is None:
                vals.append("N/A")
                continue
            key = (ref_up, ref_aw, basis)
            cell_data = grid.cells.get(key)
            if cell_data is None:
                # Try finding closest available
                vals.append("N/A")
            else:
                vals.append(getattr(cell_data, attr, 0))
        fmts_row = [None, fmt, fmt, fmt]
        # Color-code: green if positive, red if negative
        fills_row = [None, None, None, None]
        row = _data_row(ws, row, vals, fmts=fmts_row, fills=fills_row)

    row += 1

    # Winner summary
    row = _section_header(ws, row, "Investment Verdict")

    for m_idx, m in enumerate(modes):
        grid = all_results[m]["grid"]
        if grid is None:
            continue
        key = (ref_up, ref_aw, basis)
        cell_data = grid.cells.get(key)
        if cell_data is None:
            continue
        moic = cell_data.mean_moic
        p_loss = cell_data.p_loss
        verdict = (
            "Strong Buy" if moic > 2.5 and p_loss < 0.10 else
            "Attractive" if moic > 1.5 and p_loss < 0.25 else
            "Marginal" if moic > 1.0 and p_loss < 0.40 else
            "Avoid"
        )
        fill = GREEN_FILL if verdict in ("Strong Buy", "Attractive") else (
            YELLOW_FILL if verdict == "Marginal" else RED_FILL
        )
        row = _data_row(
            ws, row,
            [mode_labels[m_idx], f"E[MOIC]={moic:.2f}x", f"E[XIRR]={cell_data.mean_xirr:.1%}",
             f"P(loss)={p_loss:.1%}", verdict],
            fills=[None, None, None, None, fill],
            fonts=[BOLD_FONT, NORMAL_FONT, NORMAL_FONT, NORMAL_FONT, BOLD_FONT],
        )

    _auto_width(ws)


# ===================================================================
# Sheet 2: MOIC Comparison
# ===================================================================

def _build_moic_comparison(wb: openpyxl.Workbook, all_results: dict) -> None:
    """MOIC grids for each portfolio mode."""
    ws = wb.create_sheet("MOIC Comparison")
    row = _setup_sheet(ws, "E[MOIC] Comparison Across Portfolios")

    basis = "soc"
    modes = ["all", "siac", "domestic"]
    mode_labels = ["Full Portfolio (6 Claims)", "SIAC Portfolio (3 Claims)", "Domestic Portfolio (3 Claims)"]

    for m_idx, m in enumerate(modes):
        grid = all_results[m]["grid"]
        if grid is None:
            continue

        row = _section_header(ws, row, mode_labels[m_idx])

        # Header: Tata Tail %
        header = ["Upfront ↓ / Tail →"]
        for aw in grid.award_share_pcts:
            header.append(f"{(1.0 - aw):.0%}")
        row = _header_row(ws, row, header)

        # Body
        for up in grid.upfront_pcts:
            vals = [f"{up:.1%}"]
            fmts = [None]
            fills_row = [None]
            for aw in grid.award_share_pcts:
                key = (up, aw, basis)
                cell_data = grid.cells.get(key)
                if cell_data:
                    moic = cell_data.mean_moic
                    vals.append(moic)
                    fmts.append("0.00")
                    if moic >= 2.0:
                        fills_row.append(GREEN_FILL)
                    elif moic >= 1.0:
                        fills_row.append(YELLOW_FILL)
                    else:
                        fills_row.append(RED_FILL)
                else:
                    vals.append("N/A")
                    fmts.append(None)
                    fills_row.append(None)
            row = _data_row(ws, row, vals, fmts=fmts, fills=fills_row)

        row += 2  # gap between portfolios

    _auto_width(ws)


# ===================================================================
# Sheet 3: Risk Comparison
# ===================================================================

def _build_risk_comparison(wb: openpyxl.Workbook, all_results: dict) -> None:
    """P(Loss) grids for each portfolio."""
    ws = wb.create_sheet("Risk Comparison")
    row = _setup_sheet(ws, "P(Loss) Comparison Across Portfolios")

    basis = "soc"
    modes = ["all", "siac", "domestic"]
    mode_labels = ["Full Portfolio", "SIAC Portfolio", "Domestic Portfolio"]

    for m_idx, m in enumerate(modes):
        grid = all_results[m]["grid"]
        if grid is None:
            continue

        row = _section_header(ws, row, f"{mode_labels[m_idx]} — P(Loss)")

        header = ["Upfront ↓ / Tail →"]
        for aw in grid.award_share_pcts:
            header.append(f"{(1.0 - aw):.0%}")
        row = _header_row(ws, row, header)

        for up in grid.upfront_pcts:
            vals = [f"{up:.1%}"]
            fmts = [None]
            fills_row = [None]
            for aw in grid.award_share_pcts:
                key = (up, aw, basis)
                cell_data = grid.cells.get(key)
                if cell_data:
                    p_loss = cell_data.p_loss
                    vals.append(p_loss)
                    fmts.append(PCT_FMT)
                    if p_loss < 0.10:
                        fills_row.append(GREEN_FILL)
                    elif p_loss < 0.25:
                        fills_row.append(YELLOW_FILL)
                    else:
                        fills_row.append(RED_FILL)
                else:
                    vals.append("N/A")
                    fmts.append(None)
                    fills_row.append(None)
            row = _data_row(ws, row, vals, fmts=fmts, fills=fills_row)

        row += 2

    _auto_width(ws)


# ===================================================================
# Sheet 4: Claim Attribution
# ===================================================================

def _build_claim_attribution(wb: openpyxl.Workbook, all_results: dict) -> None:
    """Per-claim contribution to portfolio value."""
    ws = wb.create_sheet("Claim Attribution")
    row = _setup_sheet(ws, "Claim-Level Value Attribution")

    basis = "soc"
    ref_up = 0.10
    ref_aw = 0.70

    row = _section_header(ws, row, "Per-Claim Metrics at 10% Upfront / 30% Tail")
    row = _header_row(ws, row, [
        "Claim", "Jurisdiction", "SOC (₹ Cr)", "Portfolio",
        "E[MOIC]", "E[XIRR]", "P(Loss)", "E[Net Return]", "Economically Viable"
    ])

    modes = ["all", "siac", "domestic"]
    for m in modes:
        grid = all_results[m]["grid"]
        ctx = all_results[m]["context"]
        claims = all_results[m]["claims"]
        if grid is None:
            continue

        key = (ref_up, ref_aw, basis)
        cell_data = grid.cells.get(key)
        if cell_data is None:
            continue

        claim_map = {c.claim_id: c for c in claims}
        for cid in ctx.claim_ids:
            claim = claim_map.get(cid)
            per = cell_data.per_claim.get(cid, {})
            if not per:
                continue
            row = _data_row(ws, row, [
                cid,
                claim.jurisdiction.upper() if claim else "N/A",
                claim.soc_value_cr if claim else 0,
                ctx.label,
                per.get("E[MOIC]", 0),
                per.get("E[XIRR]", 0),
                per.get("P(loss)", 0),
                per.get("E[net_return_cr]", 0),
                "Yes" if per.get("economically_viable") else "No",
            ], fmts=[None, None, CR_FMT, None, MOIC_FMT, PCT_FMT, PCT_FMT, CR_FMT, None])

    _auto_width(ws)


# ===================================================================
# Sheet 5: Optimal Structures
# ===================================================================

def _build_optimal_structures(wb: openpyxl.Workbook, all_results: dict) -> None:
    """Best investment structure per portfolio (max E[MOIC] with P(loss)<25%)."""
    ws = wb.create_sheet("Optimal Structures")
    row = _setup_sheet(ws, "Optimal Investment Structures")

    basis = "soc"
    modes = ["all", "siac", "domestic"]
    mode_labels = ["Full (6 Claims)", "SIAC (3 Claims)", "Domestic (3 Claims)"]

    row = _section_header(ws, row, "Best Structure by E[MOIC] (P(loss) < 25%)")
    row = _header_row(ws, row, [
        "Portfolio", "Upfront %", "Tata Tail %", "E[MOIC]", "E[XIRR]",
        "P(Loss)", "P(IRR>30%)", "E[Net Return] (₹ Cr)", "Verdict"
    ])

    for m_idx, m in enumerate(modes):
        grid = all_results[m]["grid"]
        if grid is None:
            continue

        # Find best: max E[MOIC] where P(loss) < 25%
        best_key = None
        best_moic = 0.0
        for k, c in grid.cells.items():
            if k[2] != basis:
                continue
            if c.p_loss < 0.25 and c.mean_moic > best_moic:
                best_moic = c.mean_moic
                best_key = k

        if best_key is None:
            row = _data_row(ws, row, [mode_labels[m_idx], "N/A", "N/A", "N/A",
                                       "N/A", "N/A", "N/A", "N/A", "No viable scenario"])
            continue

        up, aw, _ = best_key
        c = grid.cells[best_key]
        tata_tail = 1.0 - aw
        verdict = (
            "Strong Buy" if c.mean_moic > 2.5 and c.p_loss < 0.10 else
            "Attractive" if c.mean_moic > 1.5 and c.p_loss < 0.25 else
            "Marginal" if c.mean_moic > 1.0 and c.p_loss < 0.40 else
            "Avoid"
        )
        fill = GREEN_FILL if verdict in ("Strong Buy", "Attractive") else YELLOW_FILL
        row = _data_row(ws, row, [
            mode_labels[m_idx], up, tata_tail, c.mean_moic, c.mean_xirr,
            c.p_loss, c.p_irr_gt_30, c.mean_net_return_cr, verdict
        ], fmts=[None, PCT_FMT, PCT_FMT, "0.00", PCT_FMT, PCT_FMT, PCT_FMT, CR_FMT, None],
           fills=[None, None, None, None, None, None, None, None, fill])

    row += 2

    # Also show best with P(loss) < 10% constraint
    row = _section_header(ws, row, "Conservative: Best E[MOIC] (P(loss) < 10%)")
    row = _header_row(ws, row, [
        "Portfolio", "Upfront %", "Tata Tail %", "E[MOIC]", "E[XIRR]",
        "P(Loss)", "Verdict"
    ])

    for m_idx, m in enumerate(modes):
        grid = all_results[m]["grid"]
        if grid is None:
            continue

        best_key = None
        best_moic = 0.0
        for k, c in grid.cells.items():
            if k[2] != basis:
                continue
            if c.p_loss < 0.10 and c.mean_moic > best_moic:
                best_moic = c.mean_moic
                best_key = k

        if best_key is None:
            row = _data_row(ws, row, [mode_labels[m_idx], "N/A", "N/A", "N/A",
                                       "N/A", "N/A", "No viable scenario"])
            continue

        up, aw, _ = best_key
        c = grid.cells[best_key]
        tata_tail = 1.0 - aw
        row = _data_row(ws, row, [
            mode_labels[m_idx], up, tata_tail, c.mean_moic, c.mean_xirr, c.p_loss, "Strong Buy"
        ], fmts=[None, PCT_FMT, PCT_FMT, "0.00", PCT_FMT, PCT_FMT, None],
           fills=[None, None, None, None, None, None, GREEN_FILL])

    _auto_width(ws)


# ===================================================================
# Sheet 6: Jurisdiction Impact
# ===================================================================

def _build_jurisdiction_impact(wb: openpyxl.Workbook, all_results: dict) -> None:
    """Compare SIAC vs domestic performance."""
    ws = wb.create_sheet("Jurisdiction Impact")
    row = _setup_sheet(ws, "Jurisdiction Performance Analysis")

    basis = "soc"

    # Compare SIAC vs Domestic at multiple upfront levels
    row = _section_header(ws, row, "SIAC vs Domestic — E[MOIC] at 30% Tata Tail")
    ref_aw = 0.70

    row = _header_row(ws, row, ["Upfront %", "SIAC E[MOIC]", "Domestic E[MOIC]",
                                 "SIAC E[XIRR]", "Domestic E[XIRR]",
                                 "SIAC P(Loss)", "Domestic P(Loss)", "Better Portfolio"])

    siac_grid = all_results.get("siac", {}).get("grid")
    dom_grid = all_results.get("domestic", {}).get("grid")

    if siac_grid and dom_grid:
        # Use SIAC upfront range (superset)
        upfront_set = sorted(set(siac_grid.upfront_pcts) | set(dom_grid.upfront_pcts))
        for up in upfront_set:
            siac_key = (up, ref_aw, basis)
            dom_key = (up, ref_aw, basis)
            siac_cell = siac_grid.cells.get(siac_key)
            dom_cell = dom_grid.cells.get(dom_key)

            s_moic = siac_cell.mean_moic if siac_cell else None
            d_moic = dom_cell.mean_moic if dom_cell else None
            s_xirr = siac_cell.mean_xirr if siac_cell else None
            d_xirr = dom_cell.mean_xirr if dom_cell else None
            s_ploss = siac_cell.p_loss if siac_cell else None
            d_ploss = dom_cell.p_loss if dom_cell else None

            if s_moic is not None and d_moic is not None:
                better = "SIAC" if s_moic > d_moic else "Domestic"
            elif s_moic is not None:
                better = "SIAC (only)"
            elif d_moic is not None:
                better = "Domestic (only)"
            else:
                better = "N/A"

            row = _data_row(ws, row, [
                up,
                s_moic if s_moic is not None else "N/A",
                d_moic if d_moic is not None else "N/A",
                s_xirr if s_xirr is not None else "N/A",
                d_xirr if d_xirr is not None else "N/A",
                s_ploss if s_ploss is not None else "N/A",
                d_ploss if d_ploss is not None else "N/A",
                better,
            ], fmts=[PCT_FMT, "0.00", "0.00", PCT_FMT, PCT_FMT, PCT_FMT, PCT_FMT, None])

    row += 2

    # SOC weights
    row = _section_header(ws, row, "SOC Contribution by Jurisdiction")
    siac_ctx = all_results.get("siac", {}).get("context")
    dom_ctx = all_results.get("domestic", {}).get("context")
    all_ctx = all_results.get("all", {}).get("context")

    if siac_ctx and dom_ctx and all_ctx:
        row = _header_row(ws, row, ["Metric", "SIAC", "Domestic", "Full"])
        row = _data_row(ws, row, ["SOC (₹ Cr)", siac_ctx.portfolio_soc_cr,
                                   dom_ctx.portfolio_soc_cr, all_ctx.portfolio_soc_cr],
                        fmts=[None, CR_FMT, CR_FMT, CR_FMT])
        total = all_ctx.portfolio_soc_cr
        row = _data_row(ws, row, ["% of Total SOC",
                                   siac_ctx.portfolio_soc_cr / total if total > 0 else 0,
                                   dom_ctx.portfolio_soc_cr / total if total > 0 else 0,
                                   1.0],
                        fmts=[None, PCT_FMT, PCT_FMT, PCT_FMT])
        row = _data_row(ws, row, ["# Claims", siac_ctx.n_claims, dom_ctx.n_claims,
                                   all_ctx.n_claims],
                        fmts=[None, INT_FMT, INT_FMT, INT_FMT])

    _auto_width(ws)


# ===================================================================
# Sheet 7: Notes & Methodology
# ===================================================================

def _build_notes(wb: openpyxl.Workbook) -> None:
    """Methodology notes and caveats."""
    ws = wb.create_sheet("Notes")
    row = _setup_sheet(ws, "Notes & Methodology")

    notes = [
        ("Comparison Methodology",
         "All three portfolio modes use the same MC simulation engine, "
         "random seed, and model parameters. Only the claim set and "
         "investment grid dimensions vary."),
        ("SIAC Portfolio",
         "3 claims under SIAC arbitration (Singapore). Legal challenge "
         "follows HC → COA path. No domestic S.34/S.37/SLP stages."),
        ("Domestic Portfolio",
         "3 claims under domestic Indian arbitration. Legal challenge "
         "follows S.34 → S.37 → SLP path. Longer legal timelines."),
        ("Full Portfolio",
         "All 6 claims combined. Diversification across jurisdictions "
         "may reduce portfolio-level risk."),
        ("Investment Grid",
         "SIAC grid: 5-35% upfront × 5-60% tail (156 combos). "
         "Domestic grid: 5-35% upfront × 10-60% tail (143 combos). "
         "Full portfolio grid: superset of SIAC (156 combos)."),
        ("E[MOIC]",
         "Expected Multiple on Invested Capital. MOIC = (Total Return) / "
         "(Total Invested). E[MOIC] > 1.0 means expected profit."),
        ("P(Loss)",
         "Probability that MOIC < 1.0 across all MC paths."),
        ("Caveats",
         "JV shares (50%/33.33%) are UNCONFIRMED placeholders. "
         "Day-count values come from Milestone_Delays file. "
         "Collection efficiency drawn independently of quantum state. "
         "No counterparty correlation or FX modeling."),
    ]

    for title, desc in notes:
        row = _section_header(ws, row, title)
        cell = ws.cell(row=row, column=2, value=desc)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT_WRAP
        ws.merge_cells(
            start_row=row, start_column=2,
            end_row=row, end_column=8,
        )
        row += 2

    _auto_width(ws)
