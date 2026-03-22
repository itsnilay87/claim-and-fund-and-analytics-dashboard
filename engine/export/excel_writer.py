"""
engine/export/excel_writer.py — Generate formatted Excel reports from dashboard JSON.
=====================================================================================

Sheets:
  1. Executive Summary — portfolio KPIs, structure info
  2. Investment Grid — upfront × tail heat-map style table
  3. Per-Claim Analysis — one row per claim with all metrics
  4. Risk Metrics — distribution percentiles, concentration, stress
  5. Model Assumptions — config inputs for auditability

Professional formatting: bold white-on-dark-blue headers, INR number formats,
conditional MOIC colouring, auto column widths.
"""

from __future__ import annotations

import os
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ── style constants ──────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1B2A4A") if _HAS_OPENPYXL else None
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11) if _HAS_OPENPYXL else None
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B2A4A") if _HAS_OPENPYXL else None
_KPI_FONT = Font(name="Calibri", bold=True, size=12) if _HAS_OPENPYXL else None
_BODY_FONT = Font(name="Calibri", size=11) if _HAS_OPENPYXL else None
_GREEN_FILL = PatternFill("solid", fgColor="C6EFCE") if _HAS_OPENPYXL else None
_RED_FILL = PatternFill("solid", fgColor="FFC7CE") if _HAS_OPENPYXL else None
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C") if _HAS_OPENPYXL else None
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
) if _HAS_OPENPYXL else None
_CENTER = Alignment(horizontal="center", vertical="center") if _HAS_OPENPYXL else None
_NUM_CR = '#,##0.00" Cr"'
_NUM_PCT = "0.0%"
_NUM_MOIC = "0.00x"


def _auto_width(ws, min_width: int = 10, max_width: int = 30) -> None:
    """Set column widths based on content."""
    for col_cells in ws.columns:
        length = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            length = max(length, min(len(val) + 4, max_width))
        ws.column_dimensions[col_letter].width = length


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _write_row(ws, row: int, values: list, formats: list[str | None] | None = None) -> None:
    """Write one data row with optional number formats."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _BODY_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if formats and col - 1 < len(formats) and formats[col - 1]:
            cell.number_format = formats[col - 1]


def _moic_fill(value: float) -> PatternFill | None:
    """Conditional fill for MOIC cells."""
    if value >= 1.5:
        return _GREEN_FILL
    elif value < 1.0:
        return _RED_FILL
    elif value < 1.2:
        return _YELLOW_FILL
    return None


# =====================================================================
# Sheet builders
# =====================================================================

def _build_executive_summary(wb: Workbook, data: dict) -> None:
    """Build 'Executive Summary' sheet with portfolio KPIs, waterfall, and return distribution."""
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = "1B2A4A"

    meta = data.get("simulation_meta", {})
    claims = data.get("claims", [])
    risk = data.get("risk", {})
    waterfall = data.get("waterfall", {})
    cashflow = data.get("cashflow_analysis", {})

    # Title
    ws.cell(row=1, column=1, value="Claim Analytics — Executive Summary").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {meta.get('generated_at', 'N/A')}").font = _BODY_FONT

    # Portfolio KPIs
    ws.cell(row=4, column=1, value="Portfolio Overview").font = _KPI_FONT
    kpi_labels = [
        ("Structure Type", meta.get("structure_type", "—")),
        ("Number of Claims", meta.get("n_claims", 0)),
        ("Total SOC (₹ Cr)", meta.get("total_soc_cr", 0)),
        ("Monte Carlo Paths", f"{meta.get('n_paths', 0):,}"),
        ("Seed", meta.get("seed", "")),
        ("Start Date", meta.get("start_date", "")),
        ("Discount Rate", meta.get("discount_rate", 0)),
        ("Risk-Free Rate", meta.get("risk_free_rate", 0)),
    ]
    for i, (label, value) in enumerate(kpi_labels):
        ws.cell(row=5 + i, column=1, value=label).font = _BODY_FONT
        ws.cell(row=5 + i, column=2, value=value).font = Font(name="Calibri", bold=True, size=11)

    # Waterfall summary
    nom = waterfall.get("nominal", {})
    pv = waterfall.get("present_value", {})
    row = 14
    ws.cell(row=row, column=1, value="Value Waterfall").font = _KPI_FONT
    _write_header_row(ws, row + 1, ["Metric", "Nominal (₹ Cr)", "Present Value (₹ Cr)"])
    wf_rows = [
        ("Total SOC", nom.get("soc_cr", 0), pv.get("pv_soc_cr", 0)),
        ("E[Collected]", nom.get("e_collected_cr", 0), pv.get("e_collected_cr", 0)),
        ("E[Legal Costs]", nom.get("legal_costs_cr", 0), pv.get("legal_costs_cr", 0)),
        ("Net After Legal", nom.get("net_after_legal_cr", 0), pv.get("net_after_legal_cr", 0)),
    ]
    for j, (label, v_nom, v_pv) in enumerate(wf_rows):
        _write_row(ws, row + 2 + j, [label, v_nom, v_pv], [None, _NUM_CR, _NUM_CR])

    # Risk KPIs
    moic_dist = risk.get("moic_distribution", {})
    row = 21
    ws.cell(row=row, column=1, value="Return Distribution").font = _KPI_FONT
    _write_header_row(ws, row + 1, ["Percentile", "MOIC"])
    for j, pctl in enumerate(["p5", "p25", "p50", "p75", "p95"]):
        val = moic_dist.get(pctl, 0)
        _write_row(ws, row + 2 + j, [pctl.upper(), val], [None, _NUM_MOIC])
        fill = _moic_fill(val)
        if fill:
            ws.cell(row=row + 2 + j, column=2).fill = fill

    _auto_width(ws)


def _build_investment_grid(wb: Workbook, data: dict) -> None:
    """Build 'Investment Grid' sheet with 2D MOIC and P(Loss) heat-map tables."""
    ws = wb.create_sheet("Investment Grid")
    ws.sheet_properties.tabColor = "6366F1"

    grid = data.get("investment_grid") or data.get("waterfall_grid") or {}
    if not grid:
        ws.cell(row=1, column=1, value="No investment grid data available.").font = _BODY_FONT
        return

    # Parse grid keys to build 2D matrix
    keys = sorted(grid.keys())
    # Keys are "upfront_tail" e.g. "5_10"
    upfronts = sorted(set(int(k.split("_")[0]) for k in keys if "_" in k))
    tails = sorted(set(int(k.split("_")[1]) for k in keys if "_" in k))

    if not upfronts or not tails:
        # 1D grid or unexpected format — write as flat table
        ws.cell(row=1, column=1, value="Investment Grid (Flat)").font = _TITLE_FONT
        headers = ["Key", "E[MOIC]", "Median MOIC", "E[XIRR]", "P(Loss)", "P(Hurdle)", "VaR₁", "CVaR₁"]
        _write_header_row(ws, 3, headers)
        for i, (k, c) in enumerate(sorted(grid.items())):
            row_data = [
                k,
                c.get("mean_moic", 0),
                c.get("median_moic", 0),
                c.get("mean_xirr", 0),
                c.get("p_loss", 0),
                c.get("p_hurdle", 0),
                c.get("var_1", 0),
                c.get("cvar_1", 0),
            ]
            fmts = [None, _NUM_MOIC, _NUM_MOIC, _NUM_PCT, _NUM_PCT, _NUM_PCT, _NUM_MOIC, _NUM_MOIC]
            _write_row(ws, 4 + i, row_data, fmts)
            fill = _moic_fill(c.get("mean_moic", 0))
            if fill:
                ws.cell(row=4 + i, column=2).fill = fill
        _auto_width(ws)
        return

    # 2D heat map: rows = upfront, cols = tail
    ws.cell(row=1, column=1, value="E[MOIC] — Upfront % (rows) × Tail % (cols)").font = _TITLE_FONT

    # Column headers (tail %)
    ws.cell(row=3, column=1, value="Upfront \\ Tail").font = _HEADER_FONT
    ws.cell(row=3, column=1).fill = _HEADER_FILL
    ws.cell(row=3, column=1).border = _THIN_BORDER
    for j, tail in enumerate(tails):
        cell = ws.cell(row=3, column=2 + j, value=f"{tail}%")
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER

    # Rows
    for i, up in enumerate(upfronts):
        ws.cell(row=4 + i, column=1, value=f"{up}%").font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=4 + i, column=1).border = _THIN_BORDER
        for j, tail in enumerate(tails):
            key = f"{up}_{tail}"
            cell_data = grid.get(key, {})
            moic = cell_data.get("mean_moic", 0)
            cell = ws.cell(row=4 + i, column=2 + j, value=moic)
            cell.number_format = "0.00"
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER
            cell.font = _BODY_FONT
            fill = _moic_fill(moic)
            if fill:
                cell.fill = fill

    # Second grid: P(Loss)
    offset = len(upfronts) + 6
    ws.cell(row=offset, column=1, value="P(Loss) — Upfront % (rows) × Tail % (cols)").font = _TITLE_FONT
    ws.cell(row=offset + 2, column=1, value="Upfront \\ Tail").font = _HEADER_FONT
    ws.cell(row=offset + 2, column=1).fill = _HEADER_FILL
    ws.cell(row=offset + 2, column=1).border = _THIN_BORDER
    for j, tail in enumerate(tails):
        cell = ws.cell(row=offset + 2, column=2 + j, value=f"{tail}%")
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER

    for i, up in enumerate(upfronts):
        ws.cell(row=offset + 3 + i, column=1, value=f"{up}%").font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=offset + 3 + i, column=1).border = _THIN_BORDER
        for j, tail in enumerate(tails):
            key = f"{up}_{tail}"
            cell_data = grid.get(key, {})
            p_loss = cell_data.get("p_loss", 0)
            cell = ws.cell(row=offset + 3 + i, column=2 + j, value=p_loss)
            cell.number_format = "0.0%"
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER
            cell.font = _BODY_FONT
            if p_loss >= 0.4:
                cell.fill = _RED_FILL
            elif p_loss <= 0.15:
                cell.fill = _GREEN_FILL

    _auto_width(ws)


def _build_per_claim(wb: Workbook, data: dict) -> None:
    """Build 'Per-Claim Analysis' sheet with one row per claim and all metrics."""
    ws = wb.create_sheet("Per-Claim Analysis")
    ws.sheet_properties.tabColor = "10B981"

    claims = data.get("claims", [])
    if not claims:
        ws.cell(row=1, column=1, value="No claim data available.").font = _BODY_FONT
        return

    ws.cell(row=1, column=1, value="Per-Claim Analysis").font = _TITLE_FONT

    headers = [
        "Claim ID", "Name", "Jurisdiction", "Type",
        "SOC (₹ Cr)", "Win Rate", "Eff. Win Rate",
        "E[Quantum] Cr", "E[Duration] mo",
        "E[Collected] Cr", "E[Legal] Cr",
        "E[Net] Cr",
    ]
    _write_header_row(ws, 3, headers)

    fmts = [
        None, None, None, None,
        _NUM_CR, _NUM_PCT, _NUM_PCT,
        _NUM_CR, "0.0",
        _NUM_CR, _NUM_CR,
        _NUM_CR,
    ]

    for i, c in enumerate(claims):
        net = (c.get("mean_collected_cr", 0) or 0) - (c.get("mean_legal_costs_cr", 0) or 0)
        row_data = [
            c.get("claim_id", ""),
            c.get("name", ""),
            c.get("jurisdiction", ""),
            c.get("claim_type", ""),
            c.get("soc_value_cr", 0),
            c.get("win_rate", 0),
            c.get("effective_win_rate", 0),
            c.get("mean_quantum_cr", 0),
            c.get("mean_duration_months", 0),
            c.get("mean_collected_cr", 0),
            c.get("mean_legal_costs_cr", 0),
            net,
        ]
        _write_row(ws, 4 + i, row_data, fmts)

    _auto_width(ws)


def _build_risk_metrics(wb: Workbook, data: dict) -> None:
    """Build 'Risk Metrics' sheet with MOIC/IRR percentiles, concentration, and stress tests."""
    ws = wb.create_sheet("Risk Metrics")
    ws.sheet_properties.tabColor = "EF4444"

    risk = data.get("risk", {})
    if not risk:
        ws.cell(row=1, column=1, value="No risk metrics data available.").font = _BODY_FONT
        return

    ws.cell(row=1, column=1, value="Risk Metrics").font = _TITLE_FONT

    # MOIC distribution
    row = 3
    ws.cell(row=row, column=1, value="MOIC Distribution").font = _KPI_FONT
    _write_header_row(ws, row + 1, ["Percentile", "MOIC"])
    moic_d = risk.get("moic_distribution", {})
    for j, pctl in enumerate(["p1", "p5", "p10", "p25", "p50", "p75", "p90", "p95", "p99"]):
        val = moic_d.get(pctl, 0)
        _write_row(ws, row + 2 + j, [pctl.upper(), val], [None, _NUM_MOIC])
        fill = _moic_fill(val)
        if fill:
            ws.cell(row=row + 2 + j, column=2).fill = fill

    # IRR distribution
    row += 13
    ws.cell(row=row, column=1, value="IRR Distribution").font = _KPI_FONT
    _write_header_row(ws, row + 1, ["Percentile", "IRR"])
    irr_d = risk.get("irr_distribution", {})
    for j, pctl in enumerate(["p1", "p5", "p10", "p25", "p50", "p75", "p90", "p95", "p99"]):
        _write_row(ws, row + 2 + j, [pctl.upper(), irr_d.get(pctl, 0)], [None, _NUM_PCT])

    # Concentration
    conc = risk.get("concentration", {})
    if conc:
        row += 13
        ws.cell(row=row, column=1, value="Concentration Metrics").font = _KPI_FONT
        _write_header_row(ws, row + 1, ["Metric", "Value"])
        conc_rows = [
            ("SOC Herfindahl Index", conc.get("soc_herfindahl", 0)),
            ("Max Claim % SOC", conc.get("max_pct_soc", 0)),
            ("Jurisdiction Count", conc.get("n_jurisdictions", 0)),
        ]
        for j, (label, val) in enumerate(conc_rows):
            _write_row(ws, row + 2 + j, [label, val])

    # Stress scenarios
    stress = risk.get("stress_scenarios", [])
    if stress:
        row += 7
        ws.cell(row=row, column=1, value="Stress Scenarios").font = _KPI_FONT
        _write_header_row(ws, row + 1, ["Scenario", "E[MOIC]", "P(Loss)", "Description"])
        for j, s in enumerate(stress):
            vals = [
                s.get("name", ""),
                s.get("mean_moic", 0),
                s.get("p_loss", 0),
                s.get("description", ""),
            ]
            _write_row(ws, row + 2 + j, vals, [None, _NUM_MOIC, _NUM_PCT, None])

    _auto_width(ws)


def _build_model_assumptions(wb: Workbook, data: dict) -> None:
    """Build 'Model Assumptions' sheet with simulation config, quantum bands, and claim inputs."""
    ws = wb.create_sheet("Model Assumptions")
    ws.sheet_properties.tabColor = "F59E0B"

    ws.cell(row=1, column=1, value="Model Assumptions & Configuration").font = _TITLE_FONT

    meta = data.get("simulation_meta", {})
    claims = data.get("claims", [])
    quantum = data.get("quantum_summary", {})

    # Simulation config
    row = 3
    ws.cell(row=row, column=1, value="Simulation Configuration").font = _KPI_FONT
    _write_header_row(ws, row + 1, ["Parameter", "Value"])
    params = [
        ("Monte Carlo Paths", f"{meta.get('n_paths', 0):,}"),
        ("Base Seed", meta.get("seed", "")),
        ("Start Date", meta.get("start_date", "")),
        ("Discount Rate", meta.get("discount_rate", 0)),
        ("Risk-Free Rate", meta.get("risk_free_rate", 0)),
        ("Structure Type", meta.get("structure_type", "")),
    ]
    for j, (label, val) in enumerate(params):
        _write_row(ws, row + 2 + j, [label, val])

    # Quantum bands
    bands = quantum.get("bands", [])
    if bands:
        row += len(params) + 4
        ws.cell(row=row, column=1, value="Quantum Bands").font = _KPI_FONT
        _write_header_row(ws, row + 1, ["Band", "Low", "High", "Probability", "Midpoint"])
        for j, b in enumerate(bands):
            _write_row(ws, row + 2 + j, [
                f"Band {j + 1}",
                b.get("low", 0),
                b.get("high", 0),
                b.get("probability", 0),
                b.get("midpoint", 0),
            ], [None, _NUM_PCT, _NUM_PCT, _NUM_PCT, _NUM_PCT])

    # Per-claim inputs summary
    row += len(bands) + 5 if bands else 12
    ws.cell(row=row, column=1, value="Claim Inputs").font = _KPI_FONT
    _write_header_row(ws, row + 1, ["Claim ID", "Jurisdiction", "SOC (₹ Cr)", "Win Rate"])
    for j, c in enumerate(claims):
        _write_row(ws, row + 2 + j, [
            c.get("claim_id", ""),
            c.get("jurisdiction", ""),
            c.get("soc_value_cr", 0),
            c.get("win_rate", 0),
        ], [None, None, _NUM_CR, _NUM_PCT])

    _auto_width(ws)


# =====================================================================
# Public API
# =====================================================================

def export_excel(data: dict, output_path: str) -> str:
    """Generate a formatted Excel workbook from dashboard JSON data.

    Parameters
    ----------
    data : dict
        The dashboard_data.json content (parsed).
    output_path : str
        File path for the .xlsx output.

    Returns
    -------
    str  The absolute path of the written file.

    Raises
    ------
    ImportError  If openpyxl is not installed.
    """
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    wb = Workbook()

    _build_executive_summary(wb, data)
    _build_investment_grid(wb, data)
    _build_per_claim(wb, data)
    _build_risk_metrics(wb, data)
    _build_model_assumptions(wb, data)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    size_kb = os.path.getsize(output_path) / 1024
    print(f"  Excel report exported -> {output_path} ({size_kb:.1f} KB)")
    return os.path.abspath(output_path)

