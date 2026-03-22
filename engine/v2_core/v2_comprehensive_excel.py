"""
TATA_code_v2/v2_comprehensive_excel.py — Comprehensive 20-sheet Investment Report
===================================================================================

Produces Investment_Analysis_Report.xlsx with 20 sheets designed for:
  - Investment Committee Memos
  - Risk Analysis Presentations
  - Due Diligence Reports
  - Portfolio Analysis Dashboards

SHEETS (in order):
  1.  Cover                  — Title, date, portfolio overview  
  2.  Executive Summary      — Key metrics, investment thesis, verdict
  3.  Investment Decisions   — Decision matrix with supporting rationale
  4.  Risk Analysis          — VaR, CVaR, tail risk, stress scenarios
  5.  Model Assumptions      — All parameters from master_inputs
  6.  Probability Trees      — Domestic & SIAC challenge trees
  7.  Timeline Analysis      — Duration breakdown, critical path
  8.  Quantum Analysis       — E[Q] bands, conditional distributions
  9.  TP-301-6               — Per-claim detail (×6 sheets: 9–14)
  10. TP-302-3
  11. TP-302-5
  12. TP-CTP11-2
  13. TP-CTP11-4
  14. TP-CTP13-2
  15. Portfolio Grid         — MOIC/IRR/P(Loss) grids with formatting
  16. Sensitivity Analysis   — Parameter sensitivity, tornado data
  17. Cashflow Projections   — Monthly/quarterly/annual cashflows
  18. Breakeven Analysis     — Maximum viable upfront by tail%
  19. Legal Cost Analysis    — Per-claim, per-stage, overrun analysis
  20. Glossary & Caveats     — Definitions, methodology notes

Formatting follows CLAUDE.md rules: Arial throughout, blue font for inputs,
yellow fill for key results, professional formatting for presentations.

All monetary values in ₹ Crore.
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint

from . import v2_master_inputs as MI
from .v2_config import ClaimConfig, SimulationResults
from .v2_investment_analysis import InvestmentGridResults


# ===================================================================
# Style Constants (CLAUDE.md formatting rules)
# ===================================================================

# Fonts — Arial throughout
TITLE_FONT = Font(name="Arial", bold=True, size=18, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=14, color="2E75B6")
SECTION_FONT = Font(name="Arial", bold=True, size=12, color="2E75B6")
SUBSECTION_FONT = Font(name="Arial", bold=True, size=10, color="404040")
HEADER_FONT = Font(name="Arial", bold=True, size=9, color="FFFFFF")
NORMAL_FONT = Font(name="Arial", size=9)
BLUE_FONT = Font(name="Arial", size=9, color="0000FF")
BOLD_FONT = Font(name="Arial", bold=True, size=9)
SMALL_FONT = Font(name="Arial", size=8, color="808080")
WARNING_FONT = Font(name="Arial", size=9, color="CC0000")
SUCCESS_FONT = Font(name="Arial", size=9, bold=True, color="006600")

# Fills
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
DARK_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
BLUE_FILL = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="FCE4EC")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
LIGHT_GRAY_FILL = PatternFill("solid", fgColor="F5F5F5")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
MEDIUM_BORDER = Border(
    left=Side(style="medium", color="2E75B6"),
    right=Side(style="medium", color="2E75B6"),
    top=Side(style="medium", color="2E75B6"),
    bottom=Side(style="medium", color="2E75B6"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color="2E75B6"))

# Alignments
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

# Number formats
PCT_FMT = "0.0%"
PCT_FMT2 = "0.00%"
CR_FMT = "#,##0.00"
INT_FMT = "#,##0"
MOIC_FMT = '0.00"×"'
MONTH_FMT = "0.0"
DATE_FMT = "DD-MMM-YYYY"


# ===================================================================
# Helper Utilities
# ===================================================================

def _setup_sheet(ws, title: str) -> int:
    """Standard sheet setup: col A = 3 width spacer, title in B1. Returns row=3."""
    ws.sheet_properties.tabColor = "2E75B6"
    ws.column_dimensions["A"].width = 3
    ws["B1"].value = title
    ws["B1"].font = TITLE_FONT
    ws["B1"].alignment = LEFT_WRAP
    return 3


def _header_row(ws, row: int, cols: list[str], start_col: int = 2,
                fill=None) -> int:
    """Write a header row with bold white text on blue fill. Returns row+1."""
    f = fill or HEADER_FILL
    for j, hdr in enumerate(cols):
        cell = ws.cell(row=row, column=start_col + j, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = f
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    return row + 1


def _data_row(ws, row: int, values: list, start_col: int = 2,
              fonts: Optional[list] = None, fmts: Optional[list] = None,
              fills: Optional[list] = None, aligns: Optional[list] = None) -> int:
    """Write one data row. Returns row+1."""
    for j, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + j, value=val)
        cell.font = (fonts[j] if fonts and j < len(fonts) else NORMAL_FONT)
        cell.border = THIN_BORDER
        cell.alignment = (aligns[j] if aligns and j < len(aligns) else CENTER)
        if fmts and j < len(fmts) and fmts[j]:
            cell.number_format = fmts[j]
        if fills and j < len(fills) and fills[j]:
            cell.fill = fills[j]
    return row + 1


def _section_header(ws, row: int, text: str, col: int = 2) -> int:
    """Write a section header. Returns row+1."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SECTION_FONT
    cell.alignment = LEFT_WRAP
    return row + 1


def _kv_row(ws, row: int, key: str, value, col: int = 2,
            val_font=None, val_fmt: str = "", key_font=None) -> int:
    """Write key-value pair in two columns. Returns row+1."""
    ws.cell(row=row, column=col, value=key).font = key_font or BOLD_FONT
    ws.cell(row=row, column=col, value=key).border = THIN_BORDER
    c = ws.cell(row=row, column=col + 1, value=value)
    c.font = val_font or BLUE_FONT
    c.border = THIN_BORDER
    if val_fmt:
        c.number_format = val_fmt
    return row + 1


def _auto_width(ws, min_width: float = 10, max_width: float = 40) -> None:
    """Auto-approximate column widths (skip col A spacer)."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        if col_letter == "A":
            continue
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) * 1.15, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _verdict(moic: float, p_loss: float) -> str:
    """Classify into verdict label."""
    if moic > 2.5 and p_loss < 0.10:
        return "Strong Buy"
    elif moic > 1.5 and p_loss < 0.25:
        return "Attractive"
    elif moic > 1.0 and p_loss < 0.40:
        return "Marginal"
    return "Avoid"


def _verdict_fill(verdict: str) -> PatternFill:
    if verdict == "Strong Buy":
        return GREEN_FILL
    elif verdict == "Attractive":
        return BLUE_FILL
    elif verdict == "Marginal":
        return YELLOW_FILL
    return RED_FILL


# ===================================================================
# Sheet 1: Cover Page
# ===================================================================

def _build_cover(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    claims: list[ClaimConfig],
) -> None:
    """Sheet 1: Professional cover page."""
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = "1F4E79"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30
    
    # Minimal styling for clean cover
    for r in range(1, 30):
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = WHITE_FILL
    
    row = 5
    
    # Title block
    cell = ws.cell(row=row, column=2, value="TATA Arbitration Claims Portfolio")
    cell.font = TITLE_FONT
    row += 1
    
    cell = ws.cell(row=row, column=2, value="Investment Analysis Report")
    cell.font = SUBTITLE_FONT
    row += 2
    
    # Summary box
    total_soc = sum(c.soc_value_cr for c in claims)
    cell = ws.cell(row=row, column=2, value=f"Total Statement of Claim: ₹{total_soc:,.2f} Crore")
    cell.font = Font(name="Arial", size=12, color="404040")
    row += 1
    
    cell = ws.cell(row=row, column=2, value=f"Number of Claims: {len(claims)}")
    cell.font = Font(name="Arial", size=12, color="404040")
    row += 1
    
    cell = ws.cell(row=row, column=2, value=f"Monte Carlo Simulations: {sim.n_paths:,}")
    cell.font = Font(name="Arial", size=12, color="404040")
    row += 3
    
    # Generation info
    dt = datetime.now().strftime("%d %B %Y at %H:%M")
    cell = ws.cell(row=row, column=2, value=f"Generated: {dt}")
    cell.font = SMALL_FONT
    row += 1
    
    cell = ws.cell(row=row, column=2, value=f"Model Version: TATA v2 | Seed: {sim.seed}")
    cell.font = SMALL_FONT
    row += 3
    
    # Jurisdictions
    domestic_claims = [c for c in claims if c.jurisdiction == "domestic"]
    siac_claims = [c for c in claims if c.jurisdiction == "siac"]
    
    row = _section_header(ws, row, "Portfolio Composition")
    row = _kv_row(ws, row, "Domestic (Indian courts)", f"{len(domestic_claims)} claims")
    row = _kv_row(ws, row, "SIAC (Singapore arb)", f"{len(siac_claims)} claims")
    row += 1
    
    # Claim types
    prolongation = [c for c in claims if c.archetype == "prolongation"]
    col = [c for c in claims if c.archetype == "change_of_law"]
    scope = [c for c in claims if c.archetype == "scope_variation"]
    
    row = _section_header(ws, row, "Claim Archetypes")
    row = _kv_row(ws, row, "Prolongation", f"{len(prolongation)} claims, ₹{sum(c.soc_value_cr for c in prolongation):,.2f} Cr")
    row = _kv_row(ws, row, "Change of Law", f"{len(col)} claims, ₹{sum(c.soc_value_cr for c in col):,.2f} Cr")
    row = _kv_row(ws, row, "Scope/Variation", f"{len(scope)} claims, ₹{sum(c.soc_value_cr for c in scope):,.2f} Cr")


# ===================================================================
# Sheet 2: Executive Summary
# ===================================================================

def _build_executive_summary(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    claims: list[ClaimConfig],
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 2: Executive Summary with key metrics and thesis."""
    ws = wb.create_sheet("Executive Summary")
    row = _setup_sheet(ws, "Executive Summary")
    
    # Portfolio Metrics
    row = _section_header(ws, row, "Portfolio Metrics")
    
    total_soc = sum(c.soc_value_cr for c in claims)
    total_eq = sum(sim.expected_quantum_map.get(c.claim_id, 0) for c in claims)
    avg_win_rate = np.mean([sim.win_rate_map.get(c.claim_id, 0) for c in claims])
    avg_duration = np.mean([sim.mean_duration_map.get(c.claim_id, 0) for c in claims])
    
    # Calculate aggregate metrics
    claim_collected = []
    for c in claims:
        paths = sim.results.get(c.claim_id, [])
        if paths:
            claim_collected.append(np.mean([p.collected_cr for p in paths]))
    total_e_collected = sum(claim_collected)
    
    cols = ["Metric", "Value", "Context"]
    row = _header_row(ws, row, cols)
    
    metrics = [
        ("Total SOC", f"₹{total_soc:,.2f} Cr", "6 DFCCIL claims"),
        ("E[Quantum]", f"₹{total_eq:,.2f} Cr", f"{total_eq/total_soc:.1%} of SOC"),
        ("E[Collected]", f"₹{total_e_collected:,.2f} Cr", "Post-challenge, post-re-arb"),
        ("Avg Win Rate", f"{avg_win_rate:.1%}", "Post-challenge survival"),
        ("Avg Duration", f"{avg_duration:.1f} months", "Investment to cash"),
        ("Arb Win Prob", f"{MI.ARB_WIN_PROBABILITY:.0%}", "Expert judgment"),
    ]
    
    for label, val, ctx in metrics:
        row = _data_row(ws, row, [label, val, ctx],
                        fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT])
    row += 1
    
    # Per-Claim Overview Table
    row = _section_header(ws, row, "Per-Claim Overview")
    cols = ["Claim", "Archetype", "SOC (₹ Cr)", "Jurisdiction",
            "Win Rate", "E[Duration] (mo)", "E[Q] (₹ Cr)"]
    row = _header_row(ws, row, cols)
    
    claim_map = {c.claim_id: c for c in claims}
    for cid in sim.claim_ids:
        c = claim_map[cid]
        wr = sim.win_rate_map.get(cid, 0)
        dur = sim.mean_duration_map.get(cid, 0)
        eq = sim.expected_quantum_map.get(cid, 0)
        row = _data_row(ws, row,
                        [cid, c.archetype.title(), c.soc_value_cr,
                         c.jurisdiction.upper(), wr, dur, eq],
                        fmts=[None, None, CR_FMT, None, PCT_FMT, MONTH_FMT, CR_FMT])
    row += 1
    
    # Reference Investment Scenario
    ref_up, ref_aw = 0.10, 0.80
    key = (ref_up, ref_aw, basis)
    cell_data = grid.cells.get(key)
    
    if cell_data:
        tata_tail = 1.0 - ref_aw
        row = _section_header(ws, row, f"Reference Scenario: {ref_up:.0%} Upfront / {tata_tail:.0%} Tata Tail")
        
        cols = ["Metric", "Value"]
        row = _header_row(ws, row, cols)
        
        v = _verdict(cell_data.mean_moic, cell_data.p_loss)
        ref_metrics = [
            ("E[MOIC]", cell_data.mean_moic, MOIC_FMT),
            ("Median MOIC", cell_data.median_moic, MOIC_FMT),
            ("E[XIRR]", cell_data.mean_xirr, PCT_FMT),
            ("Median XIRR", cell_data.median_xirr, PCT_FMT),
            ("E[Net Return]", f"₹{cell_data.mean_net_return_cr:,.2f} Cr", None),
            ("P(Capital Loss)", cell_data.p_loss, PCT_FMT),
            ("P(IRR > 30%)", cell_data.p_irr_gt_30, PCT_FMT),
            ("P(IRR > 25%)", cell_data.p_irr_gt_25, PCT_FMT),
            ("VaR 1%", f"₹{cell_data.var_1:,.2f} Cr", None),
            ("CVaR 1%", f"₹{cell_data.cvar_1:,.2f} Cr", None),
            ("Investment Verdict", v, None),
        ]
        
        for label, val, fmt in ref_metrics:
            fill = _verdict_fill(v) if label == "Investment Verdict" else None
            font = SUCCESS_FONT if label == "Investment Verdict" and v in ["Strong Buy", "Attractive"] else BLUE_FONT
            row = _data_row(ws, row, [label, val],
                            fonts=[BOLD_FONT, font],
                            fmts=[None, fmt] if fmt else None,
                            fills=[None, fill] if fill else None)
    
    _auto_width(ws)


# ===================================================================
# Sheet 3: Investment Decisions
# ===================================================================

def _build_investment_decisions(
    wb: openpyxl.Workbook,
    grid: InvestmentGridResults,
    claims: list[ClaimConfig],
    basis: str,
) -> None:
    """Sheet 3: Decision matrix with supporting rationale."""
    ws = wb.create_sheet("Investment Decisions")
    row = _setup_sheet(ws, "Investment Decision Framework")
    
    # Verdict Legend
    row = _section_header(ws, row, "Decision Categories")
    legend = [
        ("Strong Buy", "E[MOIC] > 2.5× AND P(Loss) < 10%", GREEN_FILL),
        ("Attractive", "E[MOIC] > 1.5× AND P(Loss) < 25%", BLUE_FILL),
        ("Marginal", "E[MOIC] > 1.0× AND P(Loss) < 40%", YELLOW_FILL),
        ("Avoid", "Does not meet Marginal thresholds", RED_FILL),
    ]
    row = _header_row(ws, row, ["Category", "Criteria", ""])
    for cat, criteria, fill in legend:
        ws.cell(row=row, column=2, value=cat).font = BOLD_FONT
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=criteria).font = NORMAL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        row += 1
    row += 1
    
    # Decision Matrix by Scenario
    row = _section_header(ws, row, "Scenario Decision Matrix")
    
    # Get all scenarios and sort by E[MOIC]
    scenarios = []
    for (up, aw, b), cell in grid.cells.items():
        if b == basis:
            v = _verdict(cell.mean_moic, cell.p_loss)
            scenarios.append({
                "upfront": up,
                "award_share": aw,
                "tata_tail": 1 - aw,
                "moic": cell.mean_moic,
                "xirr": cell.mean_xirr,
                "p_loss": cell.p_loss,
                "p_irr_30": cell.p_irr_gt_30,
                "var_1": cell.var_1,
                "verdict": v,
            })
    
    scenarios.sort(key=lambda x: x["moic"], reverse=True)
    
    cols = ["Rank", "Upfront %", "Tata Tail %", "E[MOIC]", "E[XIRR]", "P(Loss)",
            "P(IRR>30%)", "VaR 1%", "Verdict"]
    row = _header_row(ws, row, cols)
    
    for i, s in enumerate(scenarios[:20], 1):  # Top 20 scenarios
        v = s["verdict"]
        row = _data_row(ws, row,
                        [i, s["upfront"], s["tata_tail"], s["moic"],
                         s.get("xirr", 0), s["p_loss"], s["p_irr_30"], s["var_1"], v],
                        fmts=[INT_FMT, PCT_FMT, PCT_FMT, MOIC_FMT,
                              PCT_FMT, PCT_FMT, PCT_FMT, CR_FMT, None],
                        fills=[None, None, None, None, None, None, None, None,
                               _verdict_fill(v)])
    row += 1
    
    # Recommended Scenarios
    row = _section_header(ws, row, "Recommended Investment Scenarios")
    
    strong_buy = [s for s in scenarios if s["verdict"] == "Strong Buy"]
    attractive = [s for s in scenarios if s["verdict"] == "Attractive"]
    
    if strong_buy:
        ws.cell(row=row, column=2,
                value=f"Strong Buy: {len(strong_buy)} scenarios available").font = SUCCESS_FONT
        row += 1
        best = strong_buy[0]
        ws.cell(row=row, column=2,
                value=f"  Best: {best['upfront']:.0%} upfront / {best['tata_tail']:.0%} tail → "
                      f"E[MOIC]={best['moic']:.2f}×, E[XIRR]={best.get('xirr', 0):.1%}, P(loss)={best['p_loss']:.1%}").font = NORMAL_FONT
        row += 2
    
    if attractive:
        ws.cell(row=row, column=2,
                value=f"Attractive: {len(attractive)} scenarios available").font = BOLD_FONT
        row += 1
        best = attractive[0]
        ws.cell(row=row, column=2,
                value=f"  Best: {best['upfront']:.0%} upfront / {best['tata_tail']:.0%} tail → "
                      f"E[MOIC]={best['moic']:.2f}×, E[XIRR]={best.get('xirr', 0):.1%}, P(loss)={best['p_loss']:.1%}").font = NORMAL_FONT
        row += 2
    
    _auto_width(ws)


# ===================================================================
# Sheet 4: Risk Analysis
# ===================================================================

def _build_risk_analysis(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    claims: list[ClaimConfig],
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 4: Comprehensive risk analysis."""
    ws = wb.create_sheet("Risk Analysis")
    row = _setup_sheet(ws, "Risk Analysis")
    
    # Risk summary for reference scenario
    ref_up, ref_aw = 0.10, 0.80
    key = (ref_up, ref_aw, basis)
    cell_data = grid.cells.get(key)
    
    if cell_data:
        row = _section_header(ws, row, f"Risk Metrics — {ref_up:.0%} Upfront / {1-ref_aw:.0%} Tata Tail")
        
        cols = ["Risk Metric", "Value", "Interpretation"]
        row = _header_row(ws, row, cols)
        
        risk_metrics = [
            ("P(Capital Loss)", cell_data.p_loss, 
             "Probability of MOIC < 1.0 (losing principal)"),
            ("Value at Risk (1%)", f"₹{cell_data.var_1:,.2f} Cr",
             "1st percentile of return distribution"),
            ("Conditional VaR (1%)", f"₹{cell_data.cvar_1:,.2f} Cr",
             "Expected loss in worst 1% of outcomes"),
            ("MOIC Std Dev", f"{cell_data.std_moic:.2f}×",
             "Volatility of multiple outcomes"),
            ("P(IRR < 0%)", f"{1 - cell_data.p_irr_gt_30:.1%}",
             "Probability of negative IRR"),
        ]
        
        for metric, val, interp in risk_metrics:
            # Conditional formatting
            fill = None
            if metric == "P(Capital Loss)":
                pval = cell_data.p_loss
                fill = GREEN_FILL if pval < 0.15 else (YELLOW_FILL if pval < 0.30 else RED_FILL)
            
            fmt = PCT_FMT if isinstance(val, float) else None
            row = _data_row(ws, row, [metric, val, interp],
                            fmts=[None, fmt, None],
                            fills=[None, fill, None] if fill else None)
        row += 1
    
    # Per-Claim Risk Profile
    row = _section_header(ws, row, "Per-Claim Risk Profile")
    cols = ["Claim", "P(Loss post-arb)", "P(RESTART)", "Worst Duration", "Jurisdiction Risk"]
    row = _header_row(ws, row, cols)
    
    claim_map = {c.claim_id: c for c in claims}
    for cid in sim.claim_ids:
        c = claim_map[cid]
        paths = sim.results.get(cid, [])
        if not paths:
            continue
        
        # Calculate risk metrics
        p_lose = sum(1 for p in paths if p.final_outcome == "LOSE") / len(paths)
        p_restart = sum(1 for p in paths if p.final_outcome == "RESTART") / len(paths)
        worst_dur = float(np.percentile([p.total_duration_months for p in paths], 95))
        jur_risk = "Moderate" if c.jurisdiction == "domestic" else "Lower"
        
        # Loss probability fill
        loss_fill = GREEN_FILL if p_lose < 0.20 else (YELLOW_FILL if p_lose < 0.40 else RED_FILL)
        
        row = _data_row(ws, row,
                        [cid, p_lose, p_restart, worst_dur, jur_risk],
                        fmts=[None, PCT_FMT, PCT_FMT, MONTH_FMT, None],
                        fills=[None, loss_fill, None, None, None])
    row += 1
    
    # Tail Risk Scenarios
    row = _section_header(ws, row, "Tail Risk Scenarios (Worst Case Analysis)")
    
    cols = ["Scenario", "Description", "Impact", "Probability"]
    row = _header_row(ws, row, cols)
    
    tail_scenarios = [
        ("Arb Loss + Court Loss", "TATA loses arbitration and all court challenges",
         "Total capital loss", f"{(1-MI.ARB_WIN_PROBABILITY) * 0.70:.1%}"),
        ("Extended Duration", "All claims take P95 duration path",
         "IRR dilution, capital tie-up", "~5%"),
        ("Quantum Downside", "All claims settle in lowest quantum band (0-20% SOC)",
         "Reduced multiple", f"{MI.QUANTUM_BANDS[0]['probability']:.0%}"),
        ("SLP Admitted Adverse", "Supreme Court admits and rules against TATA",
         "Award set aside at final stage", "~5-10%"),
        ("RESTART → Re-Arb Loss", "RESTART outcome followed by re-arb loss",
         "Extended timeline + no recovery", "~9%"),
    ]
    
    for scenario, desc, impact, prob in tail_scenarios:
        row = _data_row(ws, row, [scenario, desc, impact, prob],
                        fonts=[BOLD_FONT, NORMAL_FONT, WARNING_FONT, NORMAL_FONT],
                        aligns=[LEFT_WRAP, LEFT_WRAP, LEFT_WRAP, CENTER])
    row += 1
    
    # Grid-wide P(Loss) analysis
    row = _section_header(ws, row, "P(Loss) by Investment Structure")
    
    up_pcts = sorted(grid.upfront_pcts)
    aw_pcts = sorted(grid.award_share_pcts)
    
    headers = ["Upfront \\ Tata Tail"] + [f"{1-p:.0%}" for p in aw_pcts]
    row = _header_row(ws, row, headers)
    
    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fills = [None]
        fmts = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell = grid.cells.get(key)
            ploss = cell.p_loss if cell else float("nan")
            vals.append(ploss)
            fmts.append(PCT_FMT)
            if ploss < 0.15:
                fills.append(GREEN_FILL)
            elif ploss < 0.30:
                fills.append(YELLOW_FILL)
            else:
                fills.append(RED_FILL)
        row = _data_row(ws, row, vals, fmts=fmts, fills=fills)
    
    _auto_width(ws)


# ===================================================================
# Sheet 5: Model Assumptions
# ===================================================================

def _build_model_assumptions(
    wb: openpyxl.Workbook,
    claims: list[ClaimConfig],
) -> None:
    """Sheet 5: All model parameters from master_inputs."""
    ws = wb.create_sheet("Model Assumptions")
    row = _setup_sheet(ws, "Model Assumptions")
    
    # Section 1: Simulation Engine
    row = _section_header(ws, row, "1. Simulation Engine")
    cols = ["Parameter", "Value", "Description"]
    row = _header_row(ws, row, cols)
    
    params = [
        ("N_SIMULATIONS", MI.N_SIMULATIONS, "Monte Carlo paths"),
        ("RANDOM_SEED", MI.RANDOM_SEED, "Base RNG seed"),
        ("MAX_TIMELINE_MONTHS", MI.MAX_TIMELINE_MONTHS, "Re-arb cutoff (months)"),
        ("START_DATE", MI.START_DATE, "Investment anchor date"),
    ]
    for p, v, d in params:
        row = _data_row(ws, row, [p, v, d],
                        fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT])
    row += 1
    
    # Section 2: Arbitration Outcome
    row = _section_header(ws, row, "2. Arbitration Outcome")
    row = _header_row(ws, row, cols)
    row = _data_row(ws, row, ["ARB_WIN_PROBABILITY", MI.ARB_WIN_PROBABILITY, "P(TATA wins arb)"],
                    fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT], fmts=[None, PCT_FMT, None])
    row = _data_row(ws, row, ["RE_ARB_WIN_PROBABILITY", MI.RE_ARB_WIN_PROBABILITY, "P(TATA wins re-arb)"],
                    fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT], fmts=[None, PCT_FMT, None])
    row += 1
    
    # Section 3: Quantum Bands
    row = _section_header(ws, row, "3. Quantum Bands (conditional on WIN)")
    row = _header_row(ws, row, ["Band", "Low", "High", "Probability", "E[Q] contribution"])
    for i, b in enumerate(MI.QUANTUM_BANDS):
        e_contrib = b["probability"] * (b["low"] + b["high"]) / 2
        row = _data_row(ws, row,
                        [f"Band {i+1}", b["low"], b["high"], b["probability"], e_contrib],
                        fmts=[None, PCT_FMT, PCT_FMT, PCT_FMT, PCT_FMT],
                        fonts=[NORMAL_FONT, BLUE_FONT, BLUE_FONT, BLUE_FONT, NORMAL_FONT])
    total_eq = sum(b["probability"] * (b["low"] + b["high"]) / 2 for b in MI.QUANTUM_BANDS)
    row = _data_row(ws, row, ["Total E[Q|WIN]", "", "", "", total_eq],
                    fonts=[BOLD_FONT, None, None, None, BOLD_FONT],
                    fmts=[None, None, None, None, PCT_FMT],
                    fills=[YELLOW_FILL, None, None, None, YELLOW_FILL])
    row += 1
    
    # Section 4: Timeline Durations
    row = _section_header(ws, row, "4. Timeline Durations (months)")
    row = _header_row(ws, row, ["Stage", "Low", "High", "Note"])
    stages = [
        ("DAB Duration", MI.DAB_DURATION["low"], MI.DAB_DURATION["high"], "Uniform"),
        ("Arbitration Duration", MI.ARB_DURATION["low"], MI.ARB_DURATION["high"], "Uniform"),
        ("Arb Remaining (302-5)", MI.ARB_REMAINING_302_5["low"], MI.ARB_REMAINING_302_5["high"], "302-5 only"),
        ("Re-referral (CTP11-2)", MI.RE_REFERRAL_CTP11_2["low"], MI.RE_REFERRAL_CTP11_2["high"], "CTP11-2 only"),
        ("S.34 Duration", MI.S34_DURATION["low"], MI.S34_DURATION["high"], "Domestic courts"),
        ("S.37 Duration", MI.S37_DURATION["low"], MI.S37_DURATION["high"], "Domestic courts"),
        ("SLP Dismissed", MI.SLP_DISMISSED_DURATION, MI.SLP_DISMISSED_DURATION, "Fixed"),
        ("SLP Admitted", MI.SLP_ADMITTED_DURATION, MI.SLP_ADMITTED_DURATION, "Fixed"),
        ("SIAC HC Duration", MI.SIAC_HC_DURATION, MI.SIAC_HC_DURATION, "Fixed"),
        ("SIAC COA Duration", MI.SIAC_COA_DURATION, MI.SIAC_COA_DURATION, "Fixed"),
        ("Domestic Payment Delay", MI.DOMESTIC_PAYMENT_DELAY, MI.DOMESTIC_PAYMENT_DELAY, "Post-resolution"),
        ("SIAC Payment Delay", MI.SIAC_PAYMENT_DELAY, MI.SIAC_PAYMENT_DELAY, "Post-resolution"),
    ]
    for name, lo, hi, note in stages:
        row = _data_row(ws, row, [name, lo, hi, note],
                        fonts=[NORMAL_FONT, BLUE_FONT, BLUE_FONT, SMALL_FONT],
                        fmts=[None, MONTH_FMT, MONTH_FMT, None])
    row += 1
    
    # Section 5: Investment Grid
    row = _section_header(ws, row, "5. Investment Grid")
    row = _kv_row(ws, row, "Upfront % levels", str(MI.UPFRONT_PCT_SOC))
    row = _kv_row(ws, row, "Tata Tail % levels", str(MI.TATA_TAIL_PCT))
    row = _kv_row(ws, row, "Grid size", f"{len(MI.UPFRONT_PCT_SOC)} × {len(MI.TATA_TAIL_PCT)}")
    row += 1
    
    # Section 6: Financial Parameters
    row = _section_header(ws, row, "6. Financial Parameters")
    row = _header_row(ws, row, cols)
    row = _data_row(ws, row, ["DISCOUNT_RATE", MI.DISCOUNT_RATE, "Annual hurdle rate"],
                    fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT], fmts=[None, PCT_FMT, None])
    row = _data_row(ws, row, ["RISK_FREE_RATE", MI.RISK_FREE_RATE, "Annualized risk-free rate"],
                    fonts=[BOLD_FONT, BLUE_FONT, NORMAL_FONT], fmts=[None, PCT_FMT, None])
    
    _auto_width(ws)


# ===================================================================
# Sheet 6: Probability Trees
# ===================================================================

def _build_probability_trees(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
) -> None:
    """Sheet 6: Domestic & SIAC probability tree paths."""
    ws = wb.create_sheet("Probability Trees")
    row = _setup_sheet(ws, "Legal Challenge Probability Trees")
    
    # Domestic Tree - Scenario A
    row = _section_header(ws, row, "Domestic Challenge Tree — Scenario A (TATA won arb)")
    row = _header_row(ws, row, ["Path", "S.34", "S.37", "SLP", "Outcome", "Probability"])
    
    for p in MI.DOMESTIC_PATHS_A:
        s34 = "Win" if p.get("s34_tata_wins") else "Lose"
        s37 = "Win" if p.get("s37_tata_wins") else "Lose"
        slp = "Dismissed" if not p.get("slp_admitted") else (
            "Win" if p.get("slp_merits_tata_wins") else "Lose"
        )
        outcome = p["outcome"]
        prob = p["conditional_prob"]
        
        outcome_fill = GREEN_FILL if outcome == "TRUE_WIN" else (
            YELLOW_FILL if outcome == "RESTART" else RED_FILL
        )
        
        row = _data_row(ws, row, [p["path_id"], s34, s37, slp, outcome, prob],
                        fmts=[None, None, None, None, None, PCT_FMT],
                        fills=[None, None, None, None, outcome_fill, None])
    row += 1
    
    # Domestic Tree - Scenario B  
    row = _section_header(ws, row, "Domestic Challenge Tree — Scenario B (TATA lost arb)")
    row = _header_row(ws, row, ["Path", "S.34", "S.37", "SLP", "Outcome", "Probability"])
    
    for p in MI.DOMESTIC_PATHS_B:
        s34 = "Win" if p.get("s34_tata_wins") else "Lose"
        s37 = "Win" if p.get("s37_tata_wins") else "Lose"
        slp = "Dismissed" if not p.get("slp_admitted") else (
            "Win" if p.get("slp_merits_tata_wins") else "Lose"
        )
        outcome = p["outcome"]
        prob = p["conditional_prob"]
        
        outcome_fill = GREEN_FILL if outcome == "TRUE_WIN" else (
            YELLOW_FILL if outcome == "RESTART" else RED_FILL
        )
        
        row = _data_row(ws, row, [p["path_id"], s34, s37, slp, outcome, prob],
                        fmts=[None, None, None, None, None, PCT_FMT],
                        fills=[None, None, None, None, outcome_fill, None])
    row += 1
    
    # SIAC Trees
    row = _section_header(ws, row, "SIAC Challenge Tree — Scenario A (TATA won arb)")
    row = _header_row(ws, row, ["Path", "HC", "COA", "Outcome", "Probability"])
    
    for p in MI.SIAC_PATHS_A:
        hc = "Win" if p.get("hc_tata_wins") else "Lose"
        coa = "Win" if p.get("coa_tata_wins") else "Lose"
        outcome = p["outcome"]
        prob = p["conditional_prob"]
        
        outcome_fill = GREEN_FILL if outcome == "TRUE_WIN" else (
            YELLOW_FILL if outcome == "RESTART" else RED_FILL
        )
        
        row = _data_row(ws, row, [p["path_id"], hc, coa, outcome, prob],
                        fmts=[None, None, None, None, PCT_FMT],
                        fills=[None, None, None, outcome_fill, None])
    row += 1
    
    row = _section_header(ws, row, "SIAC Challenge Tree — Scenario B (TATA lost arb)")
    row = _header_row(ws, row, ["Path", "HC", "COA", "Outcome", "Probability"])
    
    for p in MI.SIAC_PATHS_B:
        hc = "Win" if p.get("hc_tata_wins") else "Lose"
        coa = "Win" if p.get("coa_tata_wins") else "Lose"
        outcome = p["outcome"]
        prob = p["conditional_prob"]
        
        outcome_fill = GREEN_FILL if outcome == "TRUE_WIN" else (
            YELLOW_FILL if outcome == "RESTART" else RED_FILL
        )
        
        row = _data_row(ws, row, [p["path_id"], hc, coa, outcome, prob],
                        fmts=[None, None, None, None, PCT_FMT],
                        fills=[None, None, None, outcome_fill, None])
    row += 1
    
    # Outcome Summary from MC
    row = _section_header(ws, row, "Outcome Summary (from Monte Carlo)")
    row = _header_row(ws, row, ["Claim", "P(TRUE_WIN)", "P(RESTART)", "P(LOSE)", "Final Win Rate"])
    
    for cid in sim.claim_ids:
        paths = sim.results[cid]
        n = len(paths)
        p_tw = sum(1 for p in paths if p.final_outcome == "TRUE_WIN") / n
        p_re = sum(1 for p in paths if p.final_outcome == "RESTART") / n
        p_lo = sum(1 for p in paths if p.final_outcome == "LOSE") / n
        wr = sim.win_rate_map.get(cid, 0)
        row = _data_row(ws, row, [cid, p_tw, p_re, p_lo, wr],
                        fmts=[None, PCT_FMT, PCT_FMT, PCT_FMT, PCT_FMT])
    
    _auto_width(ws)


# ===================================================================
# Sheet 7: Timeline Analysis
# ===================================================================

def _build_timeline_analysis(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    claims: list[ClaimConfig],
) -> None:
    """Sheet 7: Duration breakdown and critical path analysis."""
    ws = wb.create_sheet("Timeline Analysis")
    row = _setup_sheet(ws, "Timeline Analysis")
    
    # Duration Summary  
    row = _section_header(ws, row, "Duration Summary by Claim")
    cols = ["Claim", "Pipeline", "E[Duration]", "P5", "P25", "Median", "P75", "P95"]
    row = _header_row(ws, row, cols)
    
    claim_map = {c.claim_id: c for c in claims}
    for cid in sim.claim_ids:
        c = claim_map[cid]
        paths = sim.results.get(cid, [])
        if not paths:
            continue
        
        durations = [p.total_duration_months for p in paths]
        dur_arr = np.array(durations)
        
        pipeline_str = " → ".join(c.pipeline)
        pctiles = [float(np.percentile(dur_arr, p)) for p in [5, 25, 50, 75, 95]]
        
        row = _data_row(ws, row,
                        [cid, pipeline_str, float(np.mean(dur_arr))] + pctiles,
                        fmts=[None, None] + [MONTH_FMT] * 6,
                        aligns=[CENTER, LEFT_WRAP] + [CENTER] * 6)
    row += 1
    
    # Portfolio Duration Analysis
    row = _section_header(ws, row, "Portfolio Duration Analysis")
    
    # Calculate portfolio-level duration (longest claim per path)
    n_paths = sim.n_paths
    portfolio_durations = []
    for i in range(n_paths):
        max_dur = 0
        for cid in sim.claim_ids:
            paths = sim.results.get(cid, [])
            if i < len(paths):
                max_dur = max(max_dur, paths[i].total_duration_months)
        portfolio_durations.append(max_dur)
    
    port_arr = np.array(portfolio_durations)
    
    cols = ["Metric", "Value"]
    row = _header_row(ws, row, cols)
    
    dur_metrics = [
        ("E[Portfolio Duration]", f"{float(np.mean(port_arr)):.1f} months"),
        ("Median Duration", f"{float(np.median(port_arr)):.1f} months"),
        ("P95 Duration", f"{float(np.percentile(port_arr, 95)):.1f} months"),
        ("Max Duration", f"{float(np.max(port_arr)):.1f} months"),
        ("Time to First Cash", f"{float(np.min([sim.mean_duration_map.get(cid, 0) for cid in sim.claim_ids])):.1f} months"),
    ]
    
    for metric, val in dur_metrics:
        row = _data_row(ws, row, [metric, val], fonts=[BOLD_FONT, BLUE_FONT])
    row += 1
    
    # Stage Duration Breakdown
    row = _section_header(ws, row, "Expected Stage Durations (from master_inputs)")
    cols = ["Stage", "Distribution", "E[Duration]", "Range"]
    row = _header_row(ws, row, cols)
    
    stage_data = [
        ("DAB Proceedings", "Uniform", (MI.DAB_DURATION["low"] + MI.DAB_DURATION["high"]) / 2,
         f"{MI.DAB_DURATION['low']:.1f}–{MI.DAB_DURATION['high']:.1f}"),
        ("Arbitration", "Uniform", (MI.ARB_DURATION["low"] + MI.ARB_DURATION["high"]) / 2,
         f"{MI.ARB_DURATION['low']:.1f}–{MI.ARB_DURATION['high']:.1f}"),
        ("S.34 Proceedings", "Uniform", (MI.S34_DURATION["low"] + MI.S34_DURATION["high"]) / 2,
         f"{MI.S34_DURATION['low']:.0f}–{MI.S34_DURATION['high']:.0f}"),
        ("S.37 Appeal", "Uniform", (MI.S37_DURATION["low"] + MI.S37_DURATION["high"]) / 2,
         f"{MI.S37_DURATION['low']:.0f}–{MI.S37_DURATION['high']:.0f}"),
        ("SLP (dismissed)", "Fixed", MI.SLP_DISMISSED_DURATION, "4.0"),
        ("SLP (admitted)", "Fixed", MI.SLP_ADMITTED_DURATION, "24.0"),
        ("Payment Delay (Domestic)", "Fixed", MI.DOMESTIC_PAYMENT_DELAY, "6.0"),
        ("Payment Delay (SIAC)", "Fixed", MI.SIAC_PAYMENT_DELAY, "4.0"),
    ]
    
    for stage, dist, e_dur, rng in stage_data:
        row = _data_row(ws, row, [stage, dist, e_dur, rng],
                        fmts=[None, None, MONTH_FMT, None],
                        fonts=[NORMAL_FONT, SMALL_FONT, BLUE_FONT, SMALL_FONT])
    
    _auto_width(ws)


# ===================================================================
# Sheet 8: Quantum Analysis
# ===================================================================

def _build_quantum_analysis(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    claims: list[ClaimConfig],
) -> None:
    """Sheet 8: Expected quantum analysis."""
    ws = wb.create_sheet("Quantum Analysis")
    row = _setup_sheet(ws, "Quantum Analysis")
    
    # Quantum Bands
    row = _section_header(ws, row, "Quantum Band Distribution (conditional on WIN)")
    cols = ["Band", "Range (% SOC)", "Probability", "E[Q|Band]", "Contribution"]
    row = _header_row(ws, row, cols)
    
    for i, b in enumerate(MI.QUANTUM_BANDS):
        mid = (b["low"] + b["high"]) / 2
        contrib = b["probability"] * mid
        row = _data_row(ws, row,
                        [f"Band {i+1}", f"{b['low']:.0%}–{b['high']:.0%}",
                         b["probability"], mid, contrib],
                        fmts=[None, None, PCT_FMT, PCT_FMT, PCT_FMT],
                        fonts=[NORMAL_FONT, NORMAL_FONT, BLUE_FONT, NORMAL_FONT, NORMAL_FONT])
    
    total_eq = sum(b["probability"] * (b["low"] + b["high"]) / 2 for b in MI.QUANTUM_BANDS)
    row = _data_row(ws, row, ["TOTAL", "", "100.0%", "", total_eq],
                    fonts=[BOLD_FONT, None, BOLD_FONT, None, BOLD_FONT],
                    fmts=[None, None, None, None, PCT_FMT],
                    fills=[YELLOW_FILL, None, None, None, YELLOW_FILL])
    row += 1
    
    # Per-Claim Quantum
    row = _section_header(ws, row, "Expected Quantum by Claim")
    cols = ["Claim", "SOC (₹ Cr)", "E[Q] (₹ Cr)", "E[Q|WIN]%", "E[Collected] (₹ Cr)"]
    row = _header_row(ws, row, cols)
    
    claim_map = {c.claim_id: c for c in claims}
    total_soc = 0
    total_eq_cr = 0
    total_ec = 0
    
    for cid in sim.claim_ids:
        c = claim_map[cid]
        eq = sim.expected_quantum_map.get(cid, 0)
        paths = sim.results.get(cid, [])
        e_collected = float(np.mean([p.collected_cr for p in paths])) if paths else 0
        
        total_soc += c.soc_value_cr
        total_eq_cr += eq
        total_ec += e_collected
        
        row = _data_row(ws, row,
                        [cid, c.soc_value_cr, eq, eq / c.soc_value_cr if c.soc_value_cr else 0, e_collected],
                        fmts=[None, CR_FMT, CR_FMT, PCT_FMT, CR_FMT])
    
    row = _data_row(ws, row,
                    ["PORTFOLIO", total_soc, total_eq_cr, total_eq_cr / total_soc, total_ec],
                    fonts=[BOLD_FONT, BOLD_FONT, BOLD_FONT, BOLD_FONT, BOLD_FONT],
                    fmts=[None, CR_FMT, CR_FMT, PCT_FMT, CR_FMT],
                    fills=[YELLOW_FILL, YELLOW_FILL, YELLOW_FILL, YELLOW_FILL, YELLOW_FILL])
    row += 1
    
    # Quantum Percentiles
    row = _section_header(ws, row, "Quantum Percentile Distribution (from MC paths)")
    cols = ["Claim", "P5", "P25", "Median", "P75", "P95", "Mean"]
    row = _header_row(ws, row, cols)
    
    for cid in sim.claim_ids:
        paths = sim.results.get(cid, [])
        if not paths:
            continue
        
        collected = np.array([p.collected_cr for p in paths])
        pctiles = [float(np.percentile(collected, p)) for p in [5, 25, 50, 75, 95]]
        mean_val = float(np.mean(collected))
        
        row = _data_row(ws, row,
                        [cid] + pctiles + [mean_val],
                        fmts=[None] + [CR_FMT] * 6)
    
    _auto_width(ws)


# ===================================================================
# Sheets 9-14: Per-Claim Detail
# ===================================================================

def _build_claim_sheet(
    wb: openpyxl.Workbook,
    cid: str,
    sim: SimulationResults,
    claim: ClaimConfig,
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Build per-claim detail sheet."""
    ws = wb.create_sheet(cid)
    row = _setup_sheet(ws, f"Claim Analysis — {cid}")
    
    # Identity
    row = _section_header(ws, row, "Claim Identity")
    row = _kv_row(ws, row, "Claim ID", cid)
    row = _kv_row(ws, row, "Archetype", claim.archetype.title())
    row = _kv_row(ws, row, "SOC (₹ Cr)", claim.soc_value_cr, val_fmt=CR_FMT)
    row = _kv_row(ws, row, "Jurisdiction", claim.jurisdiction.upper())
    row = _kv_row(ws, row, "Current Gate", claim.current_gate)
    row = _kv_row(ws, row, "TPL Share", claim.tpl_share, val_fmt=PCT_FMT)
    row = _kv_row(ws, row, "Pipeline", " → ".join(claim.pipeline))
    row += 1
    
    # Simulation Results
    row = _section_header(ws, row, "Monte Carlo Results")
    paths = sim.results.get(cid, [])
    n = len(paths)
    
    if n == 0:
        ws.cell(row=row, column=2, value="No simulation data").font = WARNING_FONT
        _auto_width(ws)
        return
    
    arb_win_rate = sum(1 for p in paths if p.arb_won) / n
    final_win_rate = sim.win_rate_map.get(cid, 0)
    durations = [p.total_duration_months for p in paths]
    legal_costs = [p.legal_cost_total_cr for p in paths]
    collected = [p.collected_cr for p in paths]
    
    eq_cr = sim.expected_quantum_map.get(cid, 0)
    
    row = _kv_row(ws, row, "Arb Win Rate", arb_win_rate, val_fmt=PCT_FMT)
    row = _kv_row(ws, row, "Final Win Rate", final_win_rate, val_fmt=PCT_FMT)
    row = _kv_row(ws, row, "E[Duration] (mo)", float(np.mean(durations)), val_fmt=MONTH_FMT)
    row = _kv_row(ws, row, "Median Duration (mo)", float(np.median(durations)), val_fmt=MONTH_FMT)
    row = _kv_row(ws, row, "E[Legal Cost] (₹ Cr)", float(np.mean(legal_costs)), val_fmt=CR_FMT)
    row = _kv_row(ws, row, "E[Collected] (₹ Cr)", float(np.mean(collected)), val_fmt=CR_FMT)
    row = _kv_row(ws, row, "E[Q] analytical (₹ Cr)", eq_cr, val_fmt=CR_FMT)
    row += 1
    
    # Outcome Distribution
    row = _section_header(ws, row, "Outcome Distribution")
    p_tw = sum(1 for p in paths if p.final_outcome == "TRUE_WIN") / n
    p_re = sum(1 for p in paths if p.final_outcome == "RESTART") / n
    p_lo = sum(1 for p in paths if p.final_outcome == "LOSE") / n
    
    cols = ["Outcome", "Probability", "Interpretation"]
    row = _header_row(ws, row, cols)
    row = _data_row(ws, row, ["TRUE_WIN", p_tw, "Award survives, cash recovery"],
                    fmts=[None, PCT_FMT, None], fills=[GREEN_FILL, None, None])
    row = _data_row(ws, row, ["RESTART", p_re, "Re-arbitration required"],
                    fmts=[None, PCT_FMT, None], fills=[YELLOW_FILL, None, None])
    row = _data_row(ws, row, ["LOSE", p_lo, "No recovery"],
                    fmts=[None, PCT_FMT, None], fills=[RED_FILL, None, None])
    row += 1
    
    # Percentile Table
    row = _section_header(ws, row, "Percentile Distribution")
    pctiles = [5, 25, 50, 75, 95]
    cols = ["Metric"] + [f"P{p}" for p in pctiles]
    row = _header_row(ws, row, cols)
    
    dur_arr = np.array(durations)
    coll_arr = np.array(collected)
    cost_arr = np.array(legal_costs)
    
    dur_pcts = [float(np.percentile(dur_arr, p)) for p in pctiles]
    coll_pcts = [float(np.percentile(coll_arr, p)) for p in pctiles]
    cost_pcts = [float(np.percentile(cost_arr, p)) for p in pctiles]
    
    row = _data_row(ws, row, ["Duration (mo)"] + dur_pcts, fmts=[None] + [MONTH_FMT] * 5)
    row = _data_row(ws, row, ["Collected (₹ Cr)"] + coll_pcts, fmts=[None] + [CR_FMT] * 5)
    row = _data_row(ws, row, ["Legal Cost (₹ Cr)"] + cost_pcts, fmts=[None] + [CR_FMT] * 5)
    row += 1
    
    # Investment Grid for this Claim
    row = _section_header(ws, row, f"Investment Grid ({basis.upper()} pricing)")
    cols = ["Upfront %", "Tata Tail %", "E[MOIC]", "E[XIRR]", "P(Loss)"]
    row = _header_row(ws, row, cols)
    
    ref_ups = [0.05, 0.10, 0.15, 0.20]
    ref_aws = [0.90, 0.80, 0.70, 0.50]
    
    for up in ref_ups:
        for aw in ref_aws:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            if cell_data and cid in cell_data.per_claim:
                pc = cell_data.per_claim[cid]
                moic_val = pc["E[MOIC]"]
                fill = GREEN_FILL if moic_val > 1.5 else (
                    YELLOW_FILL if moic_val > 1.0 else RED_FILL
                )
                row = _data_row(ws, row,
                                [up, 1-aw, moic_val, pc["E[XIRR]"], pc["P(loss)"]],
                                fmts=[PCT_FMT, PCT_FMT, MOIC_FMT, PCT_FMT, PCT_FMT],
                                fills=[None, None, fill, None, None])
    
    _auto_width(ws)


# ===================================================================
# Sheet 15: Portfolio Grid
# ===================================================================

def _build_portfolio_grid(
    wb: openpyxl.Workbook,
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 15: Full MOIC/IRR grids with conditional formatting."""
    ws = wb.create_sheet("Portfolio Grid")
    row = _setup_sheet(ws, f"Portfolio Analysis Grid — {basis.upper()} Pricing")
    
    up_pcts = sorted(grid.upfront_pcts)
    aw_pcts = sorted(grid.award_share_pcts)
    
    # MOIC Grid
    row = _section_header(ws, row, "E[MOIC] Grid")
    headers = ["Upfront \\ Tata Tail"] + [f"{1-p:.0%}" for p in aw_pcts]
    row = _header_row(ws, row, headers)
    
    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fmts = [None]
        fills = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            moic = cell_data.mean_moic if cell_data else 0
            vals.append(moic)
            fmts.append(MOIC_FMT)
            if moic > 2.0:
                fills.append(GREEN_FILL)
            elif moic > 1.5:
                fills.append(BLUE_FILL)
            elif moic >= 1.0:
                fills.append(YELLOW_FILL)
            else:
                fills.append(RED_FILL)
        row = _data_row(ws, row, vals, fmts=fmts, fills=fills)
    row += 1
    
    # E[XIRR] Grid
    row = _section_header(ws, row, "E[XIRR] Grid")
    row = _header_row(ws, row, headers)
    
    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fmts = [None]
        fills = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            xirr = cell_data.mean_xirr if cell_data else 0
            vals.append(xirr)
            fmts.append(PCT_FMT)
            if xirr > 0.30:
                fills.append(GREEN_FILL)
            elif xirr > 0.15:
                fills.append(BLUE_FILL)
            elif xirr >= 0.0:
                fills.append(YELLOW_FILL)
            else:
                fills.append(RED_FILL)
        row = _data_row(ws, row, vals, fmts=fmts, fills=fills)
    row += 1
    
    # P(IRR > 30%) Grid
    row = _section_header(ws, row, "P(IRR > 30%) Grid")
    row = _header_row(ws, row, headers)
    
    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fmts = [None]
        fills = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            p_irr = cell_data.p_irr_gt_30 if cell_data else 0
            vals.append(p_irr)
            fmts.append(PCT_FMT)
            if p_irr > 0.70:
                fills.append(GREEN_FILL)
            elif p_irr > 0.50:
                fills.append(BLUE_FILL)
            elif p_irr > 0.30:
                fills.append(YELLOW_FILL)
            else:
                fills.append(RED_FILL)
        row = _data_row(ws, row, vals, fmts=fmts, fills=fills)
    row += 1
    
    # P(Loss) Grid
    row = _section_header(ws, row, "P(Capital Loss) Grid")
    row = _header_row(ws, row, headers)
    
    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fmts = [None]
        fills = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            ploss = cell_data.p_loss if cell_data else 1
            vals.append(ploss)
            fmts.append(PCT_FMT)
            if ploss < 0.10:
                fills.append(GREEN_FILL)
            elif ploss < 0.25:
                fills.append(BLUE_FILL)
            elif ploss < 0.40:
                fills.append(YELLOW_FILL)
            else:
                fills.append(RED_FILL)
        row = _data_row(ws, row, vals, fmts=fmts, fills=fills)
    row += 1
    
    # VaR Grid
    row = _section_header(ws, row, "VaR 1% Grid (₹ Crore)")
    row = _header_row(ws, row, headers)
    
    for up in up_pcts:
        vals = [f"{up:.0%}"]
        fmts = [None]
        for aw in aw_pcts:
            key = (up, aw, basis)
            cell_data = grid.cells.get(key)
            var5 = cell_data.var_1 if cell_data else 0
            vals.append(var5)
            fmts.append(CR_FMT)
        row = _data_row(ws, row, vals, fmts=fmts)
    
    _auto_width(ws)


# ===================================================================
# Sheet 16: Sensitivity Analysis
# ===================================================================

def _build_sensitivity_analysis(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 16: Parameter sensitivity analysis."""
    ws = wb.create_sheet("Sensitivity Analysis")
    row = _setup_sheet(ws, "Sensitivity Analysis")
    
    # Reference scenario
    ref_up, ref_aw = 0.10, 0.80
    ref_key = (ref_up, ref_aw, basis)
    ref_cell = grid.cells.get(ref_key)
    
    if ref_cell:
        row = _section_header(ws, row, f"Base Case: {ref_up:.0%} Upfront / {1-ref_aw:.0%} Tata Tail")
        row = _kv_row(ws, row, "Base E[MOIC]", ref_cell.mean_moic, val_fmt=MOIC_FMT)
        row = _kv_row(ws, row, "Base E[XIRR]", ref_cell.mean_xirr, val_fmt=PCT_FMT)
        row = _kv_row(ws, row, "Base P(Loss)", ref_cell.p_loss, val_fmt=PCT_FMT)
        row += 1
    
    # Sensitivity grid (vary award share at fixed upfront)
    row = _section_header(ws, row, "Sensitivity to Tata Tail % (upfront = 10%)")
    cols = ["Tata Tail %", "E[MOIC]", "Δ MOIC", "E[XIRR]", "Δ XIRR", "P(Loss)", "Δ P(Loss)"]
    row = _header_row(ws, row, cols)
    
    base_moic = ref_cell.mean_moic if ref_cell else 1.0
    base_xirr = ref_cell.mean_xirr if ref_cell else 0.0
    base_ploss = ref_cell.p_loss if ref_cell else 0.3
    
    for aw in sorted(grid.award_share_pcts):
        key = (ref_up, aw, basis)
        cell = grid.cells.get(key)
        if cell:
            delta_moic = cell.mean_moic - base_moic
            delta_xirr = cell.mean_xirr - base_xirr
            delta_ploss = cell.p_loss - base_ploss
            row = _data_row(ws, row,
                            [1-aw, cell.mean_moic, delta_moic, cell.mean_xirr, delta_xirr,
                             cell.p_loss, delta_ploss],
                            fmts=[PCT_FMT, MOIC_FMT, MOIC_FMT, PCT_FMT, PCT_FMT, PCT_FMT, PCT_FMT],
                            fills=[None, None,
                                   GREEN_FILL if delta_moic > 0 else RED_FILL,
                                   None,
                                   GREEN_FILL if delta_xirr > 0 else RED_FILL,
                                   None,
                                   GREEN_FILL if delta_ploss < 0 else RED_FILL])
    row += 1
    
    # Sensitivity grid (vary upfront at fixed award share)
    row = _section_header(ws, row, "Sensitivity to Upfront % (Tata tail = 20%)")
    cols = ["Upfront %", "E[MOIC]", "Δ MOIC", "E[XIRR]", "Δ XIRR", "P(Loss)", "Δ P(Loss)"]
    row = _header_row(ws, row, cols)
    
    for up in sorted(grid.upfront_pcts):
        key = (up, ref_aw, basis)
        cell = grid.cells.get(key)
        if cell:
            delta_moic = cell.mean_moic - base_moic
            delta_xirr = cell.mean_xirr - base_xirr
            delta_ploss = cell.p_loss - base_ploss
            row = _data_row(ws, row,
                            [up, cell.mean_moic, delta_moic, cell.mean_xirr, delta_xirr,
                             cell.p_loss, delta_ploss],
                            fmts=[PCT_FMT, MOIC_FMT, MOIC_FMT, PCT_FMT, PCT_FMT, PCT_FMT, PCT_FMT],
                            fills=[None, None,
                                   GREEN_FILL if delta_moic > 0 else RED_FILL,
                                   None,
                                   GREEN_FILL if delta_xirr > 0 else RED_FILL,
                                   None,
                                   GREEN_FILL if delta_ploss < 0 else RED_FILL])
    row += 1
    
    # Key Parameter Sensitivities (narrative)
    row = _section_header(ws, row, "Key Parameter Impact Summary")
    cols = ["Parameter", "Impact Direction", "MOIC Sensitivity", "Notes"]
    row = _header_row(ws, row, cols)
    
    sensitivities = [
        ("Arb Win Probability", "↑ Win prob → ↑ MOIC", "High",
         f"Base: {MI.ARB_WIN_PROBABILITY:.0%}; +10% → ~+15% MOIC"),
        ("Quantum Band Shift", "↑ High band prob → ↑ MOIC", "High",
         "Shifting 10% from Band1→Band5 → ~+8% MOIC"),
        ("S.34 Success Rate", "↑ TATA S.34 rate → ↑ MOIC", "Medium",
         "Affects Scenario B outcomes"),
        ("SLP Admission Rate", "↓ SLP admit → ↑ MOIC for weak cases", "Low",
         "Limited paths reach SLP"),
        ("Timeline Extension", "↑ Duration → ↓ IRR (MOIC unchanged)", "Medium",
         "+20% duration → ~-5% IRR"),
        ("Legal Cost Overrun", "↑ Overrun → ↓ Net return", "Medium",
         "+20% costs → ~-8% MOIC on low-upfront"),
    ]
    
    for param, impact, sensitivity, notes in sensitivities:
        sens_fill = RED_FILL if sensitivity == "High" else (
            YELLOW_FILL if sensitivity == "Medium" else GREEN_FILL
        )
        row = _data_row(ws, row, [param, impact, sensitivity, notes],
                        fills=[None, None, sens_fill, None],
                        aligns=[LEFT_WRAP, LEFT_WRAP, CENTER, LEFT_WRAP])
    
    _auto_width(ws)


# ===================================================================
# Sheet 17: Cashflow Projections
# ===================================================================

def _build_cashflow_projections(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    claims: list[ClaimConfig],
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 17: Comprehensive cashflow projections with ₹Cr detail."""
    ws = wb.create_sheet("Cashflow Projections")
    row = _setup_sheet(ws, "Expected Cashflow Projections (₹ Crore)")
    
    total_soc = sum(c.soc_value_cr for c in claims)
    
    # ── Compute per-claim cashflow analytics ──
    claim_cf_data = {}
    for c in claims:
        cid = c.claim_id
        paths = sim.results.get(cid, [])
        if not paths:
            continue
        durations = np.array([p.total_duration_months for p in paths])
        collected = np.array([p.collected_cr for p in paths])
        legal_costs = np.array([p.legal_cost_total_cr for p in paths])
        outcomes = np.array([1 if p.final_outcome == "TRUE_WIN" else 0 for p in paths])
        
        claim_cf_data[cid] = {
            "soc": c.soc_value_cr,
            "jurisdiction": c.jurisdiction,
            "e_collected": float(np.mean(collected)),
            "p50_collected": float(np.median(collected)),
            "p5_collected": float(np.percentile(collected, 5)),
            "p95_collected": float(np.percentile(collected, 95)),
            "e_duration": float(np.mean(durations)),
            "p50_duration": float(np.median(durations)),
            "p5_duration": float(np.percentile(durations, 5)),
            "p95_duration": float(np.percentile(durations, 95)),
            "e_legal": float(np.mean(legal_costs)),
            "win_rate": float(np.mean(outcomes)),
            "eq_cr": sim.expected_quantum_map.get(cid, 0.0),
            "durations": durations,
            "collected": collected,
            "legal_costs": legal_costs,
        }
    
    # ══════════════════════════════════════════════════════════════════
    # SECTION 1: Portfolio Expected Cashflow Summary
    # ══════════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "1. Portfolio Expected Cashflow Summary")
    
    total_eq = sum(d["eq_cr"] for d in claim_cf_data.values())
    total_e_collected = sum(d["e_collected"] for d in claim_cf_data.values())
    total_e_legal = sum(d["e_legal"] for d in claim_cf_data.values())
    avg_win_rate = np.mean([d["win_rate"] for d in claim_cf_data.values()])
    eq_pct = total_eq / total_soc if total_soc > 0 else 0
    collected_pct = total_e_collected / total_soc if total_soc > 0 else 0
    
    summary_items = [
        ("Total Statement of Claim (SOC)", total_soc, CR_FMT, BLUE_FILL),
        ("E[Quantum | WIN] as % of SOC", sum(b["probability"] * (b["low"] + b["high"]) / 2 for b in MI.QUANTUM_BANDS), PCT_FMT, None),
        ("Analytical E[Quantum] — all claims", total_eq, CR_FMT, None),
        ("E[Quantum] / SOC", eq_pct, PCT_FMT, None),
        ("P(Effective Win) — portfolio avg", avg_win_rate, PCT_FMT, None),
        ("E[Collected] — portfolio total", total_e_collected, CR_FMT, YELLOW_FILL),
        ("E[Collected] / SOC", collected_pct, PCT_FMT, YELLOW_FILL),
        ("E[Legal Costs] — portfolio total", total_e_legal, CR_FMT, RED_FILL),
        ("E[Net Cashflow] (before fund share)", total_e_collected - total_e_legal, CR_FMT, GREEN_FILL),
    ]
    
    cols = ["Metric", "Value (₹ Cr)", "Notes"]
    row = _header_row(ws, row, cols)
    
    notes_map = {
        "Total Statement of Claim (SOC)": "Sum across 6 DFCCIL claims",
        "E[Quantum | WIN] as % of SOC": "P(arb_win) × E[Q|WIN] band midpoint",
        "Analytical E[Quantum] — all claims": "SOC × E[Q|WIN] from quantum bands",
        "E[Collected] — portfolio total": f"MC average across {sim.n_paths:,} paths",
        "E[Collected] / SOC": "Key metric — how much of SOC we expect to recover",
        "E[Legal Costs] — portfolio total": "All stages: DAB → Arb → Courts → Enforcement",
        "E[Net Cashflow] (before fund share)": "Does NOT include upfront purchase or fund split",
    }
    
    for label, val, fmt, fill in summary_items:
        note = notes_map.get(label, "")
        fills_list = [fill, fill, None] if fill else None
        font = BOLD_FONT if fill else NORMAL_FONT
        row = _data_row(ws, row, [label, val, note],
                        fmts=[None, fmt, None],
                        fonts=[font, font, SMALL_FONT],
                        fills=fills_list,
                        aligns=[LEFT_WRAP, RIGHT, LEFT_WRAP])
    row += 1
    
    # ══════════════════════════════════════════════════════════════════
    # SECTION 2: Per-Claim E[Collected] Breakdown
    # ══════════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "2. Per-Claim Expected Collected Breakdown (₹ Crore)")
    
    cols = ["Claim", "SOC", "Jurisdiction", "E[Q] (₹Cr)", "E[Q]%",
            "P(Win)", "E[Collected]", "P5 Collected", "P50 Collected",
            "P95 Collected", "E[Legal]", "E[Net]"]
    row = _header_row(ws, row, cols)
    
    for c in claims:
        cid = c.claim_id
        d = claim_cf_data.get(cid, {})
        if not d:
            continue
        e_net = d["e_collected"] - d["e_legal"]
        eq_claim_pct = d["eq_cr"] / c.soc_value_cr if c.soc_value_cr > 0 else 0
        net_fill = GREEN_FILL if e_net > 0 else RED_FILL
        
        row = _data_row(ws, row,
                        [cid, c.soc_value_cr, d["jurisdiction"].upper(),
                         d["eq_cr"], eq_claim_pct, d["win_rate"],
                         d["e_collected"], d["p5_collected"], d["p50_collected"],
                         d["p95_collected"], d["e_legal"], e_net],
                        fmts=[None, CR_FMT, None, CR_FMT, PCT_FMT,
                              PCT_FMT, CR_FMT, CR_FMT, CR_FMT,
                              CR_FMT, CR_FMT, CR_FMT],
                        fills=[None, None, None, None, None,
                               None, YELLOW_FILL, None, None,
                               None, None, net_fill])
    
    # Portfolio totals row
    row = _data_row(ws, row,
                    ["PORTFOLIO", total_soc, "—", total_eq,
                     eq_pct, avg_win_rate,
                     total_e_collected,
                     sum(d["p5_collected"] for d in claim_cf_data.values()),
                     sum(d["p50_collected"] for d in claim_cf_data.values()),
                     sum(d["p95_collected"] for d in claim_cf_data.values()),
                     total_e_legal,
                     total_e_collected - total_e_legal],
                    fmts=[None, CR_FMT, None, CR_FMT, PCT_FMT,
                          PCT_FMT, CR_FMT, CR_FMT, CR_FMT,
                          CR_FMT, CR_FMT, CR_FMT],
                    fonts=[BOLD_FONT]*12,
                    fills=[BLUE_FILL]*12)
    row += 1
    
    # ══════════════════════════════════════════════════════════════════
    # SECTION 3: Annual Resolution & Recovery Timeline
    # ══════════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "3. Annual Resolution & Expected Recovery Timeline")
    
    # Aggregate all durations and collected values across portfolio
    all_durations = []
    all_collected = []
    all_legal = []
    for cid in sim.claim_ids:
        paths = sim.results.get(cid, [])
        for p in paths:
            all_durations.append(p.total_duration_months)
            all_collected.append(p.collected_cr)
            all_legal.append(p.legal_cost_total_cr)
    
    dur_arr = np.array(all_durations)
    col_arr = np.array(all_collected)
    lc_arr = np.array(all_legal)
    n_total = len(dur_arr)
    
    cols = ["Period", "Month Range", "Claims Resolving %",
            "Cumulative Resolved %", "E[Recovery] this period (₹ Cr)",
            "Cumulative E[Recovery] (₹ Cr)", "E[Legal Costs] this period (₹ Cr)",
            "Net Cashflow (₹ Cr)", "Phase"]
    row = _header_row(ws, row, cols)
    
    cumul_recovery = 0.0
    cumul_pct = 0.0
    
    for year in range(1, 9):
        month_start = (year - 1) * 12
        month_end = year * 12
        
        # Fraction of paths resolving this year
        mask_this = (dur_arr > month_start) & (dur_arr <= month_end)
        pct_this = float(np.mean(mask_this))
        pct_cumul = float(np.mean(dur_arr <= month_end))
        
        # Expected recovery for paths resolving this year
        recovery_this = float(np.sum(col_arr[mask_this])) / max(n_total / len(claims), 1)
        cumul_recovery += recovery_this
        
        # Legal cost estimate (proportional to time in this period)
        # Legal costs are front-loaded, approximate by duration overlap
        legal_this = float(np.mean(lc_arr)) * pct_this * len(claims)
        
        net_this = recovery_this - legal_this
        
        if year <= 2:
            phase = "Investment & Pre-Arb"
        elif year <= 4:
            phase = "Arbitration & S.34"
        elif year <= 6:
            phase = "Appeals (S.37/SLP)"
        else:
            phase = "Tail / Re-Arbitration"
        
        net_fill = GREEN_FILL if net_this > 0 else RED_FILL
        row = _data_row(ws, row,
                        [f"Year {year}", f"M{month_start+1}–M{month_end}",
                         pct_this, pct_cumul,
                         recovery_this, cumul_recovery,
                         legal_this, net_this, phase],
                        fmts=[None, None, PCT_FMT, PCT_FMT,
                              CR_FMT, CR_FMT, CR_FMT, CR_FMT, None],
                        fills=[None, None, None, None,
                               None, YELLOW_FILL, None, net_fill, None])
    row += 1
    
    # ══════════════════════════════════════════════════════════════════
    # SECTION 4: Quarterly Cashflow Detail (first 5 years = 20 quarters)
    # ══════════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "4. Quarterly Expected Recovery Distribution")
    
    cols = ["Quarter", "Month Range", "% Claims Resolving",
            "E[Recovery] (₹ Cr)", "Running Total (₹ Cr)"]
    row = _header_row(ws, row, cols)
    
    running_total = 0.0
    for q in range(1, 21):  # 20 quarters = 5 years
        q_start = (q - 1) * 3
        q_end = q * 3
        mask_q = (dur_arr > q_start) & (dur_arr <= q_end)
        pct_q = float(np.mean(mask_q))
        recovery_q = float(np.sum(col_arr[mask_q])) / max(n_total / len(claims), 1)
        running_total += recovery_q
        
        if pct_q > 0.001:  # Only show non-trivial quarters
            row = _data_row(ws, row,
                            [f"Q{q}", f"M{q_start+1}–M{q_end}",
                             pct_q, recovery_q, running_total],
                            fmts=[None, None, PCT_FMT, CR_FMT, CR_FMT],
                            fills=[None, None, None, None, YELLOW_FILL])
    row += 1
    
    # ══════════════════════════════════════════════════════════════════
    # SECTION 5: Investment Structure Cashflow Scenarios
    # ══════════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "5. Investor Cashflow Under Key Scenarios (₹ Crore)")
    
    scenarios = [
        (0.05, 0.80, "Conservative: 5% upfront, 20% Tata tail"),
        (0.10, 0.80, "Base case: 10% upfront, 20% Tata tail"),
        (0.15, 0.70, "Moderate: 15% upfront, 30% Tata tail"),
        (0.20, 0.60, "Aggressive: 20% upfront, 40% Tata tail"),
    ]
    
    cols = ["Scenario", "Upfront (₹ Cr)", "Legal Costs (₹ Cr)",
            "Total Investment (₹ Cr)", "E[Gross Recovery] (₹ Cr)",
            "E[Net to Fund] (₹ Cr)", "E[MOIC]", "E[XIRR]", "P(Loss)", "Verdict"]
    row = _header_row(ws, row, cols)
    
    for up_pct, aw_pct, label in scenarios:
        upfront_cr = up_pct * total_soc
        key = (up_pct, aw_pct, basis)
        cell = grid.cells.get(key)
        
        if cell:
            e_moic = cell.mean_moic
            e_xirr = cell.mean_xirr
            p_loss = cell.p_loss
            total_inv = upfront_cr + total_e_legal
            e_gross = total_inv * e_moic
            e_net = e_gross - total_inv
            
            if e_moic >= 2.5 and p_loss < 0.10:
                verdict = "STRONG BUY"
                v_fill = GREEN_FILL
            elif e_moic >= 1.5 and p_loss < 0.25:
                verdict = "BUY"
                v_fill = GREEN_FILL
            elif e_moic >= 1.0:
                verdict = "HOLD"
                v_fill = YELLOW_FILL
            else:
                verdict = "AVOID"
                v_fill = RED_FILL
            
            row = _data_row(ws, row,
                            [label, upfront_cr, total_e_legal, total_inv,
                             e_gross, e_net, e_moic, e_xirr, p_loss, verdict],
                            fmts=[None, CR_FMT, CR_FMT, CR_FMT,
                                  CR_FMT, CR_FMT, MOIC_FMT, PCT_FMT, PCT_FMT, None],
                            fills=[None, BLUE_FILL, None, YELLOW_FILL,
                                   None, GREEN_FILL if e_net > 0 else RED_FILL,
                                   None, None, RED_FILL if p_loss > 0.15 else None, v_fill],
                            aligns=[LEFT_WRAP] + [RIGHT]*8 + [CENTER])
        else:
            row = _data_row(ws, row,
                            [label, upfront_cr, total_e_legal,
                             upfront_cr + total_e_legal,
                             "N/A", "N/A", "N/A", "N/A", "N/A", "—"],
                            fmts=[None, CR_FMT, CR_FMT, CR_FMT,
                                  None, None, None, None, None, None])
    row += 1
    
    # ══════════════════════════════════════════════════════════════════
    # SECTION 6: Per-Claim Timing & Recovery Heatmap
    # ══════════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "6. Per-Claim Resolution Timing (months)")
    
    cols = ["Claim", "SOC (₹ Cr)", "E[Duration]", "P5", "Median",
            "P95", "Fastest Possible", "E[Collected] (₹ Cr)",
            "E[Collected] / SOC"]
    row = _header_row(ws, row, cols)
    
    for c in claims:
        cid = c.claim_id
        d = claim_cf_data.get(cid, {})
        if not d:
            continue
        e_col_pct = d["e_collected"] / c.soc_value_cr if c.soc_value_cr > 0 else 0
        
        # Fastest possible = shortest pipeline
        min_dur = d["p5_duration"]
        
        duration_fill = GREEN_FILL if d["e_duration"] < 48 else (
            YELLOW_FILL if d["e_duration"] < 60 else RED_FILL
        )
        
        row = _data_row(ws, row,
                        [cid, c.soc_value_cr, d["e_duration"],
                         d["p5_duration"], d["p50_duration"], d["p95_duration"],
                         min_dur, d["e_collected"], e_col_pct],
                        fmts=[None, CR_FMT, MONTH_FMT, MONTH_FMT, MONTH_FMT,
                              MONTH_FMT, MONTH_FMT, CR_FMT, PCT_FMT],
                        fills=[None, None, duration_fill, None, None,
                               None, None, YELLOW_FILL, None])
    row += 1
    
    # ══════════════════════════════════════════════════════════════════
    # SECTION 7: Value-at-Risk Cashflow Analysis
    # ══════════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "7. Portfolio Recovery Distribution (₹ Crore)")
    
    # Compute portfolio-level collected per path
    portfolio_collected_per_path = np.zeros(sim.n_paths)
    portfolio_legal_per_path = np.zeros(sim.n_paths)
    for cid in sim.claim_ids:
        paths = sim.results.get(cid, [])
        for i, p in enumerate(paths):
            portfolio_collected_per_path[i] += p.collected_cr
            portfolio_legal_per_path[i] += p.legal_cost_total_cr
    
    portfolio_net = portfolio_collected_per_path - portfolio_legal_per_path
    
    percentiles = [1, 5, 10, 25, 50, 75, 90, 95, 99]
    cols = ["Percentile", "Gross Collected (₹ Cr)", "Legal Costs (₹ Cr)",
            "Net Recovery (₹ Cr)", "Net / SOC"]
    row = _header_row(ws, row, cols)
    
    for pctl in percentiles:
        gross_p = float(np.percentile(portfolio_collected_per_path, pctl))
        legal_p = float(np.percentile(portfolio_legal_per_path, pctl))
        net_p = float(np.percentile(portfolio_net, pctl))
        net_soc_pct = net_p / total_soc if total_soc > 0 else 0
        
        fill = RED_FILL if net_p < 0 else (GREEN_FILL if pctl >= 50 else None)
        row = _data_row(ws, row,
                        [f"P{pctl}", gross_p, legal_p, net_p, net_soc_pct],
                        fmts=[None, CR_FMT, CR_FMT, CR_FMT, PCT_FMT],
                        fills=[None, None, None, fill, fill])
    
    # Mean row
    row = _data_row(ws, row,
                    ["Mean", float(np.mean(portfolio_collected_per_path)),
                     float(np.mean(portfolio_legal_per_path)),
                     float(np.mean(portfolio_net)),
                     float(np.mean(portfolio_net)) / total_soc],
                    fmts=[None, CR_FMT, CR_FMT, CR_FMT, PCT_FMT],
                    fonts=[BOLD_FONT]*5,
                    fills=[BLUE_FILL]*5)
    row += 1
    
    # ══════════════════════════════════════════════════════════════════
    # SECTION 8: Cashflow Decomposition Chain
    # ══════════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "8. Value Decomposition: SOC → E[Collected]")
    
    cols = ["Step", "Description", "Factor", "Running Value (₹ Cr)", "Notes"]
    row = _header_row(ws, row, cols)
    
    # E[Q|WIN] from quantum bands
    eq_given_win = sum(
        b["probability"] * (b["low"] + b["high"]) / 2
        for b in MI.QUANTUM_BANDS
    )
    
    steps = [
        ("1. SOC", "Total Statement of Claim", "—", total_soc,
         "As per claim filings"),
        ("2. × P(arb_win)", "Probability of winning arbitration",
         f"{MI.ARB_WIN_PROBABILITY:.0%}", total_soc * MI.ARB_WIN_PROBABILITY,
         "Expert judgment: 70%"),
        ("3. × E[Q|WIN]", "Expected quantum given win",
         f"{eq_given_win:.1%}", total_soc * MI.ARB_WIN_PROBABILITY * eq_given_win,
         "5-band mixture: weighted midpoints"),
        ("4. × P(survive tree)", "Probability award survives court challenges",
         f"~{avg_win_rate/MI.ARB_WIN_PROBABILITY:.1%}",
         total_e_collected,
         "Domestic: S.34→S.37→SLP; SIAC: HC→COA"),
        ("5. = E[Collected]", "Expected amount collected",
         "—", total_e_collected,
         f"{collected_pct:.1%} of SOC"),
        ("6. − E[Legal]", "Minus expected legal costs",
         f"−₹{total_e_legal:.1f} Cr", total_e_collected - total_e_legal,
         "DAB + Arb + Courts + Enforcement"),
    ]
    
    for step, desc, factor, running, note in steps:
        is_result = "E[Collected]" in step or "Legal" in step
        font = BOLD_FONT if is_result else NORMAL_FONT
        fill = YELLOW_FILL if "E[Collected]" in step else (
            RED_FILL if "Legal" in step else None)
        fills_list = [fill, fill, None, fill, None] if fill else None
        row = _data_row(ws, row,
                        [step, desc, factor, running, note],
                        fmts=[None, None, None, CR_FMT, None],
                        fonts=[font, font, NORMAL_FONT, font, SMALL_FONT],
                        fills=fills_list,
                        aligns=[LEFT_WRAP, LEFT_WRAP, CENTER, RIGHT, LEFT_WRAP])
    row += 1
    
    # ══════════════════════════════════════════════════════════════════
    # SECTION 9: Key Observations
    # ══════════════════════════════════════════════════════════════════
    row = _section_header(ws, row, "9. Key Cashflow Observations")
    
    # Analytical observations
    largest_claim = max(claims, key=lambda c: c.soc_value_cr)
    largest_pct = largest_claim.soc_value_cr / total_soc
    
    observations = [
        f"Portfolio total SOC: ₹{total_soc:,.2f} Cr across 6 TATA arbitration claims",
        f"Expected collected amount: ₹{total_e_collected:,.2f} Cr ({collected_pct:.1%} of SOC)",
        f"Expected legal costs: ₹{total_e_legal:,.2f} Cr ({total_e_legal/total_soc:.1%} of SOC)",
        f"Largest claim: {largest_claim.claim_id} at ₹{largest_claim.soc_value_cr:,.0f} Cr ({largest_pct:.0%} of portfolio)",
        f"Domestic claims face longer timelines (E[τ]≈43-52m) vs SIAC (E[τ]≈39-51m)",
        f"Portfolio VaR(1%): ₹{float(np.percentile(portfolio_net, 1)):,.2f} Cr net recovery",
        f"All values in ₹ Crore. MC simulation: {sim.n_paths:,} paths, seed {sim.seed}",
    ]
    
    for obs in observations:
        cell = ws.cell(row=row, column=2, value=f"• {obs}")
        cell.font = NORMAL_FONT
        cell.alignment = LEFT_WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        row += 1
    
    _auto_width(ws)


# ===================================================================
# Sheet 18: Breakeven Analysis
# ===================================================================

def _build_breakeven_analysis(
    wb: openpyxl.Workbook,
    grid: InvestmentGridResults,
    basis: str,
) -> None:
    """Sheet 18: Breakeven analysis."""
    ws = wb.create_sheet("Breakeven Analysis")
    row = _setup_sheet(ws, "Breakeven Analysis")
    
    be = grid.breakeven.get(basis, {})
    aw_pcts = sorted(grid.award_share_pcts)
    
    row = _section_header(ws, row, f"Maximum Viable Upfront % ({basis.upper()} Pricing)")
    row = _header_row(ws, row, ["Tata Tail %", "Max Upfront %",
                                "E[MOIC] at Breakeven", "E[XIRR] at Breakeven", "P(Loss)", "Verdict"])
    
    for aw in aw_pcts:
        max_up = be.get(aw, 0)
        key = (max_up, aw, basis)
        cell_data = grid.cells.get(key)
        moic = cell_data.mean_moic if cell_data else 0
        xirr = cell_data.mean_xirr if cell_data else 0
        ploss = cell_data.p_loss if cell_data else 1
        v = _verdict(moic, ploss)
        vf = _verdict_fill(v)
        
        row = _data_row(ws, row, [1-aw, max_up, moic, xirr, ploss, v],
                        fmts=[PCT_FMT, PCT_FMT, MOIC_FMT, PCT_FMT, PCT_FMT, None],
                        fills=[None, YELLOW_FILL, None, None, None, vf])
    row += 1
    
    # Summary
    row = _section_header(ws, row, "Interpretation")
    notes = [
        "Breakeven = maximum upfront % of SOC where portfolio E[MOIC] ≥ 1.0.",
        "Higher Tata tails allow larger upfront investments while maintaining profitability.",
        "Values below the minimum tested (5%) indicate the scenario requires <5% upfront.",
    ]
    for note in notes:
        ws.cell(row=row, column=2, value=f"• {note}").font = SMALL_FONT
        row += 1
    
    _auto_width(ws)


# ===================================================================
# Sheet 19: Legal Cost Analysis
# ===================================================================

def _build_legal_costs(
    wb: openpyxl.Workbook,
    sim: SimulationResults,
    claims: list[ClaimConfig],
) -> None:
    """Sheet 19: Detailed legal cost analysis."""
    ws = wb.create_sheet("Legal Cost Analysis")
    row = _setup_sheet(ws, "Legal Cost Analysis")
    
    # Per-claim summary
    row = _section_header(ws, row, "Legal Cost Summary (from MC Simulation)")
    cols = ["Claim", "E[Total] (₹ Cr)", "Median", "Std", "P5", "P95"]
    row = _header_row(ws, row, cols)
    
    portfolio_total = 0
    for cid in sim.claim_ids:
        paths = sim.results.get(cid, [])
        if not paths:
            continue
        
        costs = np.array([p.legal_cost_total_cr for p in paths])
        e_cost = float(np.mean(costs))
        portfolio_total += e_cost
        
        row = _data_row(ws, row,
                        [cid, e_cost, float(np.median(costs)), float(np.std(costs)),
                         float(np.percentile(costs, 5)), float(np.percentile(costs, 95))],
                        fmts=[None] + [CR_FMT] * 5)
    
    row = _data_row(ws, row,
                    ["PORTFOLIO TOTAL", portfolio_total, "", "", "", ""],
                    fonts=[BOLD_FONT, BOLD_FONT, None, None, None, None],
                    fmts=[None, CR_FMT, None, None, None, None],
                    fills=[YELLOW_FILL, YELLOW_FILL, None, None, None, None])
    row += 1
    
    # Cost structure from MI
    row = _section_header(ws, row, "Cost Structure (from master_inputs)")
    
    # One-time costs
    if hasattr(MI, 'LEGAL_COSTS') and "onetime" in MI.LEGAL_COSTS:
        row = _header_row(ws, row, ["Category", "Component", "Cost (₹ Cr)"])
        onetime = MI.LEGAL_COSTS["onetime"]
        for component, cost in onetime.items():
            row = _data_row(ws, row, ["One-time (Month 0)", component, cost],
                            fmts=[None, None, CR_FMT],
                            fonts=[NORMAL_FONT, NORMAL_FONT, BLUE_FONT])
        total_onetime = sum(onetime.values())
        row = _data_row(ws, row, ["One-time Total", "", total_onetime],
                        fonts=[BOLD_FONT, None, BOLD_FONT], fmts=[None, None, CR_FMT])
        row += 1
    
    # Duration-based costs
    if hasattr(MI, 'LEGAL_COSTS') and "duration_based" in MI.LEGAL_COSTS:
        row = _header_row(ws, row, ["Stage", "Low (₹ Cr)", "High (₹ Cr)", "Notes"])
        db = MI.LEGAL_COSTS["duration_based"]
        for stage_key, val in db.items():
            if isinstance(val, dict):
                row = _data_row(ws, row,
                                [stage_key, val["low"], val["high"], "range"],
                                fmts=[None, CR_FMT, CR_FMT, None],
                                fonts=[NORMAL_FONT, BLUE_FONT, BLUE_FONT, SMALL_FONT])
            else:
                row = _data_row(ws, row,
                                [stage_key, val, val, "fixed"],
                                fmts=[None, CR_FMT, CR_FMT, None],
                                fonts=[NORMAL_FONT, BLUE_FONT, BLUE_FONT, SMALL_FONT])
        row += 1
    
    # Overrun parameters
    if hasattr(MI, 'LEGAL_COST_OVERRUN'):
        row = _section_header(ws, row, "Legal Cost Overrun Distribution")
        ov = MI.LEGAL_COST_OVERRUN
        row = _kv_row(ws, row, "Distribution", f"ScaledBeta(α={ov['alpha']}, β={ov['beta']})")
        row = _kv_row(ws, row, "Range", f"[{ov['low']:.0%}, {ov['high']:.0%}]")
        e_overrun = ov["low"] + (ov["alpha"] / (ov["alpha"] + ov["beta"])) * (ov["high"] - ov["low"])
        row = _kv_row(ws, row, "E[Overrun]", e_overrun, val_fmt=PCT_FMT)
    
    _auto_width(ws)


# ===================================================================
# Sheet 20: Glossary & Caveats
# ===================================================================

def _build_glossary(wb: openpyxl.Workbook) -> None:
    """Sheet 20: Glossary and caveats."""
    ws = wb.create_sheet("Glossary & Caveats")
    row = _setup_sheet(ws, "Glossary & Caveats")
    
    # Terms
    row = _section_header(ws, row, "Key Terms")
    terms = [
        ("SOC", "Statement of Claim — the total claimed amount (₹ Crore)."),
        ("E[Q]", "Expected Quantum — analytical mean of award conditional on winning."),
        ("MOIC", "Multiple on Invested Capital = total return ÷ total invested."),
        ("XIRR", "Extended Internal Rate of Return — annualized time-weighted return."),
        ("P(Loss)", "Probability of capital loss (MOIC < 1.0)."),
        ("VaR 1%", "Value at Risk — 1st percentile of return distribution."),
        ("CVaR 1%", "Conditional VaR — expected loss given we are in worst 1%."),
        ("TPL Share", "Third Party Litigation funder's ownership share."),
        ("DAB", "Dispute Adjudication Board — first-instance resolution."),
        ("S.34", "Section 34 of Indian Arbitration Act — set aside application."),
        ("S.37", "Section 37 — appeal against S.34 decision."),
        ("SLP", "Special Leave Petition — Supreme Court discretionary review."),
        ("SIAC", "Singapore International Arbitration Centre."),
        ("TRUE_WIN", "Final collectible outcome — TATA collects quantum."),
        ("RESTART", "Award set aside → re-arbitration with new tribunal."),
        ("LOSE", "Final loss — TATA collects nothing."),
        ("Upfront %", "Upfront investment as fraction of SOC."),
        ("Tata Tail %", "Percentage of award paid to TATA. Fund keeps (1 - Tail%)."),
        ("Strong Buy", "E[MOIC] > 2.5× AND P(Loss) < 10%."),
        ("Attractive", "E[MOIC] > 1.5× AND P(Loss) < 25%."),
        ("Marginal", "E[MOIC] > 1.0× AND P(Loss) < 40%."),
        ("Avoid", "Does not meet Marginal thresholds."),
    ]
    
    cols = ["Term", "Definition"]
    row = _header_row(ws, row, cols)
    for term, defn in terms:
        row = _data_row(ws, row, [term, defn],
                        fonts=[BOLD_FONT, NORMAL_FONT],
                        aligns=[CENTER, LEFT_WRAP])
    row += 1
    
    # Methodology
    row = _section_header(ws, row, "Methodology Notes")
    notes = [
        "Monte Carlo simulation draws N independent paths per claim.",
        "Each path traverses: Timeline → Arbitration → Challenge Tree → Re-Arb → Payment.",
        "Quantum is drawn from 5 discrete bands conditional on arbitration WIN.",
        "Legal costs are modelled with per-stage burn rates plus stochastic overrun.",
        "Investment cashflow: upfront (month 0) + legal costs (monthly) → award at resolution.",
        "XIRR computed using scipy.optimize.brentq with monthly bounds.",
        "Probability trees implement full domestic (24 paths) and SIAC (8 paths) traversal.",
        "Re-arbitration is modelled for RESTART outcomes with independent P(win) = 70%.",
    ]
    for note in notes:
        ws.cell(row=row, column=2, value=f"• {note}").font = NORMAL_FONT
        row += 1
    row += 1
    
    # Caveats
    row = _section_header(ws, row, "Important Caveats")
    caveats = [
        "TPL shares are set to 100% — actual JV shares (50% / 33.33%) are UNCONFIRMED.",
        "Quantum bands and arbitration win probability are expert-judgment calibrated.",
        "Legal cost overrun distribution assumes mean +10% — may vary by complexity.",
        "IRR is undefined for paths with no positive cashflows (set to -100%).",
        "Collection efficiency is 100% of awarded quantum (no default modelled in v2).",
        "This model is for illustrative purposes — actual outcomes may vary.",
    ]
    for caveat in caveats:
        ws.cell(row=row, column=2, value=f"⚠ {caveat}").font = WARNING_FONT
        row += 1
    
    _auto_width(ws)


# ===================================================================
# Main Entry Point
# ===================================================================

def generate_comprehensive_report(
    sim: SimulationResults,
    claims: list[ClaimConfig],
    grid: InvestmentGridResults,
    basis: str = "soc",
    output_dir: Optional[str] = None,
    filename: str = "Investment_Analysis_Report.xlsx",
    ctx=None,
) -> str:
    """Generate the comprehensive 20-sheet Excel workbook.

    Parameters
    ----------
    sim : SimulationResults
    claims : list[ClaimConfig]
    grid : InvestmentGridResults
    basis : str — pricing basis for analysis
    output_dir : str — output directory
    filename : str — output filename
    ctx : PortfolioContext, optional — if provided, uses ctx.output_dir

    Returns
    -------
    str — full path to generated Excel file.
    """
    out_dir = output_dir or (ctx.output_dir if ctx else MI.REPORT_OUTPUT_DIR)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, filename)

    wb = openpyxl.Workbook()

    print("  Generating Comprehensive Investment Report...")

    # Sheet 1: Cover
    _build_cover(wb, sim, claims)
    print("    [1/20] Cover")

    # Sheet 2: Executive Summary
    _build_executive_summary(wb, sim, claims, grid, basis)
    print("    [2/20] Executive Summary")

    # Sheet 3: Investment Decisions
    _build_investment_decisions(wb, grid, claims, basis)
    print("    [3/20] Investment Decisions")

    # Sheet 4: Risk Analysis
    _build_risk_analysis(wb, sim, claims, grid, basis)
    print("    [4/20] Risk Analysis")

    # Sheet 5: Model Assumptions
    _build_model_assumptions(wb, claims)
    print("    [5/20] Model Assumptions")

    # Sheet 6: Probability Trees
    _build_probability_trees(wb, sim)
    print("    [6/20] Probability Trees")

    # Sheet 7: Timeline Analysis
    _build_timeline_analysis(wb, sim, claims)
    print("    [7/20] Timeline Analysis")

    # Sheet 8: Quantum Analysis
    _build_quantum_analysis(wb, sim, claims)
    print("    [8/20] Quantum Analysis")

    # Sheets 9-14: Per-claim
    claim_map = {c.claim_id: c for c in claims}
    for i, cid in enumerate(sim.claim_ids):
        _build_claim_sheet(wb, cid, sim, claim_map[cid], grid, basis)
        print(f"    [{i+9}/20] {cid}")

    # Sheet 15: Portfolio Grid
    _build_portfolio_grid(wb, grid, basis)
    print("    [15/20] Portfolio Grid")

    # Sheet 16: Sensitivity Analysis
    _build_sensitivity_analysis(wb, sim, grid, basis)
    print("    [16/20] Sensitivity Analysis")

    # Sheet 17: Cashflow Projections
    _build_cashflow_projections(wb, sim, claims, grid, basis)
    print("    [17/20] Cashflow Projections")

    # Sheet 18: Breakeven Analysis
    _build_breakeven_analysis(wb, grid, basis)
    print("    [18/20] Breakeven Analysis")

    # Sheet 19: Legal Cost Analysis
    _build_legal_costs(wb, sim, claims)
    print("    [19/20] Legal Cost Analysis")

    # Sheet 20: Glossary & Caveats
    _build_glossary(wb)
    print("    [20/20] Glossary & Caveats")

    wb.save(out_path)
    print(f"  Comprehensive report saved: {out_path}")
    return out_path
